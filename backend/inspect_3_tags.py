import openpyxl
from pathlib import Path

source_path = Path(r"c:\vivek\PLANT\Turbine & Instrument Equipment life histroy 20-06-2026 - new.xlsx")
wb = openpyxl.load_workbook(source_path, data_only=True)

search_terms = ["FCV", "201", "502", "B.COMP-CT-06", "COMP-CT-06"]

for tab in wb.sheetnames:
    ws = wb[tab]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_str = " ".join(str(c) for c in row if c is not None)
        for term in search_terms:
            if term.lower() in row_str.lower():
                print(f"Match '{term}' in sheet '{tab}', row {row_idx}: {row_str[:140]}")

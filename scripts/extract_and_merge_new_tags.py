#!/usr/bin/env python3
"""
Extract 7 target equipment cards from 'Turbine & Instrument Equipment life histroy 20-06-2026 - new.xlsx'
and merge/update them in 'sugar-house-equipment-life-history-filtered-updated-updated.xlsx'.
"""

import sys
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Add scripts directory to path
SCRIPTS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_DIR))

from extract_shn_life_history_table import (
    find_equipment_subcards,
    extract_life_fields,
    extract_specification_rows,
    extract_maintenance_schedule,
    extract_maintenance_history,
    should_skip_sheet,
    make_output_sheet_name,
    insert_spec_meta_columns,
    insert_tag_column,
)
from shn_tag_match import matches_any_filter, norm_tag, compact_tag, tags_match

ROOT = SCRIPTS_DIR.parent.parent
SOURCE_NEW = ROOT / "Turbine & Instrument Equipment life histroy 20-06-2026 - new.xlsx"
MASTER_FILE = (
    SCRIPTS_DIR.parent
    / "backend"
    / "backlog-data"
    / "mill data"
    / "sugar-house-equipment-life-history-filtered-updated-updated.xlsx"
)

TARGET_TAGS = {
    "ZIL/SUG./001/SY_B_MOL_CV2",
    "ZIL/SUG./001 (40FCV201)",
    "ZIL/SUG./01/16.FT.075.B01",
    "ZIL/SUG./01/DIST.-TWFM-2",
    "ZIL/SUG./01/ETP-DECANTER-FM",
    "ZIL/SUG./01/Vacuum-FM-01",
    "ZIL/SUG./B.COMP-CT-05",
    "ZIL/SUG./001  (FCV 201)",
    "ZIL/SUG./017  ( PI_502)",
    "ZIL/SUG./B.COMP-CT-06",
}


def main():
    print(f"Source: {SOURCE_NEW}")
    print(f"Master Excel: {MASTER_FILE}")

    if not SOURCE_NEW.exists():
        raise FileNotFoundError(f"Source file not found: {SOURCE_NEW}")
    if not MASTER_FILE.exists():
        raise FileNotFoundError(f"Master file not found: {MASTER_FILE}")

    src_wb = openpyxl.load_workbook(SOURCE_NEW, data_only=True)

    extracted_cards = []

    for tab_name in src_wb.sheetnames:
        ws = src_wb[tab_name]
        if should_skip_sheet(tab_name, ws):
            continue
        blocks = find_equipment_subcards(ws)
        for card_index, block in enumerate(blocks, start=1):
            fields = extract_life_fields(ws, block)
            equip_tag = fields["EQUIPMENT TAG NAME/APPLICATION"]

            if matches_any_filter(equip_tag, TARGET_TAGS):
                output_name = make_output_sheet_name(tab_name, fields, card_index)
                spec_rows = extract_specification_rows(ws, block, "", output_name)
                specs_with_meta = insert_spec_meta_columns(
                    spec_rows, equip_tag, fields["NAME OF EQUIPMENT"]
                )
                sched_rows = insert_tag_column(
                    extract_maintenance_schedule(ws, block, "", output_name),
                    equip_tag,
                )
                hist_rows = insert_tag_column(
                    extract_maintenance_history(ws, block, "", output_name),
                    equip_tag,
                )

                extracted_cards.append({
                    "tab": tab_name,
                    "output_name": output_name,
                    "tag": equip_tag,
                    "fields": fields,
                    "specs": specs_with_meta,
                    "schedule": sched_rows,
                    "history": hist_rows,
                })

    print(f"Extracted {len(extracted_cards)} cards from new workbook:")
    for card in extracted_cards:
        print(f"  - Tag: {card['tag']} | Output Name: {card['output_name']}")

    # Load master workbook
    master_wb = openpyxl.load_workbook(MASTER_FILE)

    map_ws = master_wb["Sheet Map"]
    life_ws = master_wb["EQUIPMENT LIFE HISTORY CARD"]
    spec_ws = master_wb["EQUIPMENT SPECIFICATION"]
    sched_ws = master_wb["MAINTENANCE SCHEDULE"]
    hist_ws = master_wb["EQUIPMENT MAINTENANCE HISTORY"]

    # Build existing tag -> sheet_id map from Sheet Map
    tag_to_id = {}
    for row in map_ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 3 and row[1] and row[2]:
            tag_to_id[norm_tag(row[2])] = str(row[1])

    # Remove existing rows for these tags from master sheets to avoid duplication
    def remove_matching_tag_rows(ws, tag_col_idx):
        rows_to_keep = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                headers = row
                rows_to_keep.append(row)
                continue
            if not row or len(row) <= tag_col_idx:
                rows_to_keep.append(row)
                continue
            row_tag = str(row[tag_col_idx] or "").strip()
            if matches_any_filter(row_tag, TARGET_TAGS):
                continue  # skip/remove
            rows_to_keep.append(row)
        return rows_to_keep

    # Remove existing matching rows
    map_rows = remove_matching_tag_rows(map_ws, 2)
    life_rows = remove_matching_tag_rows(life_ws, 2)
    spec_rows = remove_matching_tag_rows(spec_ws, 2)
    sched_rows = remove_matching_tag_rows(sched_ws, 2)
    hist_rows = remove_matching_tag_rows(hist_ws, 2)

    # Append new cards
    for card in extracted_cards:
        norm_t = norm_tag(card["tag"])
        sheet_id = tag_to_id.get(norm_t) or uuid.uuid4().hex[:12]

        map_rows.append([card["output_name"], sheet_id, card["tag"]])
        life_rows.append([
            sheet_id,
            card["output_name"],
            card["tag"],
            card["fields"]["NAME OF EQUIPMENT"],
            card["fields"]["LOCATION"],
            card["fields"]["DATE OF COMMISSIONING"],
        ])

        for spec in card["specs"]:
            spec_rows.append([
                sheet_id,
                card["output_name"],
                card["tag"],
                card["fields"]["NAME OF EQUIPMENT"],
                spec[4],
                spec[5],
                spec[6],
                spec[7],
            ])

        for sched in card["schedule"]:
            # sched has tag inserted at index 2
            sched_copy = list(sched)
            sched_copy[0] = sheet_id
            sched_copy[1] = card["output_name"]
            sched_rows.append(sched_copy)

        for hist in card["history"]:
            hist_copy = list(hist)
            hist_copy[0] = sheet_id
            hist_copy[1] = card["output_name"]
            hist_rows.append(hist_copy)

    # Helper to rewrite worksheet
    def rewrite_sheet(ws, rows_data):
        ws.delete_rows(1, ws.max_row + 1)
        for r in rows_data:
            ws.append(list(r))

    rewrite_sheet(map_ws, map_rows)
    rewrite_sheet(life_ws, life_rows)
    rewrite_sheet(spec_ws, spec_rows)
    rewrite_sheet(sched_ws, sched_rows)
    rewrite_sheet(hist_ws, hist_rows)

    try:
        master_wb.save(MASTER_FILE)
        saved_file = MASTER_FILE
    except PermissionError:
        fallback = MASTER_FILE.with_name(f"{MASTER_FILE.stem}-updated{MASTER_FILE.suffix}")
        master_wb.save(fallback)
        saved_file = fallback
    print(f"Successfully updated master Excel workbook: {saved_file}")


if __name__ == "__main__":
    main()

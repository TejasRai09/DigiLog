#!/usr/bin/env python3
"""
Build normalized life-history workbook from BLR Machine history card.xlsx.

Output: blr-machine-life-history.xlsx (or -updated if file is open)
  - Sheet "Sheet Map": sheet name | id (random)
  - Sheet "EQUIPMENT LIFE HISTORY CARD": sheet id + life-history fields
  - Sheet "EQUIPMENT SPECIFICATION": sheet id + spec parameter rows
  - Sheet "MAINTENANCE SCHEDULE": sheet id + OEM schedule rows
  - Sheet "EQUIPMENT MAINTENANCE HISTORY": sheet id + maintenance history rows
"""

from __future__ import annotations

import re
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from equipment_history_extract_lib import (
    cell_text,
    clean_param_label,
    extract_history_rows_flexible,
    extract_schedule_rows,
    extract_specification_rows as extract_spec_rows_lib,
    norm,
    parse_schedule_layout,
    row_cells,
)

ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE = ROOT / "BLR Machine history card.xlsx"
OUTPUT = ROOT / "blr-machine-life-history.xlsx"

SKIP_SHEETS = {"index"}
SECTION_LIFE = "EQUIPMENT LIFE HISTORY CARD"
SECTION_SPEC = "EQUIPMENT SPECIFICATION"
SECTION_SCHEDULE = "MAINTENANCE SCHEDULE"
SECTION_HISTORY = "EQUIPMENT MAINTENANCE HISTORY"

SPEC_COLUMNS = [
    "sheet id",
    "sheet name",
    "section",
    "sub_section",
    "Parameter label",
    "Parameter value",
]

LIFE_COLUMNS = [
    "sheet id",
    "sheet name",
    "NAME OF EQUIPMENT",
    "LOCATION",
    "EQUIPMENT TAG NAME/APPLICATION",
    "DATE OF COMMISSIONING",
]

HISTORY_COLUMNS = [
    "sheet id",
    "sheet name",
    "Season / OFF Season",
    "Year",
    "Date of Start",
    "Date of Finish",
    "Outage/ Observation",
    "Action Taken",
    "Repair Cost (Rs.)",
    "Services (Internal / External)",
    "Responsibility ( Engineer/ Supervision)",
    "Remarks",
]

SCHEDULE_COLUMNS = [
    "sheet id",
    "sheet name",
    "Sr.No.",
    "Name of Equipment",
    "Maintenance / Inspection Activities",
    "Daily",
    "Weekly",
    "Monthly",
    "Yearly",
    "2 - Years",
    "3 - Years",
    "4 - Years",
    "Remarks",
]

BLR_SCHEDULE_INTERVALS = (
    "Daily",
    "Weekly",
    "Monthly",
    "Yearly",
    "2 - Years",
    "3 - Years",
    "4 - Years",
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def should_skip_sheet(name: str) -> bool:
    return norm(name) in {norm(s) for s in SKIP_SHEETS}


def find_section_rows(ws) -> dict[str, int]:
    found: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        b = norm(ws.cell(r, 2).value)
        if not b:
            b = norm(" ".join(cell_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)))
        for label, key in (
            (SECTION_LIFE, "life"),
            (SECTION_SPEC, "spec"),
            (SECTION_SCHEDULE, "schedule"),
            (SECTION_HISTORY, "history"),
        ):
            if key not in found and label in b:
                found[key] = r
    return found


def extract_history_rows_standard(ws, history_row: int, sheet_id: str, sheet_name: str) -> list[list[str]]:
    return extract_history_rows_flexible(
        ws,
        history_row,
        sheet_id,
        sheet_name,
        stop_markers=(SECTION_LIFE, SECTION_SPEC),
    )


def extract_history_rows_gauge_glass(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        joined = norm(" ".join(cell_text(ws.cell(r, c).value) for c in range(1, 8)))
        if "SR NO" in joined and "DATE" in joined and "MAINTENANCE" in joined:
            header_row = r
            break
    if not header_row:
        return []

    rows: list[list[str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        year = cell_text(ws.cell(r, 2).value)
        outage = cell_text(ws.cell(r, 3).value)
        nature = cell_text(ws.cell(r, 4).value)
        action = cell_text(ws.cell(r, 5).value)
        break_time = cell_text(ws.cell(r, 6).value)
        spare = cell_text(ws.cell(r, 7).value)
        if not any([year, outage, nature, action, break_time, spare]):
            continue
        observation = outage
        if nature and nature not in observation:
            observation = f"{observation} | {nature}" if observation else nature
        remarks = spare
        if break_time:
            remarks = f"Break Down Time: {break_time}" + (f" | {spare}" if spare else "")
        rows.append([
            sheet_id,
            sheet_name,
            "",
            year,
            "",
            "",
            observation,
            action,
            "",
            "",
            "",
            remarks,
        ])
    return rows


def extract_maintenance_schedule(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "schedule" not in sections:
        return []

    schedule_row = sections["schedule"]
    schedule_end = sections.get("history") or (ws.max_row + 1)
    layout = parse_schedule_layout(ws, schedule_row, schedule_end)
    if not layout:
        return []

    return extract_schedule_rows(
        ws,
        layout,
        schedule_end,
        sheet_id,
        sheet_name,
        BLR_SCHEDULE_INTERVALS,
        (SECTION_HISTORY, SECTION_SCHEDULE),
    )


def extract_maintenance_history(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "history" in sections:
        return extract_history_rows_standard(ws, sections["history"], sheet_id, sheet_name)
    if norm(ws.title) == norm("gauge glass"):
        return extract_history_rows_gauge_glass(ws, sheet_id, sheet_name)
    return []


def row_value_after_label(ws, row_idx: int, label_col: int = 2) -> str:
    """First non-empty cell to the right of the label column on the same row."""
    for c in range(label_col + 1, ws.max_column + 1):
        val = cell_text(ws.cell(row_idx, c).value)
        if val:
            return val
    return ""


def extract_specification_rows(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "spec" not in sections:
        if norm(ws.title) == norm("gauge glass"):
            make = cell_text(ws.cell(7, 2).value)
            val = cell_text(ws.cell(7, 3).value)
            if make and val:
                return [[sheet_id, sheet_name, "", "", clean_param_label(make), val]]
        return []

    spec_row = sections["spec"]
    spec_end = sections.get("schedule") or sections.get("history") or (ws.max_row + 1)
    return extract_spec_rows_lib(
        ws,
        spec_row,
        spec_end,
        sheet_id,
        sheet_name,
        (SECTION_SCHEDULE, SECTION_HISTORY, SECTION_LIFE),
    )


def load_existing_ids() -> dict[str, str]:
    candidates = (
        OUTPUT,
        OUTPUT.with_name(f"{OUTPUT.stem}-updated{OUTPUT.suffix}"),
    )
    for path in candidates:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Sheet Map" not in wb.sheetnames:
            continue
        ws = wb["Sheet Map"]
        ids = {
            cell_text(ws.cell(r, 1).value): cell_text(ws.cell(r, 2).value)
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value
        }
        if ids:
            return ids
    return {}



def write_data_sheet(out, title: str, columns: list[str], rows: list[list[str]], widths: dict[str, int]):
    if title in out.sheetnames:
        del out[title]
    ws = out.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append(row)
    style_header_row(ws)
    ws.freeze_panes = "A2"
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def extract_life_fields_standard(ws, life_row: int, spec_row: int) -> dict[str, str]:
    fields = {
        "NAME OF EQUIPMENT": "",
        "LOCATION": "",
        "EQUIPMENT TAG NAME/APPLICATION": "",
        "DATE OF COMMISSIONING": "",
    }
    for r in range(life_row + 1, spec_row):
        label = norm(ws.cell(r, 2).value)
        if not label:
            continue
        value = row_value_after_label(ws, r)
        if "NAME OF EQUIPMENT" in label:
            fields["NAME OF EQUIPMENT"] = value
        elif "LOCATION" in label:
            fields["LOCATION"] = value
        elif "EQUIPMENT TAG" in label or "TAG NAME" in label:
            fields["EQUIPMENT TAG NAME/APPLICATION"] = value
        elif "DATE OF COMMISSIONING" in label or "COMMISSIONING" in label:
            fields["DATE OF COMMISSIONING"] = value
    if not fields["NAME OF EQUIPMENT"]:
        fields["NAME OF EQUIPMENT"] = ws.title
    return fields


def extract_life_fields_gauge_glass(ws) -> dict[str, str]:
    name = cell_text(ws.cell(5, 3).value)
    location_label = cell_text(ws.cell(5, 6).value)
    location_value = cell_text(ws.cell(6, 6).value)
    location = location_value
    if location_label and location_value:
        location = f"{location_label}: {location_value}"
    elif location_label:
        location = location_label

    tag = cell_text(ws.cell(6, 3).value) or cell_text(ws.cell(6, 2).value)
    commissioning = ""
    for r in range(10, min(ws.max_row, 30) + 1):
        date_val = cell_text(ws.cell(r, 2).value)
        detail = cell_text(ws.cell(r, 3).value)
        if "COMMISSIONED" in norm(detail):
            commissioning = date_val
            break

    return {
        "NAME OF EQUIPMENT": name or ws.title,
        "LOCATION": location,
        "EQUIPMENT TAG NAME/APPLICATION": tag,
        "DATE OF COMMISSIONING": commissioning,
    }


def extract_life_fields(ws) -> dict[str, str]:
    sections = find_section_rows(ws)
    if "life" in sections and "spec" in sections:
        return extract_life_fields_standard(ws, sections["life"], sections["spec"])
    if norm(ws.title) == norm("gauge glass"):
        return extract_life_fields_gauge_glass(ws)
    return {
        "NAME OF EQUIPMENT": ws.title,
        "LOCATION": "",
        "EQUIPMENT TAG NAME/APPLICATION": "",
        "DATE OF COMMISSIONING": "",
    }


def style_header_row(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE}")

    src = openpyxl.load_workbook(SOURCE, data_only=True)
    existing_ids = load_existing_ids()

    mappings: list[tuple[str, str]] = []
    life_rows: list[list[str]] = []
    spec_rows: list[list[str]] = []
    schedule_rows: list[list[str]] = []
    history_rows: list[list[str]] = []

    for sheet_name in src.sheetnames:
        if should_skip_sheet(sheet_name):
            continue
        sheet_id = existing_ids.get(sheet_name) or uuid.uuid4().hex[:12]
        mappings.append((sheet_name, sheet_id))
        ws = src[sheet_name]
        fields = extract_life_fields(ws)
        life_rows.append([
            sheet_id,
            sheet_name,
            fields["NAME OF EQUIPMENT"],
            fields["LOCATION"],
            fields["EQUIPMENT TAG NAME/APPLICATION"],
            fields["DATE OF COMMISSIONING"],
        ])
        spec_rows.extend(extract_specification_rows(ws, sheet_id, sheet_name))
        schedule_rows.extend(extract_maintenance_schedule(ws, sheet_id, sheet_name))
        history_rows.extend(extract_maintenance_history(ws, sheet_id, sheet_name))

    out = openpyxl.Workbook()
    out.remove(out.active)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["sheet name", "id"])
    for sheet_name, sheet_id in mappings:
        map_ws.append([sheet_name, sheet_id])
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    map_ws.column_dimensions["A"].width = 36
    map_ws.column_dimensions["B"].width = 16

    write_data_sheet(
        out,
        SECTION_LIFE,
        LIFE_COLUMNS,
        life_rows,
        {"A": 16, "B": 36, "C": 34, "D": 28, "E": 34, "F": 22},
    )
    write_data_sheet(
        out,
        SECTION_SPEC,
        SPEC_COLUMNS,
        spec_rows,
        {"A": 16, "B": 36, "C": 14, "D": 18, "E": 28, "F": 36},
    )
    write_data_sheet(
        out,
        SECTION_SCHEDULE.strip(),
        SCHEDULE_COLUMNS,
        schedule_rows,
        {
            "A": 16, "B": 36, "C": 8, "D": 22, "E": 42, "F": 8, "G": 8, "H": 8,
            "I": 8, "J": 10, "K": 10, "L": 10, "M": 16,
        },
    )
    write_data_sheet(
        out,
        SECTION_HISTORY,
        HISTORY_COLUMNS,
        history_rows,
        {
            "A": 16, "B": 36, "C": 18, "D": 14, "E": 16, "F": 16,
            "G": 36, "H": 36, "I": 16, "J": 22, "K": 28, "L": 24,
        },
    )

    from datetime import datetime as dt

    stamp = dt.now().strftime("%Y%m%d-%H%M%S")
    fallback = OUTPUT.with_name(f"{OUTPUT.stem}-{stamp}{OUTPUT.suffix}")
    saved = None
    for path in (OUTPUT.with_name(f"{OUTPUT.stem}-updated{OUTPUT.suffix}"), OUTPUT, fallback):
        try:
            out.save(path)
            saved = path
            break
        except PermissionError:
            continue
    if saved is None:
        raise PermissionError(f"Could not write output. Close Excel files and retry. Tried: {fallback}")

    print(f"Source: {SOURCE}")
    print(f"Output: {saved}")
    print(
        f"Sheets: Sheet Map ({len(mappings)} rows), "
        f"{SECTION_LIFE} ({len(life_rows)} rows), "
        f"{SECTION_SPEC} ({len(spec_rows)} rows), "
        f"{SECTION_SCHEDULE.strip()} ({len(schedule_rows)} rows), "
        f"{SECTION_HISTORY} ({len(history_rows)} rows)"
    )


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Extract DS House PCC PANEL & Transformer electrical life-history cards.

Source (hierarchy in INDEX, cards in remaining sheets):
  Extraction_files-sugar/Electrical/1_DS Equipment Life History - PCC PANEL & Transformer.xlsx

Matching rule: data-sheet EQUIPMENT NO / TAG must equal INDEX "History card Tag Nos."
(normalized: trim + collapse spaces + lowercase). Unmatched data sheets are audited only.

Output (DigiLog/backend/backlog-data/mill data/migration files-24-08-26/):
  ds-pcc-panel-transformer-equipment-history-240826.xlsx
    - Hierarchy          (first sheet — importable by db:import-shn-electrical-hierarchy)
    - Sheet Map
    - EQUIPMENT LIFE HISTORY CARD
    - EQUIPMENT SPECIFICATION
    - MAINTENANCE SCHEDULE
    - EQUIPMENT MAINTENANCE HISTORY
  ds-pcc-panel-transformer-extract-audit-240826.xlsx
"""

from __future__ import annotations

import re
import shutil
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from equipment_history_extract_lib import cell_text, norm
from extract_mill_house_electrical import (
    HISTORY_COLUMNS,
    LIFE_COLUMNS,
    SCHEDULE_COLUMNS,
    SPEC_COLUMNS,
    extract_life_fields as elec_extract_life_fields,
    extract_maintenance_history,
    extract_maintenance_schedule,
    extract_specification_rows,
    insert_tag_column,
    prepend_source_file,
    style_header_row,
    write_data_sheet,
)

ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE_FILE = (
    ROOT
    / "Extraction_files-sugar"
    / "Electrical"
    / "1_DS Equipment Life History - PCC PANEL & Transformer.xlsx"
)
OUTPUT_DIR = (
    ROOT
    / "DigiLog"
    / "backend"
    / "backlog-data"
    / "mill data"
    / "migration files-24-08-26"
)
OUTPUT = OUTPUT_DIR / "ds-pcc-panel-transformer-equipment-history-240826.xlsx"
AUDIT_OUTPUT = OUTPUT_DIR / "ds-pcc-panel-transformer-extract-audit-240826.xlsx"
SOURCE_COPY = OUTPUT_DIR / "1_DS Equipment Life History - PCC PANEL & Transformer.xlsx"

SKIP_SHEETS = {"index", "summary index", "summary link", "sheet1"}
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill("solid", fgColor="FFC7CE")
YES_FONT = Font(bold=True, color="006100")
NO_FONT = Font(bold=True, color="9C0006")

HIERARCHY_COLUMNS = [
    "Sr.No.",
    "Plant",
    "Section",
    "Location",
    "Main Equipment",
    " Sub Equipment",
    "Department",
    "History card Tag Nos.",
    "History card Location",
]


def norm_tag(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def clean_name(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def should_skip_sheet(name: str) -> bool:
    return norm(name.strip()).lower() in SKIP_SHEETS


def load_hierarchy_rows(wb) -> list[dict]:
    ws = wb[wb.sheetnames[0]]
    header_row = 1
    for r in range(1, min(12, ws.max_row) + 1):
        cells = [clean_name(ws.cell(r, c).value).lower() for c in range(1, 10)]
        if any("tag" in c for c in cells) and any(
            "section" in c or "equipment" in c for c in cells
        ):
            header_row = r
            break

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        tag = clean_name(ws.cell(r, 8).value)
        sub = clean_name(ws.cell(r, 6).value)
        if not tag and not sub:
            continue
        if tag and "/" not in tag:
            continue
        rows.append(
            {
                "sr": clean_name(ws.cell(r, 1).value),
                "raw_tag": tag,
                "plant": clean_name(ws.cell(r, 2).value),
                "section": clean_name(ws.cell(r, 3).value),
                "location": clean_name(ws.cell(r, 4).value),
                "main_equipment": clean_name(ws.cell(r, 5).value),
                "sub_equipment": sub,
                "department": clean_name(ws.cell(r, 7).value),
                "hist_location": clean_name(ws.cell(r, 9).value),
                "hierarchy_row": r,
            }
        )
    return rows


def first_tag_map(rows: list[dict]) -> dict[str, dict]:
    tag_map: dict[str, dict] = {}
    for meta in rows:
        if meta["raw_tag"] and "/" in meta["raw_tag"]:
            tag_map.setdefault(norm_tag(meta["raw_tag"]), meta)
    return tag_map


def fallback_tag_from_sheet(ws) -> str:
    for r in range(1, min(ws.max_row, 40) + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            val = cell_text(ws.cell(r, c).value)
            if "ZIL/" in val.upper():
                return val
    return ""


def extract_life(ws) -> dict[str, str]:
    fields = elec_extract_life_fields(ws)
    if not fields.get("EQUIPMENT TAG NO"):
        fields["EQUIPMENT TAG NO"] = fallback_tag_from_sheet(ws)
    if not fields.get("NAME OF EQUIPMENT"):
        fields["NAME OF EQUIPMENT"] = ws.title.strip()
    return fields


def save_workbook_to(out: openpyxl.Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        out.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {path.name}. Close it in Excel and retry."
        ) from exc
    return path


def write_hierarchy_sheet(out: openpyxl.Workbook, hierarchy_rows: list[dict]) -> None:
    rows = [
        [
            meta.get("sr") or str(i),
            meta["plant"],
            meta["section"],
            meta["location"],
            meta["main_equipment"],
            meta["sub_equipment"],
            meta["department"],
            meta["raw_tag"],
            meta["hist_location"],
        ]
        for i, meta in enumerate(hierarchy_rows, start=1)
    ]
    write_data_sheet(
        out,
        "Hierarchy",
        HIERARCHY_COLUMNS,
        rows,
        {
            "A": 8,
            "B": 12,
            "C": 14,
            "D": 14,
            "E": 28,
            "F": 48,
            "G": 14,
            "H": 42,
            "I": 36,
        },
    )
    # Keep Hierarchy as the first sheet for db:import-shn-electrical-hierarchy
    sheets = out._sheets
    hier = out["Hierarchy"]
    sheets.remove(hier)
    sheets.insert(0, hier)


def write_audit(
    hierarchy_rows: list[dict],
    matched_sheets: dict[str, list[str]],
    extra_sheets: list[list[str]],
) -> Path:
    rows = []
    for meta in hierarchy_rows:
        nt = norm_tag(meta["raw_tag"])
        sheets = matched_sheets.get(nt, []) if nt else []
        found = bool(sheets)
        rows.append(
            [
                SOURCE_FILE.name,
                meta["raw_tag"] or "(no tag)",
                meta["department"],
                meta["section"],
                meta["location"],
                meta["main_equipment"],
                meta["sub_equipment"],
                meta["hist_location"],
                "Yes" if found else "No",
                ", ".join(sheets),
            ]
        )
    yes_n = sum(1 for r in rows if r[8] == "Yes")
    tagged = [m for m in hierarchy_rows if m["raw_tag"] and "/" in m["raw_tag"]]
    unique = {norm_tag(m["raw_tag"]) for m in tagged}

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_data_sheet(
        out,
        "Summary",
        ["Issue type", "Count"],
        [
            ["Hierarchy equipment rows", len(hierarchy_rows)],
            ["Hierarchy rows with tag", len(tagged)],
            ["Unique hierarchy tags", len(unique)],
            ["Unique tags found in data", len(matched_sheets)],
            ["Hierarchy rows WITH data", yes_n],
            ["Hierarchy rows WITHOUT data", len(rows) - yes_n],
            ["Data sheets not in hierarchy", len(extra_sheets)],
        ],
        {"A": 48, "B": 12},
    )
    ws = write_data_sheet(
        out,
        "All hierarchy rows",
        [
            "source file",
            "History card Tag Nos.",
            "Department",
            "Section",
            "Location",
            "Main Equipment",
            "Sub Equipment",
            "History card Location",
            "Found in data",
            "Matched sheet name(s)",
        ],
        rows,
        {
            "A": 48,
            "B": 42,
            "C": 14,
            "D": 18,
            "E": 16,
            "F": 32,
            "G": 42,
            "H": 28,
            "I": 14,
            "J": 28,
        },
    )
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 9)
        if cell.value == "Yes":
            cell.fill = YES_FILL
            cell.font = YES_FONT
        else:
            cell.fill = NO_FILL
            cell.font = NO_FONT
    write_data_sheet(
        out,
        "Data sheets not in hierarchy",
        ["source file", "sheet name", "EQUIPMENT TAG NO", "NAME OF EQUIPMENT"],
        extra_sheets,
        {"A": 48, "B": 32, "C": 42, "D": 40},
    )
    return save_workbook_to(out, AUDIT_OUTPUT)


def main() -> None:
    print(f"Source: {SOURCE_FILE}")
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(SOURCE_FILE)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_FILE, SOURCE_COPY)
    print(f"Copied source -> {SOURCE_COPY.name}")

    src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    source_name = SOURCE_FILE.name
    hierarchy_rows = load_hierarchy_rows(src)
    hierarchy_tags = first_tag_map(hierarchy_rows)
    print(f"Hierarchy rows: {len(hierarchy_rows)} | unique tags: {len(hierarchy_tags)}")

    all_mappings: list[tuple[str, str, str, str]] = []
    all_life: list[list[str]] = []
    all_spec: list[list[str]] = []
    all_schedule: list[list[str]] = []
    all_history: list[list[str]] = []
    matched_sheets: dict[str, list[str]] = {}
    extra_sheets: list[list[str]] = []

    for sheet_name in src.sheetnames[1:]:
        if should_skip_sheet(sheet_name):
            continue
        ws = src[sheet_name]
        fields = extract_life(ws)
        equip_tag = fields["EQUIPMENT TAG NO"]
        nt = norm_tag(equip_tag)
        meta = hierarchy_tags.get(nt)

        if meta is None:
            extra_sheets.append(
                [
                    source_name,
                    sheet_name,
                    equip_tag or "(blank)",
                    fields.get("NAME OF EQUIPMENT", ""),
                ]
            )
            print(f"  SKIP (tag not in hierarchy): {sheet_name}  tag={equip_tag or '(blank)'}")
            continue

        if not equip_tag:
            print(f"  SKIP (no tag): {sheet_name}")
            continue

        matched_sheets.setdefault(nt, [])
        if sheet_name not in matched_sheets[nt]:
            matched_sheets[nt].append(sheet_name)

        sheet_id = uuid.uuid4().hex[:12]
        out_tag = meta["raw_tag"]
        all_mappings.append((source_name, sheet_name, sheet_id, out_tag))
        all_life.append(
            [
                source_name,
                sheet_id,
                sheet_name,
                out_tag,
                fields["NAME OF EQUIPMENT"] or meta.get("sub_equipment", ""),
                fields["LOCATION"] or meta.get("hist_location", "") or meta.get("location", ""),
                fields["DATE OF COMMISSIONING"],
                meta.get("main_equipment", ""),
                meta.get("sub_equipment", ""),
                meta.get("department", "ELECTRICAL"),
            ]
        )
        all_spec.extend(
            prepend_source_file(
                insert_tag_column(
                    extract_specification_rows(ws, sheet_id, sheet_name), out_tag
                ),
                source_name,
            )
        )
        all_schedule.extend(
            prepend_source_file(
                insert_tag_column(
                    extract_maintenance_schedule(ws, sheet_id, sheet_name), out_tag
                ),
                source_name,
            )
        )
        all_history.extend(
            prepend_source_file(
                insert_tag_column(
                    extract_maintenance_history(ws, sheet_id, sheet_name), out_tag
                ),
                source_name,
            )
        )
        print(f"  OK: {sheet_name} -> {out_tag}")

    src.close()

    missing = [
        m
        for m in hierarchy_rows
        if m["raw_tag"] and "/" in m["raw_tag"] and norm_tag(m["raw_tag"]) not in matched_sheets
    ]
    print(f"\nExtracted cards (hierarchy-matched): {len(all_mappings)}")
    print(f"Hierarchy tags missing in data: {len(missing)}")
    for m in missing:
        print(f"  - {m['raw_tag']}  ({m['sub_equipment']})")
    print(f"Data sheets not in hierarchy: {len(extra_sheets)}")

    out = openpyxl.Workbook()
    out.remove(out.active)

    write_hierarchy_sheet(out, hierarchy_rows)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["source file", "sheet name", "id", "EQUIPMENT TAG NO"])
    for row in all_mappings:
        map_ws.append(list(row))
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    map_ws.column_dimensions["A"].width = 48
    map_ws.column_dimensions["B"].width = 32
    map_ws.column_dimensions["C"].width = 16
    map_ws.column_dimensions["D"].width = 42

    write_data_sheet(
        out,
        "EQUIPMENT LIFE HISTORY CARD",
        LIFE_COLUMNS,
        all_life,
        {
            "A": 48,
            "B": 16,
            "C": 32,
            "D": 42,
            "E": 40,
            "F": 28,
            "G": 20,
            "H": 28,
            "I": 36,
            "J": 14,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT SPECIFICATION",
        SPEC_COLUMNS,
        all_spec,
        {"A": 48, "B": 16, "C": 32, "D": 42, "E": 14, "F": 18, "G": 28, "H": 36},
    )
    write_data_sheet(
        out,
        "MAINTENANCE SCHEDULE",
        SCHEDULE_COLUMNS,
        all_schedule,
        {
            "A": 48,
            "B": 16,
            "C": 32,
            "D": 42,
            "E": 8,
            "F": 22,
            "G": 42,
            "H": 8,
            "I": 8,
            "J": 8,
            "K": 10,
            "L": 12,
            "M": 8,
            "N": 10,
            "O": 10,
            "P": 10,
            "Q": 16,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT MAINTENANCE HISTORY",
        HISTORY_COLUMNS,
        all_history,
        {
            "A": 48,
            "B": 16,
            "C": 32,
            "D": 42,
            "E": 18,
            "F": 14,
            "G": 16,
            "H": 16,
            "I": 36,
            "J": 36,
            "K": 16,
            "L": 22,
            "M": 28,
            "N": 24,
        },
    )

    saved = save_workbook_to(out, OUTPUT)
    audit_path = write_audit(hierarchy_rows, matched_sheets, extra_sheets)
    print(f"\nEXTRACTED: {saved}")
    print(f"AUDIT:     {audit_path}")
    print(
        f"Hierarchy ({len(hierarchy_rows)}), Sheet Map ({len(all_mappings)}), "
        f"LIFE ({len(all_life)}), SPEC ({len(all_spec)}), "
        f"SCHEDULE ({len(all_schedule)}), HISTORY ({len(all_history)})"
    )


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Electrical history data feed.

Reads the electrical team's life-history workbook
(``electrical-equipment-history.xlsx``) and feeds it into the Power Plant
Equipment History (new) ``ppn_*`` tables under the *electrical* department.

Flow:
  1. Parse the workbook (life-history card, specifications, OEM maintenance
     schedule, maintenance history) keyed by the team's sheet name.
  2. Map each sheet name to its frontend hierarchy card using MAPPING below
     (derived from frontend/src/config/powerPlantEquipmentHierarchy.js and the
     agreed mapping sheet). Only the mapping-sheet equipment is fed:
     150TPH ID/FD/SA Fans, HP-01/HP-02, Belt Conveyor 01-08, Slat Chain
     Carrier, Bagasse Elevator. (BFP is intentionally excluded.)
  3. Emit a feed JSON in the format consumed by ppnFeedLib, then invoke the
     existing tested Node importer (``npm run db:import-ppn-feed -- --replace``)
     so the ppn_* schema, __subsections__ meta and history scoping stay
     identical to the application.

Notes:
  * HP Heater -1 / -2 each resolve to a single ppn_equipment row that the
    frontend surfaces on BOTH the 150TPH (HP-01/HP-02) and the 30.85MW STG
    (HP Heater-1/HP Heater-2) cards, because both cards look the record up by
    the same name ("HP Heater -1" / "HP Heater -2"). They are therefore fed
    once. In this workbook they carry only an identity card (no specs).
  * SA Fan -2 shares equip_no ZIL/GSM/PP/14 with SA Fan -1 in the workbook, so
    it is fed by name only (equip_no blanked) to avoid the importer matching
    both onto the same row.

Usage (from anywhere):
  python DigiLog/scripts/electrical_history_data_feed.py            # build + import (replace)
  python DigiLog/scripts/electrical_history_data_feed.py --dry-run  # build + importer dry-run
  python DigiLog/scripts/electrical_history_data_feed.py --no-import # build JSON only
  python DigiLog/scripts/electrical_history_data_feed.py --xlsx PATH --out PATH
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with: pip install openpyxl")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "backend"))
DEFAULT_XLSX = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "..", "electrical-equipment-history.xlsx")
)
FEED_REL = os.path.join(
    "scripts", "data_feed_power_history", "feed-data", "electrical_history_feed.json"
)
DEFAULT_OUT = os.path.join(BACKEND_DIR, FEED_REL)


# Excel sheet name -> (hierarchy_card, category, subcategory, equip_no)
# equip_no is taken from the frontend hierarchy leaf; blank means "match by
# name only" (used for SA Fan -02, a named() leaf with no equip number).
MAPPING: dict[str, tuple[str, str, str, str]] = {
    "ID Fan -1": ("ID Fan -01", "150TPH BLR", "Auxiliary Equipment", "ZIL/GSM/PP/10"),
    "ID Fan-2": ("ID Fan -02", "150TPH BLR", "Auxiliary Equipment", "ZIL/GSM/PP/11"),
    "FD Fan-1": ("FD Fan -01", "150TPH BLR", "Auxiliary Equipment", "ZIL/GSM/PP/12"),
    "FD Fan-2": ("FD Fan -02", "150TPH BLR", "Auxiliary Equipment", "ZIL/GSM/PP/13"),
    "SA Fan -1": ("SA Fan -01", "150TPH BLR", "Auxiliary Equipment", "ZIL/GSM/PP/14"),
    "SA Fan -2": ("SA Fan -02", "150TPH BLR", "Auxiliary Equipment", ""),
    "HP Heater -1": ("HP-01", "150TPH BLR", "Auxiliary Equipment", "ZIL/GSM/PP/04"),
    "HP Heater -2": ("HP-02", "150TPH BLR", "Auxiliary Equipment", "ZIL/GSM/PP/05"),
    "BC-1": ("Belt Conveyor -01", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/15"),
    "BC-2": ("Belt Conveyor -02", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/16"),
    "BC-3": ("Belt Conveyor -03", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/17"),
    "BC-4": ("Belt Conveyor -04", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/18"),
    "BC-5": ("Belt Conveyor -05", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/19"),
    "BC-6": ("Belt Conveyor -06", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/20"),
    "BC-7": ("Belt Conveyor -07", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/21"),
    "BC-8": ("Belt Conveyor -08", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/22"),
    "Slat Chain": ("Slat Chain Carrier", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/24"),
    "Bagasse Elevator": ("Bagasse Elevator", "150TPH BLR", "Fuel Handling System (Phase-1)", "ZIL/GSM/PP/23"),
}

SHEET_MAP = "Sheet Map"
SHEET_CARD = "EQUIPMENT LIFE HISTORY CARD"
SHEET_SPEC = "EQUIPMENT SPECIFICATION"
SHEET_SCHED = "MAINTENANCE SCHEDULE"
SHEET_HIST = "EQUIPMENT MAINTENANCE HISTORY"

# Excel interval column -> ppn feed key. Daily / 4-Years have no ppn column;
# when marked they are appended to the activity text so nothing is lost.
SCHEDULE_INTERVALS = {
    "Weekly": "iv_W",
    "Monthly": "iv_M",
    "Quarterly": "iv_Q",
    "Half - Yearly": "iv_H",
    "Yearly": "iv_Y",
    "2 - Years": "iv_T",
    "3 - Years": "iv_3Y",
}
EXTRA_INTERVALS = {"Daily": "Daily", "4 - Years": "4-Yearly"}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, dt.date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def clean_date(value) -> str:
    """Like clean(), but normalize ISO date text (YYYY-MM-DD) to DD.MM.YYYY."""
    s = clean(value)
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        try:
            return dt.datetime.strptime(s, "%Y-%m-%d").strftime("%d.%m.%Y")
        except ValueError:
            return s
    return s


def normalize_spec_value(lbl: str, val: str) -> str:
    """Efficiency is recorded as a fraction (0.965). Show it as a percentage."""
    if "effic" not in lbl.strip().lower():
        return val
    num_text = val.strip().rstrip("%").strip()
    try:
        num = float(num_text)
    except ValueError:
        return val
    if num <= 1.5:  # fraction -> percentage
        num *= 100
    text = f"{num:.2f}".rstrip("0").rstrip(".")
    return f"{text} %"


def is_blank_row(row) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def iter_records(ws):
    """Yield each non-empty row of a sheet as a dict keyed by trimmed headers."""
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return
    keys = [clean(h) for h in header]
    for row in it:
        if is_blank_row(row):
            continue
        yield dict(zip(keys, row))


def load_workbook_data(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    cards: dict[str, dict] = {}
    for r in iter_records(wb[SHEET_CARD]):
        name = clean(r.get("sheet name"))
        if not name:
            continue
        cards[name] = {
            "image_name": clean(r.get("NAME OF EQUIPMENT")),
            "location": clean(r.get("LOCATION")) or "POWER PLANT",
            "equip_no": clean(r.get("EQUIPMENT TAG NAME/APPLICATION")),
            "commissioned": clean_date(r.get("DATE OF COMMISSIONING")),
        }

    specs: dict[str, list] = {}
    for r in iter_records(wb[SHEET_SPEC]):
        name = clean(r.get("sheet name"))
        if not name:
            continue
        lbl = clean(r.get("Parameter label"))
        val = normalize_spec_value(lbl, clean(r.get("Parameter value")))
        if not lbl and not val:
            continue
        spec = {"lbl": lbl, "val": val}
        section = clean(r.get("section"))
        sub_section = clean(r.get("sub_section"))
        if section:
            spec["section"] = section
        if sub_section:
            spec["sub_section"] = sub_section
        specs.setdefault(name, []).append(spec)

    schedule: dict[str, list] = {}
    for r in iter_records(wb[SHEET_SCHED]):
        name = clean(r.get("sheet name"))
        if not name:
            continue
        act = clean(r.get("Maintenance / Inspection Activities"))
        comp = clean(r.get("Name of Equipment"))
        if not act and not comp:
            continue
        extras = [
            label for col, label in EXTRA_INTERVALS.items()
            if clean(r.get(col))
        ]
        if extras:
            act = f"{act} [{', '.join(extras)}]" if act else f"[{', '.join(extras)}]"
        entry = {
            "no": clean(r.get("Sr.No.")),
            "comp": comp,
            "act": act,
        }
        for col, key in SCHEDULE_INTERVALS.items():
            if clean(r.get(col)):
                entry[key] = "X"
        schedule.setdefault(name, []).append(entry)

    history: dict[str, list] = {}
    for r in iter_records(wb[SHEET_HIST]):
        name = clean(r.get("sheet name"))
        if not name:
            continue
        entry = {
            "season": clean(r.get("Season / OFF Season")),
            "year": clean(r.get("Year")),
            "date_start": clean(r.get("Date of Start")),
            "date_finish": clean(r.get("Date of Finish")),
            "obs": clean(r.get("Outage/ Observation")),
            "act": clean(r.get("Action Taken")),
            "cost": clean(r.get("Repair Cost (Rs.)")),
            "svc": clean(r.get("Services (Internal / External)")),
            "resp": clean(r.get("Responsibility ( Engineer/ Supervision)")),
            "rem": clean(r.get("Remarks")),
        }
        if any(entry[k] for k in ("obs", "act", "year", "date_start")):
            history.setdefault(name, []).append(entry)

    wb.close()
    return cards, specs, schedule, history


def build_feed(xlsx_path: str):
    cards, specs, schedule, history = load_workbook_data(xlsx_path)

    equipment = []
    matched, missing_card, no_specs = [], [], []

    for sort_order, (sheet_name, (card, category, subcategory, equip_no)) in enumerate(
        MAPPING.items()
    ):
        card_info = cards.get(sheet_name)
        if card_info is None:
            missing_card.append(sheet_name)
            continue

        spec_rows = specs.get(sheet_name, [])
        if not spec_rows:
            no_specs.append(sheet_name)

        record = {
            "hierarchy_name": sheet_name,
            "hierarchy_card": card,
            "hierarchy_path": f"Power Plant > {category} > {subcategory} > {card}",
            "image_name": card_info["image_name"],
            "name": sheet_name,
            "equip_no": equip_no,
            "tag_name": "",
            "category": category,
            "subcategory": subcategory,
            "location": card_info["location"],
            "commissioned": card_info["commissioned"],
            "drive": "",
            "history_section": "electrical",
            "specs": spec_rows,
            "schedule": schedule.get(sheet_name, []),
            "history": history.get(sheet_name, []),
        }
        equipment.append(record)
        matched.append(sheet_name)

    return {"equipment": equipment}, {
        "matched": matched,
        "missing_card": missing_card,
        "no_specs": no_specs,
    }


def run_importer(out_path: str, dry_run: bool) -> int:
    rel = os.path.relpath(out_path, BACKEND_DIR)
    args = ["npm", "run", "db:import-ppn-feed", "--", "--file", rel]
    args.append("--dry-run" if dry_run else "--replace")
    print(f"\nRunning importer in {BACKEND_DIR}:\n  {' '.join(args)}\n")
    proc = subprocess.run(args, cwd=BACKEND_DIR, shell=(os.name == "nt"))
    return proc.returncode


def main() -> int:
    parser = argparse.ArgumentParser(description="Electrical history data feed")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Source workbook path")
    parser.add_argument("--out", default=DEFAULT_OUT, help="Feed JSON output path")
    parser.add_argument("--dry-run", action="store_true", help="Run importer in dry-run mode")
    parser.add_argument("--no-import", action="store_true", help="Only write JSON; skip importer")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"Workbook not found: {args.xlsx}")

    feed, report = build_feed(args.xlsx)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(feed, fh, indent=2, ensure_ascii=False)

    print(f"Wrote {len(feed['equipment'])} equipment record(s) -> {args.out}")
    for rec in feed["equipment"]:
        print(
            f"  + {rec['name']:<18} -> {rec['hierarchy_card']:<20} "
            f"specs={len(rec['specs']):<3} schedule={len(rec['schedule'])} "
            f"history={len(rec['history'])}"
        )
    if report["missing_card"]:
        print("\n  ! Not found in workbook:", ", ".join(report["missing_card"]))
    if report["no_specs"]:
        print(
            "\n  i Identity only (no electrical specs in workbook):",
            ", ".join(report["no_specs"]),
        )

    if args.no_import:
        print("\n--no-import set: skipping database import.")
        return 0

    return run_importer(args.out, args.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())

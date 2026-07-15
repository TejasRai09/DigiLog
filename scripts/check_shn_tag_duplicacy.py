#!/usr/bin/env python3
"""
Compare Inst. History card Tag nas. from Plant Instrument list
(Instrument_working_sheet only) against Turbine life history workbook.

Output: sugar-house-tag-duplicacy-check.xlsx
"""

from __future__ import annotations

import re
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from equipment_history_extract_lib import cell_text, norm

ROOT = Path(__file__).resolve().parent.parent.parent
PLANT_LIST = ROOT / "Plant Instrument Equipment List-11-07-2026.xlsx"
TURBINE_BOOK = ROOT / "Turbine & Instrument Equipment life histroy 20-06-2026.xlsx"
OUTPUT = ROOT / "sugar-house-tag-duplicacy-check.xlsx"

PLANT_SHEET = "Instrument_working_sheet"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
DUP_FILL = PatternFill("solid", fgColor="FFC7CE")


def norm_tag(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def compact_tag(value) -> str:
    return re.sub(r"\s+", "", norm_tag(value))


def row_value_after_label(ws, row_idx: int, label_col: int) -> str:
    for c in range(label_col + 1, ws.max_column + 1):
        val = cell_text(ws.cell(row_idx, c).value)
        if val:
            return val
    return ""


def extract_plant_tags(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if PLANT_SHEET not in wb.sheetnames:
        raise KeyError(
            f"Sheet {PLANT_SHEET!r} not found. Available: {wb.sheetnames}",
        )
    ws = wb[PLANT_SHEET]
    headers = [str(c.value or "").strip() for c in ws[1]]
    tag_col = headers.index("Inst. History card Tag nas.") + 1
    tags: list[str] = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, tag_col).value
        if val is not None and str(val).strip():
            tags.append(str(val).strip())
    wb.close()
    return tags


def extract_turbine_equipment_tags(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    tags: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            for c in (1, 2):
                label = norm(ws.cell(r, c).value)
                if label != "EQUIPMENT NO:":
                    continue
                val = row_value_after_label(ws, r, c)
                if val:
                    tags.append(val)
                break
    wb.close()
    return tags


def turbine_tag_matches_plant(plant_tag: str, turbine_tag: str) -> bool:
    p = norm_tag(plant_tag)
    t = norm_tag(turbine_tag)
    if not p or not t:
        return False
    if p == t:
        return True
    if t.startswith(f"{p} ") or t.startswith(f"{p}("):
        return True
    if compact_tag(plant_tag) == compact_tag(turbine_tag):
        return True
    # Turbine tag may omit trailing segment present in plant list (rare).
    if p.startswith(t + " ") or p.startswith(t + "("):
        return True
    return False


def count_in_turbine(plant_tag: str, turbine_tags: list[str]) -> int:
    return sum(1 for t in turbine_tags if turbine_tag_matches_plant(plant_tag, t))


def style_header(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def save_workbook(wb: openpyxl.Workbook) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    for path in (
        OUTPUT.with_name(f"{OUTPUT.stem}-updated{OUTPUT.suffix}"),
        OUTPUT,
        OUTPUT.with_name(f"{OUTPUT.stem}-{stamp}{OUTPUT.suffix}"),
    ):
        try:
            wb.save(path)
            return path
        except PermissionError:
            continue
    raise PermissionError(f"Could not write {OUTPUT}")


def main() -> None:
    if not PLANT_LIST.exists():
        raise FileNotFoundError(PLANT_LIST)
    if not TURBINE_BOOK.exists():
        raise FileNotFoundError(TURBINE_BOOK)

    plant_tags = extract_plant_tags(PLANT_LIST)
    turbine_tags = extract_turbine_equipment_tags(TURBINE_BOOK)

    plant_counter = Counter(norm_tag(t) for t in plant_tags)
    turbine_counter = Counter(norm_tag(t) for t in turbine_tags)

    unique_plant = sorted({norm_tag(t): t for t in plant_tags}.values(), key=norm_tag)

    summary_rows: list[list] = []
    for tag in unique_plant:
        nt = norm_tag(tag)
        plant_count = plant_counter[nt]
        turbine_count = count_in_turbine(tag, turbine_tags)
        summary_rows.append([
            tag,
            plant_count,
            turbine_count,
            "Yes" if plant_count > 1 else "No",
            "Yes" if turbine_count > 1 else "No",
            "Yes" if turbine_count == 0 else "No",
        ])

    detail_rows: list[list] = []
    for tag in plant_tags:
        nt = norm_tag(tag)
        turbine_count = count_in_turbine(tag, turbine_tags)
        detail_rows.append([tag, turbine_count, "Yes" if turbine_count > 1 else "No", "Yes" if turbine_count == 0 else "No"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Tag Count Summary")
    headers = [
        "Inst. History card Tag nas.",
        "Count in Plant Instrument List",
        "Count in Turbine Life History",
        "Duplicate in Plant List",
        "Duplicate in Turbine History",
        "Not found in Turbine History",
    ]
    ws.append(headers)
    for row in summary_rows:
        ws.append(row)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 42
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22

    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 4).value == "Yes" or ws.cell(r, 5).value == "Yes":
            for c in range(1, 7):
                ws.cell(r, c).fill = DUP_FILL

    ws2 = wb.create_sheet("All Plant Rows")
    ws2.append([
        "Inst. History card Tag nas.",
        "Count in Turbine Life History",
        "Duplicate in Turbine History",
        "Not found in Turbine History",
    ])
    for row in detail_rows:
        ws2.append(row)
    style_header(ws2)
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 42
    for col in "BCD":
        ws2.column_dimensions[col].width = 24

    ws3 = wb.create_sheet("Turbine Tag Duplicates")
    ws3.append(["Equipment tag (Turbine file)", "Count in Turbine file"])
    for tag, count in sorted(turbine_counter.items(), key=lambda x: (-x[1], x[0])):
        if count > 1:
            ws3.append([tag, count])
    style_header(ws3)
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 22

    saved = save_workbook(wb)

    not_found = sum(1 for row in summary_rows if row[5] == "Yes")
    dup_plant = sum(1 for row in summary_rows if row[3] == "Yes")
    dup_turbine = sum(1 for row in summary_rows if row[4] == "Yes")

    print(f"Plant list tags (rows): {len(plant_tags)}")
    print(f"Unique plant tags: {len(unique_plant)}")
    print(f"Turbine equipment-no entries: {len(turbine_tags)}")
    print(f"Unique turbine tags: {len(turbine_counter)}")
    print(f"Not found in turbine: {not_found} unique tags")
    print(f"Duplicate in plant list: {dup_plant} unique tags")
    print(f"Duplicate in turbine: {dup_turbine} unique tags")
    print(f"Output: {saved}")


if __name__ == "__main__":
    main()

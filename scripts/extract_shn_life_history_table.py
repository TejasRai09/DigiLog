#!/usr/bin/env python3
"""
Build normalized life-history workbook from sugar-house equipment cards.

Source: Turbine & Instrument Equipment life histroy 20-06-2026.xlsx
        (multi-card sheets — several equipment blocks per tab)

Output: sugar-house-equipment-life-history.xlsx (or -updated if file is open)
  - Sheet Map
  - EQUIPMENT LIFE HISTORY CARD
  - EQUIPMENT SPECIFICATION
  - MAINTENANCE SCHEDULE
  - EQUIPMENT MAINTENANCE HISTORY
"""

from __future__ import annotations

import argparse
import re
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from equipment_history_extract_lib import (
    cell_text,
    extract_history_rows_flexible,
    extract_schedule_rows,
    extract_specification_rows as extract_spec_rows_lib,
    norm,
    parse_schedule_layout,
)
from shn_tag_match import load_tag_filter, matches_any_filter, norm_tag

ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE = ROOT / "Turbine & Instrument Equipment life histroy 20-06-2026.xlsx"
OUTPUT = ROOT / "sugar-house-equipment-life-history.xlsx"

SECTION_LIFE = "EQUIPMENT LIFE HISTORY CARD"
SECTION_SPEC = "EQUIPMENT SPECIFICATION"
SECTION_SCHEDULE_MARKERS = ("MAINTENANCE SCHEDULE", "OEM MAINTENANCE SCHEDULE")
SECTION_HISTORY = "EQUIPMENT MAINTENANCE HISTORY"

TAG_COLUMN = "EQUIPMENT TAG NAME/APPLICATION"

SPEC_COLUMNS = [
    "sheet id",
    "sheet name",
    TAG_COLUMN,
    "NAME OF EQUIPMENT",
    "section",
    "sub_section",
    "Parameter label",
    "Parameter value",
]

LIFE_COLUMNS = [
    "sheet id",
    "sheet name",
    TAG_COLUMN,
    "NAME OF EQUIPMENT",
    "LOCATION",
    "DATE OF COMMISSIONING",
]

HISTORY_COLUMNS = [
    "sheet id",
    "sheet name",
    TAG_COLUMN,
    "Season / OFF Season",
    "Year",
    "Date of Start",
    "Date of Finish",
    "Outage/ Observation",
    "Action Taken",
    "Repair Cost (Rs.)",
    "Services (Internal / External)",
    "Responsibility ( Engineer/ Supervision)",
    "Remarks",
]

SCHEDULE_INTERVALS = [
    "Daily",
    "Weekly",
    "Monthly",
    "Quarterly",
    "Half - Yearly",
    "Yearly",
    "2 - Years",
    "3 - Years",
    "4 - Years",
]

SCHEDULE_COLUMNS = [
    "sheet id",
    "sheet name",
    TAG_COLUMN,
    "Sr.No.",
    "Name of Equipment",
    "Maintenance / Inspection Activities",
    *SCHEDULE_INTERVALS,
    "Remarks",
]

SPEC_SECTION_HEADERS = (
    ("MECHANICAL PART", "mechanical"),
    ("CIVIL PART", "civil"),
    ("INSTRUMENT PART", "instrument"),
    ("ELECTRICALT PART", "electrical"),
    ("ELECTRIC PART", "electrical"),
    ("ELECTRICAL PART", "electrical"),
    ("ELECTRONIC PART", "electrical"),
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def is_life_card_header(text: str) -> bool:
    t = norm(text)
    return t == SECTION_LIFE or (SECTION_LIFE in t and len(t) < 50)


def sheet_has_life_cards(ws) -> bool:
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            if is_life_card_header(ws.cell(r, c).value):
                return True
    return False


def should_skip_sheet(name: str, ws) -> bool:
    if not sheet_has_life_cards(ws):
        return True
    n = norm(name)
    if n in {"SHEET1"}:
        return True
    return False


def row_value_after_label(ws, row_idx: int, label_col: int) -> str:
    for c in range(label_col + 1, ws.max_column + 1):
        val = cell_text(ws.cell(row_idx, c).value)
        if val:
            return val
    return ""


def value_from_labeled_cell(raw) -> tuple[str, str]:
    """Split ``LABEL : value`` / ``LABEL:`` in one cell into (label_norm_hint, value)."""
    text = cell_text(raw)
    if not text:
        return "", ""
    # Prefer splitting on the first colon used as a label separator.
    if ":" in text:
        left, right = text.split(":", 1)
        return left.strip(), right.strip()
    return text.strip(), ""


def labeled_row_value(ws, row_idx: int, label_cols: tuple[int, ...] = (1, 2)) -> tuple[str, str]:
    """Return (normalized label key fragment, value) for a life-card field row."""
    for col in label_cols:
        raw = ws.cell(row_idx, col).value
        if raw is None or str(raw).strip() == "":
            continue
        label_part, inline_value = value_from_labeled_cell(raw)
        label_n = norm(label_part)
        if not label_n:
            continue
        value = inline_value or row_value_after_label(ws, row_idx, col)
        return label_n, value
    return "", ""


def is_equip_no_label(text) -> bool:
    return norm(text) == "EQUIPMENT NO:" or norm(text).startswith("EQUIPMENT NO")


def find_name_row(ws, equip_row: int, label_col: int) -> int | None:
    for r in range(equip_row - 1, max(1, equip_row - 10), -1):
        for c in (label_col, 1, 2):
            label = norm(ws.cell(r, c).value)
            if "NAME OF EQUIPMENT" in label:
                return r
    return None


def find_sections_in_range(ws, start_row: int, end_row: int) -> dict[str, int]:
    sections: dict[str, int] = {}
    for r in range(start_row, end_row + 1):
        for c in (1, 2):
            text = norm(ws.cell(r, c).value)
            if not text:
                continue
            if SECTION_SPEC in text and "spec" not in sections:
                sections["spec"] = r
            elif any(marker in text for marker in SECTION_SCHEDULE_MARKERS) and "schedule" not in sections:
                sections["schedule"] = r
            elif SECTION_HISTORY in text and "history" not in sections:
                sections["history"] = r
    return sections


def find_equipment_subcards(ws) -> list[dict]:
    """One sub-card per EQUIPMENT NO row (handles stacked cards without repeated life headers)."""
    anchors: list[tuple[int, int]] = []
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            if is_equip_no_label(ws.cell(r, c).value):
                if row_value_after_label(ws, r, c):
                    anchors.append((r, c))
                break

    if not anchors:
        return []

    subcards: list[dict] = []
    for i, (equip_row, label_col) in enumerate(anchors):
        name_row = find_name_row(ws, equip_row, label_col)
        if i + 1 < len(anchors):
            next_name = find_name_row(ws, anchors[i + 1][0], anchors[i + 1][1])
            block_end = (next_name - 1) if next_name else anchors[i + 1][0] - 1
        else:
            block_end = ws.max_row
            for r in range(equip_row + 1, ws.max_row + 1):
                for c in (1, 2):
                    if is_life_card_header(ws.cell(r, c).value):
                        block_end = r - 1
                        break
                else:
                    continue
                break

        scan_start = name_row or max(1, equip_row - 4)
        sections = find_sections_in_range(ws, equip_row, block_end)
        life_header_row = scan_start
        for r in range(scan_start, equip_row):
            for c in (1, 2):
                if is_life_card_header(ws.cell(r, c).value):
                    life_header_row = r
                    break

        subcards.append({
            "life_row": life_header_row,
            "name_row": scan_start,
            "equip_row": equip_row,
            "label_col": label_col,
            "block_end": block_end,
            **sections,
        })
    return subcards


def find_equipment_blocks(ws) -> list[dict]:
    """Backward-compatible alias."""
    return find_equipment_subcards(ws)


def extract_life_fields(ws, block: dict) -> dict[str, str]:
    fields = {
        "NAME OF EQUIPMENT": "",
        "LOCATION": "",
        "EQUIPMENT TAG NAME/APPLICATION": "",
        "DATE OF COMMISSIONING": "",
    }
    scan_start = block.get("name_row", block.get("life_row", block.get("equip_row", 1)))
    scan_end = block.get("spec", block["block_end"] + 1)
    equip_no = ""

    for r in range(scan_start, scan_end):
        label, value = labeled_row_value(ws, r)
        if not label:
            continue
        value = re.sub(r"\s+", " ", value).strip() if value else ""
        if "NAME OF EQUIPMENT" in label:
            if value:
                fields["NAME OF EQUIPMENT"] = value
        elif label.startswith("LOCATION") or label == "LOCATION":
            if value:
                fields["LOCATION"] = value
        elif "EQUIPMENT TAG" in label or label.startswith("TAG NO") or label == "TAG NAME":
            if value:
                fields["EQUIPMENT TAG NAME/APPLICATION"] = value
        elif "EQUIPMENT NO" in label and "TAG" not in label:
            if value:
                equip_no = value
        elif "DATE OF COMMISSIONING" in label or "COMMISSIONING" in label:
            if value:
                fields["DATE OF COMMISSIONING"] = value

    if not fields["EQUIPMENT TAG NAME/APPLICATION"] and equip_no:
        fields["EQUIPMENT TAG NAME/APPLICATION"] = equip_no
    # Do not fall back to the Excel tab title — leave blank if the card has no name.
    return fields


def make_output_sheet_name(tab_name: str, fields: dict[str, str], card_index: int) -> str:
    tag = fields.get("EQUIPMENT TAG NAME/APPLICATION") or ""
    name = fields.get("NAME OF EQUIPMENT") or ""
    if tag:
        return f"{tab_name} :: {tag}"
    if name:
        return f"{tab_name} :: {name}"
    return f"{tab_name} :: card-{card_index}"


def insert_tag_column(rows: list[list[str]], equip_tag: str) -> list[list[str]]:
    """Insert equipment tag after sheet id and sheet name."""
    return [row[:2] + [equip_tag] + row[2:] for row in rows]


def insert_spec_meta_columns(
    rows: list[list[str]],
    equip_tag: str,
    equipment_name: str,
) -> list[list[str]]:
    """Insert tag + NAME OF EQUIPMENT after sheet id and sheet name."""
    return [row[:2] + [equip_tag, equipment_name] + row[2:] for row in rows]


def extract_specification_rows(ws, block: dict, sheet_id: str, sheet_name: str) -> list[list[str]]:
    if "spec" not in block:
        return []

    spec_row = block["spec"]
    spec_end = block["block_end"] + 1
    for key in ("schedule", "history"):
        if key in block and block[key] > spec_row:
            spec_end = min(spec_end, block[key])

    return extract_spec_rows_lib(
        ws,
        spec_row,
        spec_end,
        sheet_id,
        sheet_name,
        (SECTION_HISTORY, SECTION_LIFE, *SECTION_SCHEDULE_MARKERS),
        section_headers=SPEC_SECTION_HEADERS,
    )


def extract_maintenance_schedule(ws, block: dict, sheet_id: str, sheet_name: str) -> list[list[str]]:
    if "schedule" not in block:
        return []

    schedule_row = block["schedule"]
    schedule_end = block.get("history", block["block_end"] + 1)
    layout = parse_schedule_layout(ws, schedule_row, schedule_end)
    if not layout:
        return []

    if layout.data_start > block["block_end"]:
        return []

    return extract_schedule_rows(
        ws,
        layout,
        min(schedule_end, block["block_end"] + 1),
        sheet_id,
        sheet_name,
        tuple(SCHEDULE_INTERVALS),
        (SECTION_HISTORY, *SECTION_SCHEDULE_MARKERS, SECTION_LIFE),
    )


def extract_maintenance_history(ws, block: dict, sheet_id: str, sheet_name: str) -> list[list[str]]:
    if "history" not in block:
        return []

    history_row = block["history"]
    return extract_history_rows_flexible(
        ws,
        history_row,
        sheet_id,
        sheet_name,
        stop_markers=(SECTION_LIFE, SECTION_SPEC, *SECTION_SCHEDULE_MARKERS),
    )


def style_header_row(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_data_sheet(out, title: str, columns: list[str], rows: list[list[str]], widths: dict[str, int]):
    if title in out.sheetnames:
        del out[title]
    ws = out.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append(row)
    style_header_row(ws)
    ws.freeze_panes = "A2"
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def load_existing_ids(output: Path) -> dict[str, str]:
    candidates = (
        output,
        output.with_name(f"{output.stem}-updated{output.suffix}"),
    )
    for path in candidates:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Sheet Map" not in wb.sheetnames:
            continue
        ws = wb["Sheet Map"]
        ids = {
            cell_text(ws.cell(r, 1).value): cell_text(ws.cell(r, 2).value)
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value
        }
        if ids:
            return ids
    return {}


def save_workbook(out: openpyxl.Workbook, output: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    candidates = (
        output.with_name(f"{output.stem}-updated{output.suffix}"),
        output,
        output.with_name(f"{output.stem}-{stamp}{output.suffix}"),
    )
    for path in candidates:
        try:
            out.save(path)
            return path
        except PermissionError:
            continue
    raise PermissionError(f"Could not write output. Close Excel and retry. Tried: {candidates[-1]}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract sugar-house equipment life history cards to normalized Excel.",
    )
    parser.add_argument(
        "--tags-file",
        type=Path,
        help="Only extract equipment whose tag matches a line in this file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT,
        help=f"Output workbook path (default: {OUTPUT.name})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = args.output.resolve()
    tag_filter: set[str] | None = None
    if args.tags_file:
        tag_filter = load_tag_filter(args.tags_file.resolve())

    if not SOURCE.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE}")

    src = openpyxl.load_workbook(SOURCE, data_only=True)
    existing_ids = load_existing_ids(output_path)

    mappings: list[tuple[str, str, str]] = []
    life_rows: list[list[str]] = []
    spec_rows: list[list[str]] = []
    schedule_rows: list[list[str]] = []
    history_rows: list[list[str]] = []

    skipped_tabs: list[str] = []
    matched_tags: set[str] = set()

    for tab_name in src.sheetnames:
        ws = src[tab_name]
        if should_skip_sheet(tab_name, ws):
            skipped_tabs.append(tab_name)
            continue

        blocks = find_equipment_subcards(ws)
        for card_index, block in enumerate(blocks, start=1):
            fields = extract_life_fields(ws, block)
            equip_tag = fields["EQUIPMENT TAG NAME/APPLICATION"]
            if tag_filter is not None and not matches_any_filter(equip_tag, tag_filter):
                continue

            if tag_filter is not None:
                for allowed in tag_filter:
                    if matches_any_filter(equip_tag, {allowed}):
                        matched_tags.add(allowed)

            output_name = make_output_sheet_name(tab_name, fields, card_index)
            sheet_id = existing_ids.get(output_name) or uuid.uuid4().hex[:12]
            mappings.append((output_name, sheet_id, equip_tag))

            life_rows.append([
                sheet_id,
                output_name,
                equip_tag,
                fields["NAME OF EQUIPMENT"],
                fields["LOCATION"],
                fields["DATE OF COMMISSIONING"],
            ])
            spec_rows.extend(insert_spec_meta_columns(
                extract_specification_rows(ws, block, sheet_id, output_name),
                equip_tag,
                fields["NAME OF EQUIPMENT"],
            ))
            schedule_rows.extend(insert_tag_column(
                extract_maintenance_schedule(ws, block, sheet_id, output_name),
                equip_tag,
            ))
            history_rows.extend(insert_tag_column(
                extract_maintenance_history(ws, block, sheet_id, output_name),
                equip_tag,
            ))

    out = openpyxl.Workbook()
    out.remove(out.active)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["sheet name", "id", TAG_COLUMN])
    for sheet_name, sheet_id, equip_tag in mappings:
        map_ws.append([sheet_name, sheet_id, equip_tag])
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    map_ws.column_dimensions["A"].width = 52
    map_ws.column_dimensions["B"].width = 16
    map_ws.column_dimensions["C"].width = 34

    write_data_sheet(
        out, SECTION_LIFE, LIFE_COLUMNS, life_rows,
        {"A": 16, "B": 52, "C": 34, "D": 34, "E": 36, "F": 22},
    )
    write_data_sheet(
        out, SECTION_SPEC, SPEC_COLUMNS, spec_rows,
        {"A": 16, "B": 52, "C": 34, "D": 34, "E": 14, "F": 18, "G": 28, "H": 36},
    )
    write_data_sheet(
        out, "MAINTENANCE SCHEDULE", SCHEDULE_COLUMNS, schedule_rows,
        {
            "A": 16, "B": 52, "C": 34, "D": 8, "E": 22, "F": 42,
            "G": 8, "H": 8, "I": 8, "J": 10, "K": 12, "L": 8, "M": 10, "N": 10, "O": 10, "P": 16,
        },
    )
    write_data_sheet(
        out, SECTION_HISTORY, HISTORY_COLUMNS, history_rows,
        {
            "A": 16, "B": 52, "C": 34, "D": 18, "E": 14, "F": 16, "G": 16,
            "H": 36, "I": 36, "J": 16, "K": 22, "L": 28, "M": 24,
        },
    )

    saved = save_workbook(out, output_path)
    print(f"Source: {SOURCE}")
    print(f"Output: {saved}")
    if tag_filter is not None:
        missing = sorted(tag_filter - matched_tags, key=norm_tag)
        print(f"Tag filter: {len(tag_filter)} requested, {len(matched_tags)} matched, {len(missing)} not found")
        if missing:
            print("Tags not found in source (first 20):")
            for tag in missing[:20]:
                print(f"  - {tag}")
            if len(missing) > 20:
                print(f"  ... and {len(missing) - 20} more")
    print(f"Skipped tabs ({len(skipped_tabs)}): {', '.join(skipped_tabs)}")
    print(
        f"Sheets: Sheet Map ({len(mappings)} cards), "
        f"{SECTION_LIFE} ({len(life_rows)}), "
        f"{SECTION_SPEC} ({len(spec_rows)}), "
        f"MAINTENANCE SCHEDULE ({len(schedule_rows)}), "
        f"{SECTION_HISTORY} ({len(history_rows)})"
    )


if __name__ == "__main__":
    main()

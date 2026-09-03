#!/usr/bin/env python3
"""
Extract equipment sections from BLR Machine history card.xlsx into a flat workbook.

Input:  ../BLR Machine history card.xlsx  (project root)
Output: ../blr-machine-history-extract.xlsx

Columns: sheetname, equipment name, EQUIPMENT LIFE HISTORY CARD,
         EQUIPMENT SPECIFICATION, EQUIPMENT MAINTENANCE HISTORY
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter



ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE = ROOT / "BLR Machine history card.xlsx"
OUTPUT = ROOT / "blr-machine-history-extract.xlsx"

SKIP_SHEETS = {"index"}


def should_skip_sheet(name: str) -> bool:
    return norm(name) in {norm(s) for s in SKIP_SHEETS}

SECTION_LIFE = "EQUIPMENT LIFE HISTORY CARD"
SECTION_SPEC = "EQUIPMENT SPECIFICATION"
SECTION_SCHEDULE = "MAINTENANCE SCHEDULE"
SECTION_HISTORY = "EQUIPMENT MAINTENANCE HISTORY"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def norm(text) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip()).upper()


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def row_values(ws, row_idx: int, max_col: int | None = None) -> list[str]:
    last = max_col or ws.max_column
    return [cell_text(ws.cell(row_idx, c).value) for c in range(1, last + 1)]


def format_row(vals: list[str]) -> str:
    parts: list[str] = []
    # Label col B + value col E; optional second pair col H + col K.
    for label_idx, value_idx in ((1, 4), (7, 10)):
        if label_idx < len(vals) and value_idx < len(vals):
            label, value = vals[label_idx], vals[value_idx]
            if label and value:
                label = label.rstrip()
                if label.endswith(":"):
                    parts.append(f"{label} {value}")
                else:
                    parts.append(f"{label}: {value}")
    if parts:
        return " | ".join(parts)

    non_empty = [v for v in vals if v]
    if not non_empty:
        return ""
    return " | ".join(non_empty)


def find_section_rows(ws) -> dict[str, int]:
    """Return 1-based row numbers for section header rows (column B first, then scan row)."""
    found: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        b = norm(ws.cell(r, 2).value)
        if not b:
            row_join = norm(" ".join(row_values(ws, r)))
            b = row_join
        for label, key in (
            (SECTION_LIFE, "life"),
            (SECTION_SPEC, "spec"),
            (SECTION_SCHEDULE, "schedule"),
            (SECTION_HISTORY, "history"),
        ):
            if key not in found and label in b:
                found[key] = r
    return found


def extract_section(ws, start_row: int, end_row: int) -> str:
    """Extract rows between start (exclusive header) and end (exclusive)."""
    lines: list[str] = []
    for r in range(start_row + 1, end_row):
        line = format_row(row_values(ws, r))
        if line:
            lines.append(line)
    return "\n".join(lines)


def extract_equipment_name(ws, life_header_row: int, spec_row: int) -> str:
    for r in range(life_header_row + 1, spec_row):
        label = norm(ws.cell(r, 2).value)
        if "NAME OF EQUIPMENT" in label:
            # Value usually column E (5)
            val = cell_text(ws.cell(r, 5).value)
            if val:
                return val
            vals = row_values(ws, r)
            for i, v in enumerate(vals):
                if v and "NAME OF EQUIPMENT" not in norm(v):
                    if i > 0:
                        return v
            break
    return cell_text(ws.title)


def parse_alternate_sheet(ws) -> dict[str, str] | None:
    """Fallback for sheets that use the older machine history card layout."""
    title = cell_text(ws.title)
    life_lines: list[str] = []
    spec_lines: list[str] = []
    history_lines: list[str] = []
    history_start: int | None = None

    for r in range(1, min(ws.max_row, 20) + 1):
        vals = row_values(ws, r)
        row_join = norm(" ".join(vals))
        if "SR NO" in row_join and "DATE" in row_join:
            history_start = r
            break
        if r <= 8 and "SR NO" not in row_join:
            line = format_row(vals)
            if not line:
                if vals[0] and vals[2]:
                    line = f"{vals[0]} {vals[2]}"
                elif vals[0] and vals[5]:
                    extra = f" {vals[6]}" if len(vals) > 6 and vals[6] else ""
                    line = f"{vals[0]} {vals[5]}{extra}"
            if line:
                if "FORMAT" in row_join or "RATING" in row_join:
                    spec_lines.append(line)
                elif any(k in row_join for k in ("M/C", "MODEL", "MAKE")):
                    life_lines.append(line)

    if history_start:
        header = " | ".join(v for v in row_values(ws, history_start) if v)
        if header:
            history_lines.append(header)
        for hr in range(history_start + 1, ws.max_row + 1):
            hline = format_row(row_values(ws, hr))
            if hline:
                history_lines.append(hline)

    if not life_lines and not history_lines:
        return None

    equipment_name = cell_text(ws.cell(5, 3).value) or title

    return {
        "sheetname": title,
        "equipment name": equipment_name,
        SECTION_LIFE: "\n".join(life_lines),
        SECTION_SPEC: "\n".join(spec_lines),
        SECTION_HISTORY: "\n".join(history_lines),
    }


def parse_sheet(ws) -> dict[str, str] | None:
    sections = find_section_rows(ws)
    if "life" not in sections or "spec" not in sections:
        return parse_alternate_sheet(ws)
    life_row = sections.get("life", 1)
    spec_row = sections.get("spec", life_row + 1)
    schedule_row = sections.get("schedule")
    history_row = sections.get("history")

    spec_end = schedule_row or history_row or (ws.max_row + 1)
    history_end = ws.max_row + 1

    equipment_name = extract_equipment_name(ws, life_row, spec_row)

    return {
        "sheetname": ws.title,
        "equipment name": equipment_name,
        SECTION_LIFE: extract_section(ws, life_row, spec_row),
        SECTION_SPEC: extract_section(ws, spec_row, spec_end),
        SECTION_HISTORY: extract_section(ws, history_row, history_end)
        if history_row
        else "",
    }


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths: dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE}")

    src = openpyxl.load_workbook(SOURCE, data_only=True)

    columns = [
        "sheetname",
        "equipment name",
        SECTION_LIFE,
        SECTION_SPEC,
        SECTION_HISTORY,
    ]

    rows: list[dict[str, str]] = []
    for name in src.sheetnames:
        if should_skip_sheet(name):
            continue
        parsed = parse_sheet(src[name])
        if parsed:
            rows.append(parsed)

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Extracted"
    ws.append(columns)
    for row in rows:
        ws.append([row[c] for c in columns])

    style_header(ws)
    ws.freeze_panes = "A2"
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    autosize(ws, {1: 28, 2: 32, 3: 48, 4: 48, 5: 64})

    out.save(OUTPUT)
    print(f"Source:  {SOURCE}")
    print(f"Output:  {OUTPUT}")
    print(f"Sheets:  {len(rows)} equipment sheets extracted (Index skipped)")


if __name__ == "__main__":
    main()

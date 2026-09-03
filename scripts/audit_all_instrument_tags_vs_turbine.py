#!/usr/bin/env python3
"""
Audit ALL tagged rows (not just yellow) of
  Plant Instrument Equipment List-21-08-2026.xlsx
against every card in
  Turbine & Instrument Equipment life histroy 20-06-2026 (2).xlsx

Same tag-only matching as extract_turbine_instrument_life_history.py
(exact tag, then punctuation-loose tag). No name-based fallback.

Output (DigiLog/backend/backlog-data/mill data/migration files-24-08-26/):
  all-instrument-tags-vs-turbine-life-audit-260826.xlsx
    - Summary
    - All tagged rows (Yes/No + Yellow flag + matched card)
    - Found (has data)
    - Not found (no data)
    - Cards not in hierarchy list
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from extract_mill_house_electrical import style_header_row, write_data_sheet
from extract_shn_life_history_table import (
    extract_life_fields as block_extract_life_fields,
    find_equipment_subcards,
)
from extract_turbine_instrument_life_history import (
    HIERARCHY_FILE,
    SOURCE_FILE,
    clean_name,
    is_yellow,
    loose_tag,
    norm_tag,
    should_skip_sheet,
)

ROOT = Path(__file__).resolve().parent.parent.parent
OUTPUT_DIR = (
    ROOT / "DigiLog" / "backend" / "backlog-data" / "mill data" / "migration files-24-08-26"
)
AUDIT_OUTPUT = OUTPUT_DIR / "all-instrument-tags-vs-turbine-life-audit-260826.xlsx"

YES_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill("solid", fgColor="FFC7CE")
YES_FONT = Font(bold=True, color="006100")
NO_FONT = Font(bold=True, color="9C0006")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")


def load_all_tagged_rows() -> list[dict]:
    wb = openpyxl.load_workbook(HIERARCHY_FILE, data_only=True)
    ws = wb.active

    rows: list[dict] = []
    for r in range(4, ws.max_row + 1):  # row 3 is the header row
        tag = clean_name(ws.cell(r, 8).value)
        sub = clean_name(ws.cell(r, 6).value)
        if not tag:
            continue
        yellow = any(is_yellow(ws.cell(r, c)) for c in range(1, 10))
        rows.append(
            {
                "sr": clean_name(ws.cell(r, 1).value),
                "raw_tag": tag,
                "plant": clean_name(ws.cell(r, 2).value),
                "section": clean_name(ws.cell(r, 3).value),
                "location": clean_name(ws.cell(r, 4).value),
                "main_equipment": clean_name(ws.cell(r, 5).value),
                "sub_equipment": sub,
                "department": clean_name(ws.cell(r, 7).value),
                "hist_location": clean_name(ws.cell(r, 9).value),
                "hierarchy_row": r,
                "yellow": yellow,
                "_key": f"{r}",
            }
        )
    wb.close()
    return rows


def group_by(rows: list[dict], keyfn) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for meta in rows:
        if meta["raw_tag"]:
            grouped[keyfn(meta["raw_tag"])].append(meta)
    return grouped


def main() -> None:
    print(f"Hierarchy: {HIERARCHY_FILE}")
    print(f"Data:      {SOURCE_FILE}")

    rows = load_all_tagged_rows()
    yellow_n = sum(1 for r in rows if r["yellow"])
    print(f"Tagged/named hierarchy rows: {len(rows)}  (yellow: {yellow_n})")

    by_tag = group_by(rows, norm_tag)
    by_loose = group_by(rows, loose_tag)

    src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    matched_by_key: dict[str, str] = {}
    used_keys: set[str] = set()
    extra_cards: list[list[str]] = []
    total_cards = 0

    for sheet_name in src.sheetnames:
        if should_skip_sheet(sheet_name):
            continue
        ws = src[sheet_name]
        blocks = find_equipment_subcards(ws)
        if not blocks:
            continue

        for card_index, block in enumerate(blocks, start=1):
            total_cards += 1
            fields = block_extract_life_fields(ws, block)
            equip_tag = clean_name(
                fields.get("EQUIPMENT TAG NAME/APPLICATION")
                or fields.get("EQUIPMENT TAG NO")
                or ""
            )
            equip_name = clean_name(fields.get("NAME OF EQUIPMENT") or "")
            card_label = f"{sheet_name} :: {equip_name or equip_tag or card_index}"

            if not equip_tag:
                extra_cards.append([sheet_name, "(blank)", equip_name, "No tag on card"])
                continue

            candidates = [m for m in by_tag.get(norm_tag(equip_tag), []) if m["_key"] not in used_keys]
            how = "exact-tag"
            if not candidates:
                candidates = [m for m in by_loose.get(loose_tag(equip_tag), []) if m["_key"] not in used_keys]
                how = "loose-tag"
            if not candidates:
                extra_cards.append([sheet_name, equip_tag, equip_name, "Tag not in hierarchy list"])
                continue

            meta = candidates[0]
            used_keys.add(meta["_key"])
            matched_by_key[meta["_key"]] = f"{card_label} [{how}]"

    src.close()

    found_rows = [r for r in rows if r["_key"] in matched_by_key]
    missing_rows = [r for r in rows if r["_key"] not in matched_by_key]

    print(f"\nTotal cards scanned in data file: {total_cards}")
    print(f"Hierarchy rows WITH data: {len(found_rows)} / {len(rows)}")
    print(f"  of which yellow: {sum(1 for r in found_rows if r['yellow'])} / {yellow_n}")
    print(f"Hierarchy rows WITHOUT data: {len(missing_rows)}")
    print(f"Cards with no matching hierarchy row: {len(extra_cards)}")

    out = openpyxl.Workbook()
    out.remove(out.active)

    write_data_sheet(
        out,
        "Summary",
        ["Issue type", "Count"],
        [
            ["Hierarchy source", HIERARCHY_FILE.name],
            ["Data source", SOURCE_FILE.name],
            ["Total tagged/named hierarchy rows", len(rows)],
            ["  ...of which yellow-highlighted", yellow_n],
            ["Hierarchy rows WITH data (any row)", len(found_rows)],
            ["  ...of which yellow-highlighted", sum(1 for r in found_rows if r["yellow"])],
            ["  ...of which NOT yellow", sum(1 for r in found_rows if not r["yellow"])],
            ["Hierarchy rows WITHOUT data", len(missing_rows)],
            ["  ...of which yellow-highlighted", sum(1 for r in missing_rows if r["yellow"])],
            ["Total cards scanned in data file", total_cards],
            ["Cards with no matching hierarchy row", len(extra_cards)],
        ],
        {"A": 44, "B": 60},
    )

    def row_out(meta: dict) -> list:
        sheet = matched_by_key.get(meta["_key"], "")
        return [
            meta["raw_tag"] or "(no tag)",
            "Yes" if meta["yellow"] else "No",
            meta["department"],
            meta["section"],
            meta["location"],
            meta["main_equipment"],
            meta["sub_equipment"],
            meta["hist_location"],
            "Yes" if sheet else "No",
            sheet,
        ]

    columns = [
        "History card Tag Nos.",
        "Yellow highlighted",
        "Department",
        "Section",
        "Location",
        "Main Equipment",
        "Sub Equipment",
        "History card Location",
        "Found in data",
        "Matched sheet / card",
    ]
    widths = {
        "A": 42, "B": 14, "C": 14, "D": 18, "E": 16,
        "F": 32, "G": 42, "H": 28, "I": 14, "J": 46,
    }

    all_rows_out = [row_out(m) for m in rows]
    ws_all = write_data_sheet(out, "All tagged rows", columns, all_rows_out, widths)
    for r in range(2, ws_all.max_row + 1):
        yellow_cell = ws_all.cell(r, 2)
        if yellow_cell.value == "Yes":
            yellow_cell.fill = YELLOW_FILL
        found_cell = ws_all.cell(r, 9)
        if found_cell.value == "Yes":
            found_cell.fill = YES_FILL
            found_cell.font = YES_FONT
        else:
            found_cell.fill = NO_FILL
            found_cell.font = NO_FONT

    write_data_sheet(
        out, "Found (has data)", columns, [row_out(m) for m in found_rows], widths
    )
    write_data_sheet(
        out, "Not found (no data)", columns, [row_out(m) for m in missing_rows], widths
    )
    write_data_sheet(
        out,
        "Cards not in hierarchy list",
        ["sheet name", "EQUIPMENT TAG NO", "NAME OF EQUIPMENT", "Issue"],
        extra_cards,
        {"A": 36, "B": 42, "C": 40, "D": 40},
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out.save(AUDIT_OUTPUT)
    print(f"\nAUDIT: {AUDIT_OUTPUT}")


if __name__ == "__main__":
    main()

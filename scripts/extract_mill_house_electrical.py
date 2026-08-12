#!/usr/bin/env python3
"""
Extract normalized equipment history for Electrical department (Sugar House).

Sources (each has its hierarchy in its own sheet or a separate file):
  1. Mill House:  hierarchy file + data file (separate)
  2. 70 TPH Boiler:  hierarchy in first sheet, data in remaining sheets (single file)
  3. Power House:  hierarchy in first sheet, data in remaining sheets (single file)
  4. DS House Motors:  hierarchy in first sheet (INDEX), data in remaining sheets
  5. DS House VFD Panels: hierarchy in first sheet (INDEX), data in remaining sheets

Output:  DigiLog/backend/backlog-data/mill data/mill-house-electrical-equipment-history.xlsx
  - Sheet Map
  - EQUIPMENT LIFE HISTORY CARD
  - EQUIPMENT SPECIFICATION (section / sub_section when present)
  - MAINTENANCE SCHEDULE (OEM; dynamic interval columns)
  - EQUIPMENT MAINTENANCE HISTORY

Also writes audit workbook:
  mill-house-electrical-extract-audit.xlsx
  - All hierarchy tags (Found in data = Yes/No for every tag)
  - Missing in data (hierarchy tags with no matching data sheet)
  - Duplicate tags (same tag already extracted from another/same source)
  - Not in hierarchy (data sheets whose tag is not in that source hierarchy)
  - Hierarchy duplicates (same tag listed more than once in hierarchy)
"""

from __future__ import annotations

import re
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from equipment_history_extract_lib import (
    cell_text,
    extract_history_rows_flexible,
    extract_schedule_rows,
    extract_specification_rows as extract_spec_rows_lib,
    is_history_header_row,
    norm,
    parse_schedule_layout,
    row_cells,
)

ROOT = Path(__file__).resolve().parent.parent.parent
BACKLOG_DIR = ROOT / "DigiLog" / "backend" / "backlog-data"
OUTPUT_DIR = BACKLOG_DIR / "mill data"
OUTPUT = OUTPUT_DIR / "mill-house-electrical-equipment-history.xlsx"
AUDIT_OUTPUT = OUTPUT_DIR / "mill-house-electrical-extract-audit.xlsx"

SOURCES = [
    {
        "label": "Mill House Electrical",
        "hierarchy_file": BACKLOG_DIR / "electrical -mill house-herarchy-30072026.xlsx",
        "data_file": ROOT / "electrical Mill house life history Data-30072026.xlsx",
        "skip_sheets": {"summary index", "link", "sheet1"},
    },
    {
        "label": "70 TPH Boiler Electrical",
        "hierarchy_file": BACKLOG_DIR / "2_70 TPH boiler life history.xlsx",
        "data_file": BACKLOG_DIR / "2_70 TPH boiler life history.xlsx",
        "skip_sheets": {"summary index", "summary link", "sheet1"},
    },
    {
        "label": "Power House Electrical",
        "hierarchy_file": BACKLOG_DIR / "3_Power House life history.xlsx",
        "data_file": BACKLOG_DIR / "3_Power House life history.xlsx",
        "skip_sheets": {"summary index", "summary link", "sheet1"},
    },
    {
        "label": "DS House Motors Electrical",
        "hierarchy_file": BACKLOG_DIR / "4_DS Electrical Equipment Life History - MOTORS.xlsx",
        "data_file": BACKLOG_DIR / "4_DS Electrical Equipment Life History - MOTORS.xlsx",
        "skip_sheets": {"index", "summary index", "summary link", "sheet1"},
    },
    {
        "label": "DS House VFD Panel Electrical",
        "hierarchy_file": BACKLOG_DIR / "5_DS Electrical Equipment Live  History - VFD PANEL.xlsx",
        "data_file": BACKLOG_DIR / "5_DS Electrical Equipment Live  History - VFD PANEL.xlsx",
        "skip_sheets": {"index", "summary index", "summary link", "sheet1"},
    },
]

SECTION_LIFE = "EQUIPMENT LIFE HISTORY CARD"
SECTION_SPEC = "EQUIPMENT SPECIFICATION"
SECTION_SCHEDULE_MARKERS = (
    "MAINTENANCE SCHEDULE",
    "OEM MAINTENANCE SCHEDULE",
    "GSM MAINTENANCE SCHEDULE",
)
SECTION_HISTORY = "EQUIPMENT MAINTENANCE HISTORY"

LIFE_COLUMNS = [
    "source file",
    "sheet id",
    "sheet name",
    "EQUIPMENT TAG NO",
    "NAME OF EQUIPMENT",
    "LOCATION",
    "DATE OF COMMISSIONING",
    "Main Equipment",
    "Sub Equipment",
    "Department",
]

SPEC_COLUMNS = [
    "source file",
    "sheet id",
    "sheet name",
    "EQUIPMENT TAG NO",
    "section",
    "sub_section",
    "Parameter label",
    "Parameter value",
]

HISTORY_COLUMNS = [
    "source file",
    "sheet id",
    "sheet name",
    "EQUIPMENT TAG NO",
    "Season / OFF Season",
    "Year",
    "Date of Start",
    "Date of Finish",
    "Outage/ Observation",
    "Action Taken",
    "Repair Cost (Rs.)",
    "Services (Internal / External)",
    "Responsibility ( Engineer/ Supervision)",
    "Remarks",
]

SCHEDULE_INTERVALS = [
    "Daily",
    "Weekly",
    "Monthly",
    "Quarterly",
    "Half - Yearly",
    "Yearly",
    "2 - Years",
    "3 - Years",
    "4 - Years",
]

SCHEDULE_COLUMNS = [
    "source file",
    "sheet id",
    "sheet name",
    "EQUIPMENT TAG NO",
    "Sr.No.",
    "Name of Equipment",
    "Maintenance / Inspection Activities",
    *SCHEDULE_INTERVALS,
    "Remarks",
]

SPEC_SECTION_HEADERS = (
    ("MECHANICAL PART", "mechanical"),
    ("INSTRUMENT PART", "instrument"),
    ("ELECTRICALT PART", "electrical"),
    ("ELECTRIC PART", "electrical"),
    ("ELECTRICAL PART", "electrical"),
    ("ELECTRONIC PART", "electrical"),
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def norm_tag(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def tags_match(hierarchy_tag: str, data_tag: str) -> bool:
    """Exact match after trim-all-spaces + lowercase. No fuzzy suffix matching."""
    ht = norm_tag(hierarchy_tag)
    dt = norm_tag(data_tag)
    return bool(ht and dt and ht == dt)


def load_hierarchy_tags(hierarchy_file: Path) -> tuple[dict[str, dict], list[dict]]:
    """Return (tag_map, hierarchy_duplicate_rows).

    tag_map keeps the first occurrence of each normalized tag.
    hierarchy_duplicate_rows lists later repeats within the same hierarchy sheet.
    """
    if not hierarchy_file.exists():
        raise FileNotFoundError(f"Hierarchy file not found: {hierarchy_file}")

    wb = openpyxl.load_workbook(hierarchy_file, data_only=True)
    ws = wb[wb.sheetnames[0]]
    tag_map: dict[str, dict] = {}
    duplicates: list[dict] = []

    tag_col = 8
    loc_col = 4
    hist_loc_col = 9

    for r in range(2, ws.max_row + 1):
        raw_tag = str(ws.cell(r, tag_col).value or "").strip()
        if not raw_tag:
            continue
        # Skip header / title rows (e.g. "History card Tag Nos.")
        if "/" not in raw_tag:
            continue
        meta = {
            "raw_tag": raw_tag,
            "plant": str(ws.cell(r, 2).value or "").strip(),
            "section": str(ws.cell(r, 3).value or "").strip(),
            "location": str(ws.cell(r, loc_col).value or "").strip(),
            "main_equipment": str(ws.cell(r, 5).value or "").strip(),
            "sub_equipment": str(ws.cell(r, 6).value or "").strip(),
            "department": str(ws.cell(r, 7).value or "").strip(),
            "hist_location": str(ws.cell(r, hist_loc_col).value or "").strip()
            if ws.max_column >= hist_loc_col
            else "",
            "hierarchy_row": r,
        }
        nt = norm_tag(raw_tag)
        if nt in tag_map:
            first = tag_map[nt]
            duplicates.append({
                "tag": raw_tag,
                "main_equipment": meta["main_equipment"],
                "sub_equipment": meta["sub_equipment"],
                "location": meta["location"],
                "department": meta["department"],
                "hierarchy_row": r,
                "first_row": first.get("hierarchy_row", ""),
                "first_sub_equipment": first.get("sub_equipment", ""),
            })
            continue
        tag_map[nt] = meta

    return tag_map, duplicates


def should_skip_sheet(name: str, skip_set: set[str]) -> bool:
    return norm(name.strip()).lower() in skip_set


def find_section_rows(ws) -> dict[str, tuple[int, int]]:
    found: dict[str, tuple[int, int]] = {}
    life_count = 0
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            text = norm(ws.cell(r, c).value)
            if not text:
                continue
            if SECTION_LIFE in text:
                life_count += 1
                if life_count == 1:
                    found["life"] = (r, c)
                elif life_count == 2 and "history" not in found:
                    found["history"] = (r, c)
                continue
            if SECTION_SPEC in text and "spec" not in found:
                found["spec"] = (r, c)
                continue
            if any(marker in text for marker in SECTION_SCHEDULE_MARKERS) and "schedule" not in found:
                found["schedule"] = (r, c)
                continue
            if SECTION_HISTORY in text and "history" not in found:
                found["history"] = (r, c)

    # Cardian / Rotary style: no "EQUIPMENT MAINTENANCE HISTORY" title — history
    # starts at a "Season / OFF Season" header row after the schedule block.
    if "history" not in found:
        start = 1
        if "schedule" in found:
            start = found["schedule"][0] + 1
        elif "spec" in found:
            start = found["spec"][0] + 1
        for r in range(start, ws.max_row + 1):
            if is_history_header_row(row_cells(ws, r)):
                found["history"] = (r, 1)
                break
    return found


def row_value_after_label(ws, row_idx: int, label_col: int) -> str:
    for c in range(label_col + 1, ws.max_column + 1):
        val = cell_text(ws.cell(row_idx, c).value)
        if val:
            return val
    return ""


def extract_life_fields(ws) -> dict[str, str]:
    sections = find_section_rows(ws)
    fields = {
        "NAME OF EQUIPMENT": ws.title.strip(),
        "LOCATION": "",
        "EQUIPMENT TAG NO": "",
        "DATE OF COMMISSIONING": "",
    }
    if "life" not in sections or "spec" not in sections:
        return fields

    life_row, label_col = sections["life"]
    spec_row, _ = sections["spec"]
    equip_no = ""

    for r in range(life_row + 1, spec_row):
        label = norm(ws.cell(r, label_col).value)
        if not label:
            continue
        value = row_value_after_label(ws, r, label_col)
        if "NAME OF EQUIPMENT" in label:
            fields["NAME OF EQUIPMENT"] = value or fields["NAME OF EQUIPMENT"]
        elif "LOCATION" in label:
            fields["LOCATION"] = value
        elif "EQUIPMENT TAG" in label or label.startswith("TAG NO") or label == "TAG NAME":
            fields["EQUIPMENT TAG NO"] = value
        elif "EQUIPMENT NO" in label and "TAG" not in label:
            equip_no = value
        elif "DATE OF COMMISSIONING" in label or "COMMISSIONING" in label:
            fields["DATE OF COMMISSIONING"] = value

    if not fields["EQUIPMENT TAG NO"] and equip_no:
        fields["EQUIPMENT TAG NO"] = equip_no
    if not fields["NAME OF EQUIPMENT"]:
        fields["NAME OF EQUIPMENT"] = ws.title.strip()
    return fields


def extract_specification_rows(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "spec" not in sections:
        return []

    spec_row, _ = sections["spec"]
    spec_end = sections.get("schedule", (ws.max_row + 1, 2))[0]
    if sections.get("history") and sections["history"][0] < spec_end:
        hist_before = sections["history"][0]
        if hist_before > spec_row:
            spec_end = min(spec_end, hist_before)

    return extract_spec_rows_lib(
        ws,
        spec_row,
        spec_end,
        sheet_id,
        sheet_name,
        (SECTION_HISTORY, SECTION_LIFE, *SECTION_SCHEDULE_MARKERS),
        section_headers=SPEC_SECTION_HEADERS,
    )


def extract_maintenance_schedule(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "schedule" not in sections:
        return []

    schedule_row, _ = sections["schedule"]
    schedule_end = sections.get("history", (ws.max_row + 1, 2))[0]
    layout = parse_schedule_layout(ws, schedule_row, schedule_end)
    if not layout:
        return []

    return extract_schedule_rows(
        ws,
        layout,
        schedule_end,
        sheet_id,
        sheet_name,
        tuple(SCHEDULE_INTERVALS),
        (SECTION_HISTORY, *SECTION_SCHEDULE_MARKERS),
    )


def extract_maintenance_history(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "history" not in sections:
        return []

    history_row, _ = sections["history"]
    return extract_history_rows_flexible(
        ws,
        history_row,
        sheet_id,
        sheet_name,
        stop_markers=(SECTION_LIFE, SECTION_SPEC),
    )


def insert_tag_column(rows: list[list[str]], equip_tag: str) -> list[list[str]]:
    return [row[:2] + [equip_tag] + row[2:] for row in rows]


def prepend_source_file(rows: list[list[str]], source_file: str) -> list[list[str]]:
    return [[source_file] + row for row in rows]


def style_header_row(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_data_sheet(out, title: str, columns: list[str], rows: list[list[str]], widths: dict[str, int]):
    if title in out.sheetnames:
        del out[title]
    ws = out.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append(row)
    style_header_row(ws)
    ws.freeze_panes = "A2"
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def save_workbook_to(out: openpyxl.Workbook, path: Path) -> Path:
    try:
        out.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {path.name}. Close it in Excel and retry."
        ) from exc
    return path


def save_workbook(out: openpyxl.Workbook) -> Path:
    """Always write the single canonical output path (no -updated / timestamp copies)."""
    return save_workbook_to(out, OUTPUT)


def process_source(
    source: dict,
    all_mappings: list,
    all_life: list,
    all_spec: list,
    all_schedule: list,
    all_history: list,
    seen_tags: dict[str, dict],
    audit: dict[str, list],
    cross_hierarchy_tags: dict[str, dict],
):
    label = source["label"]
    hierarchy_file = source["hierarchy_file"]
    data_file = source["data_file"]
    skip_sheets = source["skip_sheets"]

    print(f"\n{'='*60}")
    print(f"Processing: {label}")
    print(f"  Hierarchy: {hierarchy_file.name}")
    print(f"  Data:      {data_file.name}")

    if not data_file.exists():
        print(f"  SKIPPED — data file not found: {data_file}")
        return

    hierarchy_tags, hier_dups = load_hierarchy_tags(hierarchy_file)
    print(f"  Hierarchy tags: {len(hierarchy_tags)}")

    hierarchy_file_name = hierarchy_file.name
    source_file_name = data_file.name

    for dup in hier_dups:
        audit["hierarchy_duplicates"].append([
            source_file_name,
            hierarchy_file_name,
            dup["tag"],
            dup["main_equipment"],
            dup["sub_equipment"],
            dup["location"],
            dup["department"],
            dup["hierarchy_row"],
            (
                f"Duplicate within hierarchy sheet "
                f"(first at row {dup['first_row']}: {dup['first_sub_equipment']})"
            ),
        ])

    for nt, meta in hierarchy_tags.items():
        if nt in cross_hierarchy_tags:
            first = cross_hierarchy_tags[nt]
            audit["hierarchy_duplicates"].append([
                source_file_name,
                hierarchy_file_name,
                meta["raw_tag"],
                meta["main_equipment"],
                meta["sub_equipment"],
                meta["location"],
                meta["department"],
                meta.get("hierarchy_row", ""),
                (
                    f"Same tag already in hierarchy of {first['hierarchy_file']} "
                    f"({first['sub_equipment']})"
                ),
            ])
        else:
            cross_hierarchy_tags[nt] = {
                "hierarchy_file": hierarchy_file_name,
                "source_file": source_file_name,
                "raw_tag": meta["raw_tag"],
                "sub_equipment": meta["sub_equipment"],
            }

    src = openpyxl.load_workbook(data_file, data_only=True)

    matched_tags: set[str] = set()
    matched_sheet_by_tag: dict[str, str] = {}
    skipped_sheets: list[str] = []
    unmatched_count = 0
    duplicate_count = 0

    for sheet_name in src.sheetnames:
        if should_skip_sheet(sheet_name, skip_sheets):
            skipped_sheets.append(sheet_name)
            continue

        ws = src[sheet_name]
        fields = extract_life_fields(ws)
        equip_tag = fields["EQUIPMENT TAG NO"]

        nt = norm_tag(equip_tag)
        meta = hierarchy_tags.get(nt)
        if meta is None:
            for h_nt, h_meta in hierarchy_tags.items():
                if tags_match(h_meta["raw_tag"], equip_tag):
                    nt = h_nt
                    meta = h_meta
                    break
        if meta is None:
            unmatched_count += 1
            audit["not_in_hierarchy"].append([
                source_file_name,
                hierarchy_file_name,
                sheet_name,
                equip_tag or "(blank)",
                fields.get("NAME OF EQUIPMENT", ""),
                "Data sheet tag not found in this source hierarchy",
            ])
            continue

        if nt in seen_tags:
            duplicate_count += 1
            first = seen_tags[nt]
            audit["duplicates"].append([
                source_file_name,
                hierarchy_file_name,
                sheet_name,
                equip_tag,
                meta["main_equipment"],
                meta["sub_equipment"],
                meta["location"],
                (
                    f"Duplicate tag — already extracted from "
                    f"{first['source_file']} / sheet '{first['sheet_name']}'"
                ),
            ])
            continue

        matched_tags.add(nt)
        matched_sheet_by_tag[nt] = sheet_name
        seen_tags[nt] = {
            "source_file": source_file_name,
            "sheet_name": sheet_name,
            "tag": equip_tag,
        }

        sheet_id = uuid.uuid4().hex[:12]
        all_mappings.append((source_file_name, sheet_name, sheet_id, equip_tag))

        all_life.append([
            source_file_name,
            sheet_id,
            sheet_name,
            equip_tag,
            fields["NAME OF EQUIPMENT"],
            meta["location"],
            fields["DATE OF COMMISSIONING"],
            meta["main_equipment"],
            meta["sub_equipment"],
            meta["department"],
        ])
        all_spec.extend(prepend_source_file(
            insert_tag_column(
                extract_specification_rows(ws, sheet_id, sheet_name),
                equip_tag,
            ),
            source_file_name,
        ))
        all_schedule.extend(prepend_source_file(
            insert_tag_column(
                extract_maintenance_schedule(ws, sheet_id, sheet_name),
                equip_tag,
            ),
            source_file_name,
        ))
        all_history.extend(prepend_source_file(
            insert_tag_column(
                extract_maintenance_history(ws, sheet_id, sheet_name),
                equip_tag,
            ),
            source_file_name,
        ))

    missing_nts = sorted(set(hierarchy_tags.keys()) - matched_tags)
    # Full tag checklist (found + missing) for this source.
    for nt_all, m in sorted(hierarchy_tags.items(), key=lambda item: item[1]["raw_tag"]):
        found = nt_all in matched_tags
        audit["all_tags"].append([
            source_file_name,
            hierarchy_file_name,
            m["raw_tag"],
            m["main_equipment"],
            m["sub_equipment"],
            m["location"],
            m["department"],
            m.get("hist_location", ""),
            "Yes" if found else "No",
            matched_sheet_by_tag.get(nt_all, ""),
        ])

    for nt_m in missing_nts:
        m = hierarchy_tags[nt_m]
        if nt_m in seen_tags and nt_m not in matched_tags:
            # Present in another source's extract already — still missing from this data file.
            first = seen_tags[nt_m]
            issue = (
                f"Hierarchy tag has no data sheet in this file "
                f"(tag already extracted from {first['source_file']})"
            )
        else:
            issue = "Hierarchy tag has no matching data sheet"
        audit["missing_in_data"].append([
            source_file_name,
            hierarchy_file_name,
            m["raw_tag"],
            m["main_equipment"],
            m["sub_equipment"],
            m["location"],
            m["department"],
            m.get("hist_location", ""),
            issue,
        ])

    print(f"  Matched: {len(matched_tags)} | Not in data: {len(missing_nts)}")
    if missing_nts:
        print(f"  Tags from hierarchy NOT found in data ({len(missing_nts)}):")
        for nt_m in missing_nts[:10]:
            m = hierarchy_tags[nt_m]
            print(f"    - {m['raw_tag']}  ({m['sub_equipment']})")
    if unmatched_count:
        print(f"  Data sheets NOT in hierarchy: {unmatched_count}")
    if duplicate_count:
        print(f"  Duplicate tags skipped: {duplicate_count}")
    print(f"  Skipped sheets: {', '.join(skipped_sheets)}")


def write_audit_workbook(audit: dict[str, list]) -> Path:
    out = openpyxl.Workbook()
    out.remove(out.active)

    found_yes = sum(1 for row in audit["all_tags"] if row[8] == "Yes")
    found_no = sum(1 for row in audit["all_tags"] if row[8] == "No")
    summary_rows = [
        ["All hierarchy tags", len(audit["all_tags"])],
        ["Found in data (Yes)", found_yes],
        ["Not found in data (No)", found_no],
        ["Missing in data", len(audit["missing_in_data"])],
        ["Duplicate tags", len(audit["duplicates"])],
        ["Not in hierarchy", len(audit["not_in_hierarchy"])],
        ["Hierarchy duplicates", len(audit["hierarchy_duplicates"])],
    ]
    write_data_sheet(
        out,
        "Summary",
        ["Issue type", "Count"],
        summary_rows,
        {"A": 28, "B": 10},
    )
    write_data_sheet(
        out,
        "All hierarchy tags",
        [
            "source file",
            "hierarchy file",
            "EQUIPMENT TAG NO",
            "Main Equipment",
            "Sub Equipment",
            "Location",
            "Department",
            "History card Location",
            "Found in data",
            "Matched sheet name",
        ],
        audit["all_tags"],
        {
            "A": 52, "B": 52, "C": 34, "D": 28, "E": 40,
            "F": 20, "G": 14, "H": 36, "I": 14, "J": 36,
        },
    )
    write_data_sheet(
        out,
        "Missing in data",
        [
            "source file",
            "hierarchy file",
            "EQUIPMENT TAG NO",
            "Main Equipment",
            "Sub Equipment",
            "Location",
            "Department",
            "History card Location",
            "Issue",
        ],
        audit["missing_in_data"],
        {
            "A": 52, "B": 52, "C": 34, "D": 28, "E": 40,
            "F": 20, "G": 14, "H": 36, "I": 55,
        },
    )
    write_data_sheet(
        out,
        "Duplicate tags",
        [
            "source file",
            "hierarchy file",
            "sheet name",
            "EQUIPMENT TAG NO",
            "Main Equipment",
            "Sub Equipment",
            "Location",
            "Issue",
        ],
        audit["duplicates"],
        {
            "A": 52, "B": 52, "C": 36, "D": 34,
            "E": 28, "F": 40, "G": 20, "H": 60,
        },
    )
    write_data_sheet(
        out,
        "Not in hierarchy",
        [
            "source file",
            "hierarchy file",
            "sheet name",
            "EQUIPMENT TAG NO",
            "NAME OF EQUIPMENT",
            "Issue",
        ],
        audit["not_in_hierarchy"],
        {"A": 52, "B": 52, "C": 36, "D": 34, "E": 40, "F": 50},
    )
    write_data_sheet(
        out,
        "Hierarchy duplicates",
        [
            "source file",
            "hierarchy file",
            "EQUIPMENT TAG NO",
            "Main Equipment",
            "Sub Equipment",
            "Location",
            "Department",
            "Hierarchy row",
            "Issue",
        ],
        audit["hierarchy_duplicates"],
        {
            "A": 52, "B": 52, "C": 34, "D": 28, "E": 40,
            "F": 20, "G": 14, "H": 14, "I": 60,
        },
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return save_workbook_to(out, AUDIT_OUTPUT)


def main() -> None:
    all_mappings: list[tuple[str, str, str, str]] = []
    all_life: list[list[str]] = []
    all_spec: list[list[str]] = []
    all_schedule: list[list[str]] = []
    all_history: list[list[str]] = []
    seen_tags: dict[str, dict] = {}
    cross_hierarchy_tags: dict[str, dict] = {}
    audit: dict[str, list] = {
        "all_tags": [],
        "missing_in_data": [],
        "duplicates": [],
        "not_in_hierarchy": [],
        "hierarchy_duplicates": [],
    }

    for source in SOURCES:
        process_source(
            source,
            all_mappings,
            all_life,
            all_spec,
            all_schedule,
            all_history,
            seen_tags,
            audit,
            cross_hierarchy_tags,
        )

    # --- Write combined output workbook ---
    out = openpyxl.Workbook()
    out.remove(out.active)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["source file", "sheet name", "id", "EQUIPMENT TAG NO"])
    for source_file_name, sheet_name, sheet_id, equip_tag in all_mappings:
        map_ws.append([source_file_name, sheet_name, sheet_id, equip_tag])
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    map_ws.column_dimensions["A"].width = 52
    map_ws.column_dimensions["B"].width = 44
    map_ws.column_dimensions["C"].width = 16
    map_ws.column_dimensions["D"].width = 30

    write_data_sheet(
        out, SECTION_LIFE, LIFE_COLUMNS, all_life,
        {
            "A": 52, "B": 16, "C": 44, "D": 30, "E": 34, "F": 20,
            "G": 22, "H": 30, "I": 40, "J": 14,
        },
    )
    write_data_sheet(
        out, SECTION_SPEC, SPEC_COLUMNS, all_spec,
        {"A": 52, "B": 16, "C": 44, "D": 30, "E": 14, "F": 18, "G": 28, "H": 36},
    )
    write_data_sheet(
        out, "MAINTENANCE SCHEDULE", SCHEDULE_COLUMNS, all_schedule,
        {
            "A": 52, "B": 16, "C": 44, "D": 30, "E": 8, "F": 22, "G": 42,
            "H": 8, "I": 8, "J": 8, "K": 10, "L": 12, "M": 8,
            "N": 10, "O": 10, "P": 10, "Q": 16,
        },
    )
    write_data_sheet(
        out, SECTION_HISTORY, HISTORY_COLUMNS, all_history,
        {
            "A": 52, "B": 16, "C": 44, "D": 30, "E": 18, "F": 14, "G": 16,
            "H": 16, "I": 36, "J": 36, "K": 16, "L": 22, "M": 28, "N": 24,
        },
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    saved = save_workbook(out)
    audit_path = write_audit_workbook(audit)

    print(f"\n{'='*60}")
    print(f"COMBINED OUTPUT: {saved}")
    print(f"AUDIT OUTPUT:    {audit_path}")
    print(
        f"\nTotal: Sheet Map ({len(all_mappings)}), "
        f"{SECTION_LIFE} ({len(all_life)}), "
        f"{SECTION_SPEC} ({len(all_spec)}), "
        f"MAINTENANCE SCHEDULE ({len(all_schedule)}), "
        f"{SECTION_HISTORY} ({len(all_history)})"
    )
    print(
        f"Audit: All tags ({len(audit['all_tags'])}), "
        f"Yes ({sum(1 for r in audit['all_tags'] if r[8] == 'Yes')}), "
        f"No ({sum(1 for r in audit['all_tags'] if r[8] == 'No')}), "
        f"Duplicates ({len(audit['duplicates'])}), "
        f"Not in hierarchy ({len(audit['not_in_hierarchy'])}), "
        f"Hierarchy duplicates ({len(audit['hierarchy_duplicates'])})"
    )


if __name__ == "__main__":
    main()

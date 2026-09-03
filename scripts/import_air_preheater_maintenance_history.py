#!/usr/bin/env python3
"""
Extract maintenance history from Air Pre Heater Excel and write a PPN feed JSON.

Input (default):  ../../folder/air preheater data.xlsx
Output (default):  ../backend/scripts/data_feed_power_history/feed-data/air-pre-heater-mechanical-history.json

Usage:
  python import_air_preheater_maintenance_history.py
  python import_air_preheater_maintenance_history.py --input "path/to/file.xlsx"
  python import_air_preheater_maintenance_history.py --sub-section "General"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_INPUT = ROOT / "folder" / "air preheater data.xlsx"
DEFAULT_OUTPUT = (
    Path(__file__).resolve().parent.parent
    / "backend"
    / "scripts"
    / "data_feed_power_history"
    / "feed-data"
    / "air-pre-heater-mechanical-history.json"
)

SECTION_HISTORY = "EQUIPMENT MAINTENANCE HISTORY"


def norm(text) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip()).upper()


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_season(raw: str) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    u = norm(s)
    if "OFF" in u and "SEASON" in u:
        return "Off-Season"
    if u == "SEASON":
        return "Season"
    return s


def row_cells(ws, row_idx: int) -> list[str]:
    return [cell_text(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]


def find_history_section_row(ws) -> int | None:
    for r in range(1, ws.max_row + 1):
        joined = norm(" ".join(row_cells(ws, r)))
        if SECTION_HISTORY in joined:
            return r
    return None


def is_history_header_row(row: list[str]) -> bool:
    joined = norm(" ".join(row))
    return "SEASON" in joined and ("YEAR" in joined or "OUTAGE" in joined or "OBSERVATION" in joined)


def map_history_columns(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(header_row, c).value)
        if not h:
            continue
        if "SEASON" in h and "OFF" in h:
            mapping["season"] = c
        elif h.startswith("YEAR"):
            mapping["year"] = c
        elif "DATE OF START" in h:
            mapping["date_start"] = c
        elif "DATE OF FINISH" in h:
            mapping["date_finish"] = c
        elif "OUTAGE" in h or "OBSERVATION" in h:
            mapping["obs"] = c
        elif "ACTION TAKEN" in h:
            mapping["act"] = c
        elif "REPAIR COST" in h:
            mapping["cost"] = c
        elif "SERVICE" in h:
            mapping["svc"] = c
        elif "RESPONSIBILITY" in h:
            mapping["resp"] = c
        elif "REMARKS" in h:
            mapping["rem"] = c
    return mapping


def extract_history_from_blr_sheet(ws) -> list[dict]:
    history_row = find_history_section_row(ws)
    if not history_row:
        return []

    header_row = None
    for r in range(history_row + 1, min(history_row + 8, ws.max_row + 1)):
        if is_history_header_row(row_cells(ws, r)):
            header_row = r
            break
    if header_row is None:
        return []

    cols = map_history_columns(ws, header_row)
    records: list[dict] = []

    for r in range(header_row + 1, ws.max_row + 1):
        row = row_cells(ws, r)
        if not any(row):
            continue
        if is_history_header_row(row):
            continue
        joined = norm(" ".join(row))
        if SECTION_HISTORY in joined or "EQUIPMENT LIFE HISTORY" in joined:
            break

        def val(key: str) -> str:
            col = cols.get(key)
            return cell_text(ws.cell(r, col).value) if col else ""

        season = normalize_season(val("season"))
        year = val("year")
        date_start = val("date_start")
        date_finish = val("date_finish")
        obs = val("obs")
        act = val("act")

        if not any([season, year, date_start, date_finish, obs, act]):
            continue
        if norm(season) in {"SEASON / OFF SEASON", "SR.NO.", "SR NO"}:
            continue

        if not year and date_start and len(date_start) >= 4:
            year = date_start[:4]

        records.append({
            "season": season or None,
            "year": year or None,
            "date_start": date_start or None,
            "date_finish": date_finish or None,
            "obs": obs or None,
            "act": act or None,
            "cost": val("cost") or None,
            "svc": val("svc") or None,
            "resp": val("resp") or None,
            "rem": val("rem") or None,
        })

    return records


def extract_history_flat_sheet(ws) -> list[dict]:
    """Sheet row 1 = headers (export table format)."""
    header = row_cells(ws, 1)
    if not is_history_header_row(header):
        return []

    cols = map_history_columns(ws, 1)
    records: list[dict] = []

    for r in range(2, ws.max_row + 1):
        if not any(row_cells(ws, r)):
            continue

        def val(key: str) -> str:
            col = cols.get(key)
            return cell_text(ws.cell(r, col).value) if col else ""

        season = normalize_season(val("season"))
        year = val("year")
        date_start = val("date_start")
        date_finish = val("date_finish")
        obs = val("obs")
        act = val("act")

        if not any([season, year, date_start, date_finish, obs, act]):
            continue

        if not year and date_start and len(date_start) >= 4:
            year = date_start[:4]

        records.append({
            "season": season or None,
            "year": year or None,
            "date_start": date_start or None,
            "date_finish": date_finish or None,
            "obs": obs or None,
            "act": act or None,
            "cost": val("cost") or None,
            "svc": val("svc") or None,
            "resp": val("resp") or None,
            "rem": val("rem") or None,
        })

    return records


def extract_all_history(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    all_records: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        blr_rows = extract_history_from_blr_sheet(ws)
        flat_rows = extract_history_flat_sheet(ws) if not blr_rows else []
        rows = blr_rows or flat_rows
        all_records.extend(rows)

    return all_records


def build_feed(history: list[dict], sub_section: str) -> dict:
    scoped = []
    for row in history:
        scoped.append({
            **row,
            "section": "mechanical",
            "sub_section": sub_section,
            "equipment_refs": [{"section": "mechanical", "sub_section": sub_section}],
        })

    return {
        "equipment": [{
            "hierarchy_name": "Air Preheater",
            "hierarchy_card": "Air Pre Heater",
            "hierarchy_path": "Power Plant > 150TPH BLR > Auxiliary Equipment > Air Pre Heater",
            "name": "Air Preheater",
            "category": "150TPH BLR",
            "subcategory": "Auxiliary Equipment",
            "history_only": True,
            "history": scoped,
        }]
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract Air Pre Heater maintenance history to PPN feed JSON")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Source .xlsx path")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output feed JSON path")
    parser.add_argument("--sub-section", default="General", help="Mechanical equipment card name in specs")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Input file not found: {args.input}", file=sys.stderr)
        print("Place the workbook at: folder/air preheater data.xlsx", file=sys.stderr)
        return 1

    history = extract_all_history(args.input)
    if not history:
        print(f"No maintenance history rows found in {args.input}", file=sys.stderr)
        return 1

    feed = build_feed(history, args.sub_section)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(feed, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Extracted {len(history)} history row(s) from {args.input.name}")
    print(f"Wrote {args.output}")
    print("\nImport into database:")
    print("  cd DigiLog/backend")
    print("  npm run db:import-aph-history -- --replace-history")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""
Extract boiling-house mechanical life history cards.

Filter tags from:
  DigiLog/backend/backlog-data/BOILING HOUSE MECHANICAL  EQUIPMENT LIST -30-07-2026.xlsx
  - Department = MECHANICAL, or INSTRUMENT + MECHANICAL
  - If a cell has both instrument and mechanical tags, keep ONLY the mechanical tag
    (prefer ZIL/GSM/... over ZIL/SUG...)

Data source:
  HISTORY SHEET 28.07.2026.xlsx

Output (extraction only):
  DigiLog/backend/backlog-data/boiling-house-mechanical-equipment-history.xlsx
  DigiLog/backend/backlog-data/boiling-house-mechanical-extract-audit.xlsx
"""

from __future__ import annotations

import re
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from equipment_history_extract_lib import cell_text, norm
from extract_mill_house_electrical import (
    HISTORY_COLUMNS as ELEC_HISTORY_COLUMNS,
    LIFE_COLUMNS as ELEC_LIFE_COLUMNS,
    SCHEDULE_COLUMNS as ELEC_SCHEDULE_COLUMNS,
    SPEC_COLUMNS as ELEC_SPEC_COLUMNS,
    extract_maintenance_history,
    extract_maintenance_schedule,
    extract_specification_rows,
    insert_tag_column,
    norm_tag,
    prepend_source_file,
    style_header_row,
    write_data_sheet,
)

ROOT = Path(__file__).resolve().parent.parent.parent
BACKLOG = ROOT / "DigiLog" / "backend" / "backlog-data"
HIERARCHY_FILE = BACKLOG / "BOILING HOUSE MECHANICAL  EQUIPMENT LIST -30-07-2026.xlsx"
DATA_FILE = ROOT / "HISTORY SHEET 28.07.2026.xlsx"
OUTPUT_DIR = BACKLOG
OUTPUT = OUTPUT_DIR / "boiling-house-mechanical-equipment-history.xlsx"
AUDIT_OUTPUT = OUTPUT_DIR / "boiling-house-mechanical-extract-audit.xlsx"

SKIP_SHEETS = {"index", "summary index", "summary link", "sheet1"}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Reuse electrical column layouts (include source file).
LIFE_COLUMNS = list(ELEC_LIFE_COLUMNS)
SPEC_COLUMNS = list(ELEC_SPEC_COLUMNS)
SCHEDULE_COLUMNS = list(ELEC_SCHEDULE_COLUMNS)
HISTORY_COLUMNS = list(ELEC_HISTORY_COLUMNS)


def split_tags(raw: str) -> list[str]:
    parts: list[str] = []
    for line in re.split(r"[\n\r;]+", str(raw or "")):
        t = line.strip()
        if t and "/" in t:
            parts.append(t)
    return parts


def is_mechanical_dept(dept: str) -> bool:
    return "MECHANICAL" in str(dept or "").upper()


def is_instrument_style_tag(tag: str) -> bool:
    """Instrument companion tags in dual cells (e.g. ZIL/SUG./014 ...)."""
    t = norm_tag(tag)
    return t.startswith("zil/sug") or "/sug./" in t or t.startswith("zil/sug.")


def is_mechanical_style_tag(tag: str) -> bool:
    """Mechanical history-card tags (e.g. ZIL/GSM/RR/JH/06)."""
    t = norm_tag(tag)
    return t.startswith("zil/gsm/") or "/gsm/" in t


def pick_mechanical_tag(tags: list[str]) -> str | None:
    """From one hierarchy cell, keep only the mechanical tag.

    Dual-department cells often store:
      ZIL/SUG./...   (instrument)
      ZIL/GSM/...    (mechanical)
    """
    cleaned = [t.strip() for t in tags if t and "/" in t]
    if not cleaned:
        return None
    if len(cleaned) == 1:
        return cleaned[0]

    mech = [t for t in cleaned if is_mechanical_style_tag(t)]
    if mech:
        return mech[0]

    non_inst = [t for t in cleaned if not is_instrument_style_tag(t)]
    if non_inst:
        return non_inst[0]

    # Last resort: last tag in the cell (historically the GSM line).
    return cleaned[-1]


def load_mechanical_hierarchy_tags(path: Path) -> dict[str, dict]:
    """Return norm_tag -> meta for Mechanical / Instrument+Mechanical rows."""
    if not path.exists():
        raise FileNotFoundError(f"Hierarchy file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    tag_map: dict[str, dict] = {}

    for r in range(4, ws.max_row + 1):
        dept = str(ws.cell(r, 7).value or "").strip()
        if not is_mechanical_dept(dept):
            continue

        raw_tags = split_tags(str(ws.cell(r, 8).value or ""))
        mech_tag = pick_mechanical_tag(raw_tags)
        if not mech_tag:
            continue

        nt = norm_tag(mech_tag)
        meta = {
            "raw_tag": mech_tag,
            "all_tags_in_cell": raw_tags,
            "plant": str(ws.cell(r, 2).value or "").strip(),
            "section": str(ws.cell(r, 3).value or "").strip(),
            "location": str(ws.cell(r, 4).value or "").strip(),
            "main_equipment": str(ws.cell(r, 5).value or "").strip(),
            "sub_equipment": str(ws.cell(r, 6).value or "").strip(),
            "department": " / ".join(
                p.strip() for p in re.split(r"[\n\r]+", dept) if p.strip()
            ),
            "hist_location": str(ws.cell(r, 9).value or "").strip(),
            "hierarchy_row": r,
        }
        # Keep first occurrence.
        tag_map.setdefault(nt, meta)

    wb.close()
    return tag_map


def should_skip_sheet(name: str) -> bool:
    return norm(name.strip()).lower() in SKIP_SHEETS


def row_value_after_label(ws, row_idx: int, label_col: int) -> str:
    for c in range(label_col + 1, ws.max_column + 1):
        val = cell_text(ws.cell(row_idx, c).value)
        if val:
            return val
    return ""


def extract_life_fields(ws) -> dict[str, str]:
    """Read life-card fields from HISTORY SHEET layouts.

    Some cards use:
      - full banner ``EQUIPMENT LIFE HISTORY CARD``
      - short banner ``EQUIPMENT LIFE HISTORY``
      - no banner (labels start at row 1)
    Tags are often under ``EQUIPMENT NO:`` in column E (merged).
    """
    fields = {
        "NAME OF EQUIPMENT": ws.title.strip(),
        "LOCATION": "",
        "EQUIPMENT TAG NO": "",
        "DATE OF COMMISSIONING": "",
    }

    life_row = None
    life_col = 1
    spec_row = ws.max_row + 1

    for r in range(1, min(ws.max_row, 40) + 1):
        for c in (1, 2, 3, 4, 5):
            text = norm(ws.cell(r, c).value)
            if not text:
                continue
            if life_row is None and "EQUIPMENT LIFE HISTORY" in text:
                life_row = r
                life_col = c
            if "EQUIPMENT SPECIFICATION" in text:
                spec_row = r
                break
        if spec_row <= ws.max_row:
            # keep scanning a little for life banner above spec only
            pass

    scan_start = (life_row + 1) if life_row is not None else 1
    scan_end = spec_row if spec_row <= ws.max_row else min(ws.max_row, 20)
    equip_no = ""

    for r in range(scan_start, scan_end):
        # Labels may sit in col A even when life banner was found elsewhere.
        label = ""
        label_col = 1
        for c in range(1, min(6, ws.max_column + 1)):
            raw = norm(ws.cell(r, c).value)
            if raw:
                label = raw
                label_col = c
                break
        if not label:
            continue
        value = row_value_after_label(ws, r, label_col)
        if "NAME OF EQUIPMENT" in label:
            fields["NAME OF EQUIPMENT"] = value or fields["NAME OF EQUIPMENT"]
        elif "LOCATION" in label:
            fields["LOCATION"] = value
        elif "EQUIPMENT TAG" in label or label.startswith("TAG NO") or label == "TAG NAME":
            fields["EQUIPMENT TAG NO"] = value
        elif "EQUIPMENT NO" in label and "TAG" not in label:
            equip_no = value
        elif "DATE OF COMMISSIONING" in label or "COMMISSIONING" in label:
            fields["DATE OF COMMISSIONING"] = value

    if not fields["EQUIPMENT TAG NO"] and equip_no:
        fields["EQUIPMENT TAG NO"] = equip_no
    if not fields["NAME OF EQUIPMENT"]:
        fields["NAME OF EQUIPMENT"] = ws.title.strip()
    return fields


def save_workbook_to(out: openpyxl.Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        out.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {path.name}. Close it in Excel and retry."
        ) from exc
    return path


def write_audit(hierarchy_tags: dict[str, dict], matched: dict[str, str]) -> Path:
    rows = []
    for nt, meta in sorted(hierarchy_tags.items(), key=lambda item: item[1]["raw_tag"]):
        found = nt in matched
        rows.append([
            DATA_FILE.name,
            HIERARCHY_FILE.name,
            meta["raw_tag"],
            " | ".join(meta.get("all_tags_in_cell") or []),
            meta["department"],
            meta["section"],
            meta["main_equipment"],
            meta["sub_equipment"],
            meta["location"],
            "Yes" if found else "No",
            matched.get(nt, ""),
        ])

    yes_n = sum(1 for r in rows if r[9] == "Yes")
    no_n = len(rows) - yes_n
    out = openpyxl.Workbook()
    out.remove(out.active)
    write_data_sheet(
        out,
        "Summary",
        ["Issue type", "Count"],
        [
            ["Mechanical tags (GSM-only for dual cells)", len(rows)],
            ["Found in HISTORY SHEET (Yes)", yes_n],
            ["Not found (No)", no_n],
        ],
        {"A": 48, "B": 10},
    )
    write_data_sheet(
        out,
        "All hierarchy tags",
        [
            "source file",
            "hierarchy file",
            "Mechanical tag used",
            "All tags in cell",
            "Department",
            "Section",
            "Main Equipment",
            "Sub Equipment",
            "Location",
            "Found in data",
            "Matched sheet name",
        ],
        rows,
        {
            "A": 40, "B": 52, "C": 30, "D": 44, "E": 28,
            "F": 18, "G": 28, "H": 36, "I": 22, "J": 14, "K": 28,
        },
    )
    return save_workbook_to(out, AUDIT_OUTPUT)


def main() -> None:
    print(f"Hierarchy: {HIERARCHY_FILE.name}")
    print(f"Data:      {DATA_FILE.name}")

    hierarchy_tags = load_mechanical_hierarchy_tags(HIERARCHY_FILE)
    print(f"Mechanical tags (after dual-tag filter): {len(hierarchy_tags)}")

    dual = sum(1 for m in hierarchy_tags.values() if len(m.get("all_tags_in_cell") or []) > 1)
    print(f"  of which from dual-tag cells: {dual}")

    if not DATA_FILE.exists():
        raise FileNotFoundError(f"Data file not found: {DATA_FILE}")

    src = openpyxl.load_workbook(DATA_FILE, data_only=True)
    source_name = DATA_FILE.name

    all_mappings: list[tuple[str, str, str, str]] = []
    all_life: list[list[str]] = []
    all_spec: list[list[str]] = []
    all_schedule: list[list[str]] = []
    all_history: list[list[str]] = []
    matched: dict[str, str] = {}
    unmatched_sheets: list[str] = []

    for sheet_name in src.sheetnames:
        if should_skip_sheet(sheet_name):
            continue

        ws = src[sheet_name]
        fields = extract_life_fields(ws)
        equip_tag = fields["EQUIPMENT TAG NO"]
        nt = norm_tag(equip_tag)
        meta = hierarchy_tags.get(nt)
        if meta is None:
            unmatched_sheets.append(f"{sheet_name} (tag={equip_tag})")
            continue
        if nt in matched:
            continue

        matched[nt] = sheet_name
        sheet_id = uuid.uuid4().hex[:12]
        # Prefer hierarchy mechanical tag spelling.
        out_tag = meta["raw_tag"]

        all_mappings.append((source_name, sheet_name, sheet_id, out_tag))
        all_life.append([
            source_name,
            sheet_id,
            sheet_name,
            out_tag,
            fields["NAME OF EQUIPMENT"] or meta["sub_equipment"],
            meta["location"] or fields["LOCATION"],
            fields["DATE OF COMMISSIONING"],
            meta["main_equipment"],
            meta["sub_equipment"],
            meta["department"],
        ])
        all_spec.extend(prepend_source_file(
            insert_tag_column(
                extract_specification_rows(ws, sheet_id, sheet_name),
                out_tag,
            ),
            source_name,
        ))
        all_schedule.extend(prepend_source_file(
            insert_tag_column(
                extract_maintenance_schedule(ws, sheet_id, sheet_name),
                out_tag,
            ),
            source_name,
        ))
        all_history.extend(prepend_source_file(
            insert_tag_column(
                extract_maintenance_history(ws, sheet_id, sheet_name),
                out_tag,
            ),
            source_name,
        ))

    src.close()

    missing = sorted(set(hierarchy_tags) - set(matched))
    print(f"Matched: {len(matched)} | Missing in HISTORY SHEET: {len(missing)}")
    if missing:
        print("Missing tags:")
        for nt in missing[:20]:
            m = hierarchy_tags[nt]
            print(f"  - {m['raw_tag']}  ({m['sub_equipment']})")

    out = openpyxl.Workbook()
    out.remove(out.active)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["source file", "sheet name", "id", "EQUIPMENT TAG NO"])
    for source_file_name, sheet_name, sheet_id, equip_tag in all_mappings:
        map_ws.append([source_file_name, sheet_name, sheet_id, equip_tag])
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    map_ws.column_dimensions["A"].width = 40
    map_ws.column_dimensions["B"].width = 36
    map_ws.column_dimensions["C"].width = 16
    map_ws.column_dimensions["D"].width = 30

    write_data_sheet(
        out, "EQUIPMENT LIFE HISTORY CARD", LIFE_COLUMNS, all_life,
        {
            "A": 40, "B": 16, "C": 36, "D": 30, "E": 34, "F": 22,
            "G": 20, "H": 28, "I": 36, "J": 22,
        },
    )
    write_data_sheet(
        out, "EQUIPMENT SPECIFICATION", SPEC_COLUMNS, all_spec,
        {"A": 40, "B": 16, "C": 36, "D": 30, "E": 14, "F": 18, "G": 28, "H": 36},
    )
    write_data_sheet(
        out, "MAINTENANCE SCHEDULE", SCHEDULE_COLUMNS, all_schedule,
        {
            "A": 40, "B": 16, "C": 36, "D": 30, "E": 8, "F": 22, "G": 42,
            "H": 8, "I": 8, "J": 8, "K": 10, "L": 12, "M": 8,
            "N": 10, "O": 10, "P": 10, "Q": 16,
        },
    )
    write_data_sheet(
        out, "EQUIPMENT MAINTENANCE HISTORY", HISTORY_COLUMNS, all_history,
        {
            "A": 40, "B": 16, "C": 36, "D": 30, "E": 18, "F": 14, "G": 16,
            "H": 16, "I": 36, "J": 36, "K": 16, "L": 22, "M": 28, "N": 24,
        },
    )

    saved = save_workbook_to(out, OUTPUT)
    audit_path = write_audit(hierarchy_tags, matched)

    print(f"\nCOMBINED OUTPUT: {saved}")
    print(f"AUDIT OUTPUT:    {audit_path}")
    print(
        f"Total: Sheet Map ({len(all_mappings)}), "
        f"LIFE ({len(all_life)}), SPEC ({len(all_spec)}), "
        f"SCHEDULE ({len(all_schedule)}), HISTORY ({len(all_history)})"
    )


if __name__ == "__main__":
    main()

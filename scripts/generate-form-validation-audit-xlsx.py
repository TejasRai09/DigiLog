#!/usr/bin/env python3
"""Generate form-validation-audit.xlsx from GSMA frontend validation review."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "form-validation-audit.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, min_width=10, max_width=48):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = min_width
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = length


def write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize_columns(ws)
    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb,
        "Summary",
        ["Severity", "Count", "Description"],
        [
            ["Wrong (UI ≠ JS)", 3, "Fields marked required in UI but not enforced in validate()"],
            ["Weak (metadata only)", 17, "Only date and/or shift checked; measurements can be empty"],
            ["Reasonable", 4, "Required fields match what the UI implies"],
        ],
    )

    write_sheet(
        wb,
        "All Forms",
        [
            "#",
            "Form",
            "Category",
            "File",
            "JS Validates",
            "Can Submit Empty Data",
            "Status",
            "Notes",
        ],
        [
            [1, "Near Miss / Incident", "EHS", "ehs/EhsNearMiss.jsx", "Date, Name, Severity", "Yes", "OK for * fields", "Description, location optional"],
            [2, "Accident Register", "EHS", "ehs/EhsAccident.jsx", "Date, Injured Person, Type", "Yes", "OK for * fields", "Description optional"],
            [3, "CPU Water Recycle", "EHS", "ehs/EhsWaterCpu.jsx", "Date only", "Yes", "Weak", "All readings optional"],
            [4, "ETP Working", "EHS", "ehs/EhsWaterEtp.jsx", "Date only", "Yes", "Weak", "All readings optional"],
            [5, "Ground Water Abstraction", "EHS", "ehs/EhsWaterGwa.jsx", "Date only", "Yes", "Weak", "All readings optional"],
            [6, "Power Details", "Power", "power/PhPower.jsx", "Date only", "Yes", "Wrong", "Time marked required but not validated"],
            [7, "Steam Details", "Power", "power/PhSteam.jsx", "Date only", "Yes", "Wrong", "Time marked required but not validated"],
            [8, "Power Stoppage", "Power", "power/PhStoppage.jsx", "Date, times, Section, Sub-Section, Machinery, Category, Remark", "No", "Good", "Most complete stoppage form"],
            [9, "Mill Stoppages", "Mill", "mill/MillStoppages.jsx", "Date, From, To", "Yes", "Wrong", "Section, Machinery, Remark marked * but not in JS"],
            [10, "Equipment Temperature", "Mill", "mill/EquipmentTemp.jsx", "Date, Shift", "Yes", "Weak", "All temps optional"],
            [11, "Lube Pressure & Roller Temp", "Mill", "mill/LubePressure.jsx", "Date, Shift", "Yes", "Weak", "All readings optional"],
            [12, "Shredder & OTG", "Mill", "mill/ShreddarOTG.jsx", "Date, Shift", "Yes", "Weak", "All readings optional"],
            [13, "RS Logbook", "Lab", "lab/RSLogbook.jsx", "Date, Shift", "Yes", "Weak", "Pol/Brix/IU/pH optional"],
            [14, "SA Logbook", "Lab", "lab/SALogbook.jsx", "Date, Shift", "Yes", "Weak", "Retention/moisture/color optional"],
            [15, "DS Logbook", "Lab", "lab/DSLogbook.jsx", "Date, Shift", "Yes", "Weak", "Pol/Brix optional"],
            [16, "Ops Logbook", "Lab", "lab/OpsLogbook.jsx", "Date, Shift", "Yes", "Weak", "Crush/FBD/hopper temps optional"],
            [17, "Syrup Logbook", "Lab", "lab/SyrupLogbook.jsx", "Date, Shift", "Yes", "Weak", "Production/diversion optional"],
            [18, "Stoppage Logbook", "Lab", "lab/StoppageLogbook.jsx", "Date, From, To", "Yes", "Weak", "Reason/remark optional"],
            [19, "Clarification Log", "Production", "production/ProdClarification.jsx", "Date + ≥1 hourly row", "No", "Good", "Requires at least one row"],
            [20, "Decanter Log", "Production", "production/ProdDecanter.jsx", "Date + ≥1 hourly row", "No", "Good", "Requires at least one row"],
            [21, "Pan Log Book", "Production", "production/ProdPanLogbook.jsx", "Date only", "Yes", "Weak", "All strike data optional"],
            [22, "Shift Chemist Log", "Production", "production/ProdShiftChemist.jsx", "Date only", "Yes", "Weak", "All shift jobs optional"],
            [23, "Centrifugal Stoppage", "Production", "production/ProdCentrifugal.jsx", "Date only", "Yes", "Weak", "Machine stoppage rows optional"],
            [24, "Distillery Operations", "Distillery", "distillery/DistilleryOperations.jsx", "Date only", "Yes", "Weak", "Operation mode and quantities optional"],
        ],
    )

    write_sheet(
        wb,
        "Wrong Validation",
        ["#", "Form", "File", "UI Marked Required", "JS Actually Validates", "Missing in JS"],
        [
            [1, "Mill Stoppages", "mill/MillStoppages.jsx", "Date, From, To, Section, Machinery, Remark", "Date, From, To", "Section, Machinery, Remark"],
            [2, "Power Details", "power/PhPower.jsx", "Report Date + time", "Date only", "Time"],
            [3, "Steam Details", "power/PhSteam.jsx", "Report Date + time", "Date only", "Time"],
        ],
    )

    write_sheet(
        wb,
        "Missing Types",
        ["Validation Type", "Forms Affected", "Count", "Example from Field Labels"],
        [
            ["At least one measurement/row required", "Most logbooks & ops forms", 17, "Pol, Brix, Motor Temp, Steam Generation, Meter Reading"],
            ["Date/time ordering (End > Start)", "Mill Stoppages, Lab Stoppage, Power Stoppage", 3, "From / To"],
            ["Regulatory range (limits in labels)", "CPU Water, ETP Working", "2-3", "pH (5.5–8.5), TSS <30 ppm, COD <250 ppm"],
            ["Numeric min/max / reasonableness", "Lab, Mill, Power, Production logbooks", "~18", "Pol 0–100, pH 0–14, non-negative pressure/temp"],
            ["Conditional required", "Near Miss", "1-2", "Person Type = Other → Specify Other"],
            ["Format validation (phone etc.)", "Near Miss", 1, "Contact No."],
            ["Cross-field / business logic", "Distillery, Centrifugal, Pan Log, Shift Chemist, GWA", "5+", "Operation Mode, From/To/Duration, meter vs extracted totals"],
            ["HTML required not enforced", "All 24 GSMA forms", 24, "preventDefault() runs before browser validation"],
        ],
    )

    write_sheet(
        wb,
        "EHS Water Limits",
        ["Form", "Field Label", "Implied Rule", "Enforced"],
        [
            ["CPU Water", "pH (5.5–8.5)", "5.5 ≤ pH ≤ 8.5", "No"],
            ["CPU Water", "TSS <30 ppm", "Max 30", "No"],
            ["CPU Water", "COD <250 ppm", "Max 250", "No"],
            ["CPU Water", "BOD <30 ppm", "Max 30", "No"],
            ["CPU Water", "TDS <2100 mg/L", "Max 2100", "No"],
            ["CPU Water", "Oil & Grease <10", "Max 10", "No"],
            ["CPU Water", "Transmittance >85", "Min 85", "No"],
            ["ETP Working", "pH / TSS / COD / BOD / TDS", "Similar limits in labels", "No"],
            ["GWA", "Meter vs Extracted (KL)", "Logical consistency", "No"],
        ],
    )

    write_sheet(
        wb,
        "Incident Gaps",
        ["Form", "Field", "Expected Rule", "Enforced"],
        [
            ["Near Miss", "Person Type = Other", "Specify Other required", "No"],
            ["Near Miss", "Contact No.", "Phone format", "No"],
            ["Near Miss", "Description / Cause", "Required for harm incidents", "No"],
            ["Accident", "Description", "Should accompany Major/Fatal", "No"],
        ],
    )

    write_sheet(
        wb,
        "Fix Priority",
        ["Priority", "Forms", "Fix"],
        [
            ["P1", "Mill Stoppages, PhPower, PhSteam", "Align JS validate() with * labels"],
            ["P2", "Mill Stoppages, Lab Stoppage, Power Stoppage", "Add End > Start check"],
            ["P3", "CPU Water, ETP Working", "Enforce limits shown in labels"],
            ["P4", "17 weak logbooks", "Require at least one filled measurement row"],
            ["P5", "Lab/Mill numeric forms", "Add Pol/Brix/pH/temp/pressure ranges"],
        ],
    )

    write_sheet(
        wb,
        "Equipment Forms",
        ["Form", "Required in UI", "Validated on Save", "Notes"],
        [
            ["Equipment Life History Card", "Name only (HTML)", "Name only", "Tag No., Location, Commissioning optional"],
            ["OEM Maintenance Schedule", "Component names, steps", "Partial toast check", "Light validation"],
            ["Login / Admin Login", "Email, password", "Yes (loginValidation.js)", "Good"],
        ],
    )

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()

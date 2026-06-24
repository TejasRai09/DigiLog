#!/usr/bin/env python3
"""
Extract normalized equipment history from File for Electrical.xlsx.

Output: electrical-equipment-history.xlsx
  - Sheet Map
  - EQUIPMENT LIFE HISTORY CARD
  - EQUIPMENT SPECIFICATION (section / sub_section when present)
  - MAINTENANCE SCHEDULE (OEM or standard; dynamic interval columns)
  - EQUIPMENT MAINTENANCE HISTORY
"""

from __future__ import annotations

import re
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from equipment_history_extract_lib import (
    cell_text,
    extract_history_rows_flexible,
    extract_schedule_rows,
    extract_specification_rows as extract_spec_rows_lib,
    norm,
    parse_schedule_layout,
    row_cells,
)

ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE = ROOT / "File for Electrical.xlsx"
OUTPUT = ROOT / "electrical-equipment-history.xlsx"

SKIP_SHEETS = {"index file", "sheet1"}

# Unfilled template default copied across many col-A layout sheets.
NAME_BOILERPLATE = "30.85 MW GENERATOR SET"

SECTION_LIFE = "EQUIPMENT LIFE HISTORY CARD"
SECTION_SPEC = "EQUIPMENT SPECIFICATION"
SECTION_SCHEDULE_MARKERS = ("MAINTENANCE SCHEDULE", "OEM MAINTENANCE SCHEDULE")
SECTION_HISTORY = "EQUIPMENT MAINTENANCE HISTORY"

SPEC_COLUMNS = [
    "sheet id",
    "sheet name",
    "section",
    "sub_section",
    "Parameter label",
    "Parameter value",
]

LIFE_COLUMNS = [
    "sheet id",
    "sheet name",
    "NAME OF EQUIPMENT",
    "LOCATION",
    "EQUIPMENT TAG NAME/APPLICATION",
    "DATE OF COMMISSIONING",
]

HISTORY_COLUMNS = [
    "sheet id",
    "sheet name",
    "Season / OFF Season",
    "Year",
    "Date of Start",
    "Date of Finish",
    "Outage/ Observation",
    "Action Taken",
    "Repair Cost (Rs.)",
    "Services (Internal / External)",
    "Responsibility ( Engineer/ Supervision)",
    "Remarks",
]

SCHEDULE_INTERVALS = [
    "Daily",
    "Weekly",
    "Monthly",
    "Quarterly",
    "Half - Yearly",
    "Yearly",
    "2 - Years",
    "3 - Years",
    "4 - Years",
]

SCHEDULE_COLUMNS = [
    "sheet id",
    "sheet name",
    "Sr.No.",
    "Name of Equipment",
    "Maintenance / Inspection Activities",
    *SCHEDULE_INTERVALS,
    "Remarks",
]

SPEC_SECTION_HEADERS = (
    ("MECHANICAL PART", "mechanical"),
    ("INSTRUMENT PART", "instrument"),
    ("ELECTRICALT PART", "electrical"),
    ("ELECTRIC PART", "electrical"),
    ("ELECTRICAL PART", "electrical"),
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def should_skip_sheet(name: str) -> bool:
    return norm(name) in {norm(s) for s in SKIP_SHEETS}


def find_section_rows(ws) -> dict[str, tuple[int, int]]:
    """Return section key -> (row, label_column). Scans columns A and B."""
    found: dict[str, tuple[int, int]] = {}
    life_count = 0
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            text = norm(ws.cell(r, c).value)
            if not text:
                continue
            if SECTION_LIFE in text:
                life_count += 1
                if life_count == 1:
                    found["life"] = (r, c)
                elif life_count == 2 and "history" not in found:
                    found["history"] = (r, c)
                continue
            if SECTION_SPEC in text and "spec" not in found:
                found["spec"] = (r, c)
                continue
            if any(marker in text for marker in SECTION_SCHEDULE_MARKERS) and "schedule" not in found:
                found["schedule"] = (r, c)
                continue
            if SECTION_HISTORY in text and "history" not in found:
                found["history"] = (r, c)
    return found


def row_value_after_label(ws, row_idx: int, label_col: int) -> str:
    for c in range(label_col + 1, ws.max_column + 1):
        val = cell_text(ws.cell(row_idx, c).value)
        if val:
            return val
    return ""


def resolve_equipment_name(extracted: str, sheet_title: str) -> str:
    if not extracted.strip() or norm(extracted) == norm(NAME_BOILERPLATE):
        return sheet_title
    return extracted


def extract_specification_rows(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "spec" not in sections:
        return []

    spec_row, _ = sections["spec"]
    spec_end = sections.get("schedule", (ws.max_row + 1, 2))[0]
    if sections.get("history") and sections["history"][0] < spec_end:
        hist_before = sections["history"][0]
        if hist_before > spec_row:
            spec_end = min(spec_end, hist_before)

    return extract_spec_rows_lib(
        ws,
        spec_row,
        spec_end,
        sheet_id,
        sheet_name,
        (SECTION_HISTORY, SECTION_LIFE, *SECTION_SCHEDULE_MARKERS),
        section_headers=SPEC_SECTION_HEADERS,
    )


def extract_life_fields(ws) -> dict[str, str]:
    sections = find_section_rows(ws)
    fields = {
        "NAME OF EQUIPMENT": ws.title,
        "LOCATION": "",
        "EQUIPMENT TAG NAME/APPLICATION": "",
        "DATE OF COMMISSIONING": "",
    }
    if "life" not in sections or "spec" not in sections:
        return fields

    life_row, label_col = sections["life"]
    spec_row, _ = sections["spec"]
    equip_no = ""

    for r in range(life_row + 1, spec_row):
        label = norm(ws.cell(r, label_col).value)
        if not label:
            continue
        value = row_value_after_label(ws, r, label_col)
        if "NAME OF EQUIPMENT" in label:
            fields["NAME OF EQUIPMENT"] = value or fields["NAME OF EQUIPMENT"]
        elif "LOCATION" in label:
            fields["LOCATION"] = value
        elif "EQUIPMENT TAG" in label or label.startswith("TAG NO") or label == "TAG NAME":
            fields["EQUIPMENT TAG NAME/APPLICATION"] = value
        elif "EQUIPMENT NO" in label and "TAG" not in label:
            equip_no = value
        elif "DATE OF COMMISSIONING" in label or "COMMISSIONING" in label:
            fields["DATE OF COMMISSIONING"] = value
        elif label.startswith("DRIVE") and "DATE" not in label:
            pass

    if not fields["EQUIPMENT TAG NAME/APPLICATION"] and equip_no:
        fields["EQUIPMENT TAG NAME/APPLICATION"] = equip_no
    fields["NAME OF EQUIPMENT"] = resolve_equipment_name(fields["NAME OF EQUIPMENT"], ws.title)
    return fields


ELECTRICAL_SCHEDULE_INTERVALS = tuple(SCHEDULE_INTERVALS)


def extract_maintenance_schedule(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "schedule" not in sections:
        return []

    schedule_row, _ = sections["schedule"]
    schedule_end = sections.get("history", (ws.max_row + 1, 2))[0]
    layout = parse_schedule_layout(ws, schedule_row, schedule_end)
    if not layout:
        return []

    return extract_schedule_rows(
        ws,
        layout,
        schedule_end,
        sheet_id,
        sheet_name,
        ELECTRICAL_SCHEDULE_INTERVALS,
        (SECTION_HISTORY, *SECTION_SCHEDULE_MARKERS),
    )


def extract_maintenance_history(ws, sheet_id: str, sheet_name: str) -> list[list[str]]:
    sections = find_section_rows(ws)
    if "history" not in sections:
        return []

    history_row, _ = sections["history"]
    return extract_history_rows_flexible(
        ws,
        history_row,
        sheet_id,
        sheet_name,
        stop_markers=(SECTION_LIFE, SECTION_SPEC),
    )


def style_header_row(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_data_sheet(out, title: str, columns: list[str], rows: list[list[str]], widths: dict[str, int]):
    if title in out.sheetnames:
        del out[title]
    ws = out.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append(row)
    style_header_row(ws)
    ws.freeze_panes = "A2"
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def load_existing_ids() -> dict[str, str]:
    candidates = (
        OUTPUT,
        OUTPUT.with_name(f"{OUTPUT.stem}-updated{OUTPUT.suffix}"),
    )
    for path in candidates:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Sheet Map" not in wb.sheetnames:
            continue
        ws = wb["Sheet Map"]
        ids = {
            cell_text(ws.cell(r, 1).value): cell_text(ws.cell(r, 2).value)
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value
        }
        if ids:
            return ids
    return {}


def save_workbook(out: openpyxl.Workbook) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    candidates = (
        OUTPUT,
        OUTPUT.with_name(f"{OUTPUT.stem}-updated{OUTPUT.suffix}"),
        OUTPUT.with_name(f"{OUTPUT.stem}-{stamp}{OUTPUT.suffix}"),
    )
    for path in candidates:
        try:
            out.save(path)
            return path
        except PermissionError:
            continue
    raise PermissionError(f"Could not write output. Close Excel and retry. Tried: {candidates[-1]}")


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE}")

    src = openpyxl.load_workbook(SOURCE, data_only=True)

    existing_ids = load_existing_ids()
    mappings: list[tuple[str, str]] = []
    life_rows: list[list[str]] = []
    spec_rows: list[list[str]] = []
    schedule_rows: list[list[str]] = []
    history_rows: list[list[str]] = []

    for sheet_name in src.sheetnames:
        if should_skip_sheet(sheet_name):
            continue
        sheet_id = existing_ids.get(sheet_name) or uuid.uuid4().hex[:12]
        mappings.append((sheet_name, sheet_id))
        ws = src[sheet_name]

        fields = extract_life_fields(ws)
        life_rows.append([
            sheet_id,
            sheet_name,
            fields["NAME OF EQUIPMENT"],
            fields["LOCATION"],
            fields["EQUIPMENT TAG NAME/APPLICATION"],
            fields["DATE OF COMMISSIONING"],
        ])
        spec_rows.extend(extract_specification_rows(ws, sheet_id, sheet_name))
        schedule_rows.extend(extract_maintenance_schedule(ws, sheet_id, sheet_name))
        history_rows.extend(extract_maintenance_history(ws, sheet_id, sheet_name))

    out = openpyxl.Workbook()
    out.remove(out.active)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["sheet name", "id"])
    for sheet_name, sheet_id in mappings:
        map_ws.append([sheet_name, sheet_id])
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    map_ws.column_dimensions["A"].width = 36
    map_ws.column_dimensions["B"].width = 16

    write_data_sheet(
        out, SECTION_LIFE, LIFE_COLUMNS, life_rows,
        {"A": 16, "B": 36, "C": 34, "D": 28, "E": 34, "F": 22},
    )
    write_data_sheet(
        out, SECTION_SPEC, SPEC_COLUMNS, spec_rows,
        {"A": 16, "B": 36, "C": 14, "D": 18, "E": 28, "F": 36},
    )
    write_data_sheet(
        out, "MAINTENANCE SCHEDULE", SCHEDULE_COLUMNS, schedule_rows,
        {
            "A": 16, "B": 36, "C": 8, "D": 22, "E": 42,
            "F": 8, "G": 8, "H": 8, "I": 10, "J": 12, "K": 8, "L": 10, "M": 10, "N": 10, "O": 16,
        },
    )
    write_data_sheet(
        out, SECTION_HISTORY, HISTORY_COLUMNS, history_rows,
        {"A": 16, "B": 36, "C": 18, "D": 14, "E": 16, "F": 16, "G": 36, "H": 36, "I": 16, "J": 22, "K": 28, "L": 24},
    )

    saved = save_workbook(out)
    print(f"Source: {SOURCE}")
    print(f"Output: {saved}")
    print(
        f"Sheets: Sheet Map ({len(mappings)}), "
        f"{SECTION_LIFE} ({len(life_rows)}), "
        f"{SECTION_SPEC} ({len(spec_rows)}), "
        f"MAINTENANCE SCHEDULE ({len(schedule_rows)}), "
        f"{SECTION_HISTORY} ({len(history_rows)})"
    )


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Extract Sugar Plant Instrument/Turbine life-history cards.

Hierarchy (yellow-highlighted rows only):
  Extraction_files-sugar/instrument and mechanical/
    Plant Instrument Equipment List-21-08-2026.xlsx

Data (multi-card sheets, several equipment blocks per tab):
  Extraction_files-sugar/instrument and mechanical/
    Turbine & Instrument Equipment life histroy 20-06-2026 (2).xlsx

Matching (hierarchy row <-> card), same approach as the other 24-08-26
migration extracts:
  1. EQUIPMENT NO / TAG vs "History card Tag Nos." (exact, then loose)
  2. If several hierarchy rows share that tag, or the tag doesn't match,
     fall back to matching Sub Equipment / Main Equipment name against the
     card's NAME OF EQUIPMENT / LOCATION / sheet name (one card -> one leaf).

Output (DigiLog/backend/backlog-data/mill data/migration files-24-08-26/):
  yellow-instrument-turbine-equipment-history-260826.xlsx
    - Hierarchy (first sheet, yellow rows only)
    - Sheet Map
    - EQUIPMENT LIFE HISTORY CARD / SPECIFICATION / SCHEDULE / HISTORY
  yellow-instrument-turbine-extract-audit-260826.xlsx
"""

from __future__ import annotations

import re
import shutil
import uuid
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from extract_mill_house_electrical import (
    HISTORY_COLUMNS,
    LIFE_COLUMNS,
    SCHEDULE_COLUMNS,
    SPEC_COLUMNS,
    style_header_row,
    write_data_sheet,
)
from extract_mill_house_mechanical import equipment_name_key, names_match
from extract_shn_life_history_table import (
    extract_life_fields as block_extract_life_fields,
    extract_maintenance_history as block_extract_history,
    extract_maintenance_schedule as block_extract_schedule,
    extract_specification_rows as block_extract_specs,
    find_equipment_subcards,
    insert_tag_column,
)

ROOT = Path(__file__).resolve().parent.parent.parent
SRC_DIR = ROOT / "Extraction_files-sugar" / "instrument and mechanical"
HIERARCHY_FILE = SRC_DIR / "Plant Instrument Equipment List-21-08-2026.xlsx"
SOURCE_FILE = SRC_DIR / "Turbine & Instrument Equipment life histroy 20-06-2026 (2).xlsx"

OUTPUT_DIR = (
    ROOT / "DigiLog" / "backend" / "backlog-data" / "mill data" / "migration files-24-08-26"
)
OUTPUT = OUTPUT_DIR / "yellow-instrument-turbine-equipment-history-260826.xlsx"
AUDIT_OUTPUT = OUTPUT_DIR / "yellow-instrument-turbine-extract-audit-260826.xlsx"
SOURCE_COPY = OUTPUT_DIR / "Turbine & Instrument Equipment life histroy 20-06-2026 (2).xlsx"

SKIP_SHEETS = {"index", "summary index", "summary link", "sheet1", "hierarchy"}
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill("solid", fgColor="FFC7CE")
YES_FONT = Font(bold=True, color="006100")
NO_FONT = Font(bold=True, color="9C0006")

HIERARCHY_COLUMNS = [
    "Sr.No.",
    "Plant",
    "Section",
    "Location",
    "Main Equipment",
    " Sub Equipment",
    "Department",
    "History card Tag Nos.",
    "History card Location",
]


def norm_tag(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def loose_tag(value: str) -> str:
    """Collapse punctuation/space differences but keep the '/' path shape."""
    s = norm_tag(value)
    s = re.sub(r"[^a-z0-9/]+", "", s)
    return s


def tag_suffix(value: str) -> str:
    parts = [p for p in loose_tag(value).split("/") if p]
    if len(parts) >= 2:
        return "/".join(parts[-2:])
    return parts[-1] if parts else ""


def clean_name(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def should_skip_sheet(name: str) -> bool:
    return clean_name(name).lower() in SKIP_SHEETS


def is_yellow(cell) -> bool:
    fill = cell.fill
    if not fill or fill.patternType in (None, "none"):
        return False
    fg = fill.fgColor
    if fg is None or getattr(fg, "type", None) != "rgb" or not fg.rgb:
        return False
    return str(fg.rgb).upper().endswith("FFFF00")


def load_hierarchy_rows() -> list[dict]:
    """Yellow-highlighted rows only, from the Plant Instrument Equipment List."""
    if not HIERARCHY_FILE.exists():
        raise FileNotFoundError(HIERARCHY_FILE)
    wb = openpyxl.load_workbook(HIERARCHY_FILE, data_only=True)
    ws = wb.active

    rows: list[dict] = []
    for r in range(4, ws.max_row + 1):  # row 3 is the header row
        yellow = any(is_yellow(ws.cell(r, c)) for c in range(1, 10))
        if not yellow:
            continue
        tag = clean_name(ws.cell(r, 8).value)
        sub = clean_name(ws.cell(r, 6).value)
        if not tag and not sub:
            continue
        rows.append(
            {
                "sr": clean_name(ws.cell(r, 1).value),
                "raw_tag": tag,
                "plant": clean_name(ws.cell(r, 2).value),
                "section": clean_name(ws.cell(r, 3).value),
                "location": clean_name(ws.cell(r, 4).value),
                "main_equipment": clean_name(ws.cell(r, 5).value),
                "sub_equipment": sub,
                "department": clean_name(ws.cell(r, 7).value) or "INSTRUMENT",
                "hist_location": clean_name(ws.cell(r, 9).value),
                "hierarchy_row": r,
                "_key": f"{r}:{norm_tag(tag)}:{equipment_name_key(sub)}",
            }
        )
    wb.close()
    return rows


def soft_names_match(a: str, b: str) -> bool:
    if names_match(a, b):
        return True
    ka = equipment_name_key(a)
    kb = equipment_name_key(b)
    if not ka or not kb:
        return False
    if ka in kb or kb in ka:
        return True
    noise = {
        "motor", "pump", "for", "the", "and", "m", "c", "no", "fld", "inst",
        "field", "instrument", "valve", "control",
    }
    ta = {t for t in ka.split() if len(t) > 2 and t not in noise}
    tb = {t for t in kb.split() if len(t) > 2 and t not in noise}
    if not ta or not tb:
        return False
    overlap = ta & tb
    return len(overlap) >= max(2, min(len(ta), len(tb)) - 1)


def group_by_tag(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for meta in rows:
        if meta["raw_tag"]:
            grouped[norm_tag(meta["raw_tag"])].append(meta)
    return grouped


def group_by_loose_tag(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for meta in rows:
        if meta["raw_tag"]:
            grouped[loose_tag(meta["raw_tag"])].append(meta)
    return grouped


def pick_by_name(candidates: list[dict], equip_name: str, sheet_name: str, location: str) -> dict | None:
    if not candidates:
        return None

    name_hits = [
        m for m in candidates
        if soft_names_match(equip_name, m["sub_equipment"])
        or soft_names_match(sheet_name, m["sub_equipment"])
        or soft_names_match(location, m["sub_equipment"])
        or soft_names_match(location, m["hist_location"])
        or soft_names_match(equip_name, m["main_equipment"])
        or soft_names_match(sheet_name, m["main_equipment"])
    ]
    if not name_hits:
        return None
    if len(name_hits) == 1:
        return name_hits[0]

    for source in (location, equip_name, sheet_name):
        m = re.search(r"(?:no[.\s-]*)?(\d+)\s*$", equipment_name_key(source))
        if not m:
            continue
        needle = m.group(1)
        numbered = [
            h for h in name_hits
            if re.search(
                rf"(?:^|[^0-9]){re.escape(needle)}(?:[^0-9]|$)",
                equipment_name_key(h["sub_equipment"]),
            )
            or tag_suffix(h["raw_tag"]).endswith(needle)
        ]
        if len(numbered) == 1:
            return numbered[0]
        if numbered:
            return numbered[0]
    return name_hits[0]


def resolve_hierarchy_meta(
    hierarchy_rows: list[dict],
    by_tag: dict[str, list[dict]],
    by_loose: dict[str, list[dict]],
    equip_tag: str,
    equip_name: str,
    sheet_name: str,
    location: str,
    used_keys: set[str],
) -> tuple[dict | None, str]:
    """
    Match order: tag only.
      1) exact tag (space/case-insensitive)
      2) loose tag (punctuation differences, e.g. "17 BCV 020B01" vs "17.BCV.020.B01")

    No name-based fallback: every yellow hierarchy row here has a unique tag,
    and the cards' "NAME OF EQUIPMENT" is a generic term (Control valve, Flow
    meter, ...) that cannot reliably disambiguate equipment — only the tag can.
    Sub Equipment / Main Equipment name is still used to break a tie when a
    tag legitimately maps to more than one hierarchy row.
    """
    exact = [m for m in by_tag.get(norm_tag(equip_tag), []) if m["_key"] not in used_keys]
    if exact:
        if len(exact) == 1:
            return exact[0], "exact-tag"
        picked = pick_by_name(exact, equip_name, sheet_name, location)
        return (picked, "exact-tag") if picked else (exact[0], "exact-tag")

    loose = [m for m in by_loose.get(loose_tag(equip_tag), []) if m["_key"] not in used_keys]
    if loose:
        if len(loose) == 1:
            return loose[0], "loose-tag"
        picked = pick_by_name(loose, equip_name, sheet_name, location)
        return (picked, "loose-tag") if picked else (loose[0], "loose-tag")

    return None, "unmatched"


def prepend_source_file(rows: list[list[str]], source_file: str) -> list[list[str]]:
    return [[source_file] + row for row in rows]


def save_workbook_to(out: openpyxl.Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        out.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {path.name}. Close it in Excel and retry."
        ) from exc
    return path


def write_hierarchy_sheet(out: openpyxl.Workbook, hierarchy_rows: list[dict]) -> None:
    rows = [
        [
            meta.get("sr") or str(i),
            meta["plant"],
            meta["section"],
            meta["location"],
            meta["main_equipment"],
            meta["sub_equipment"],
            meta["department"],
            meta["raw_tag"],
            meta["hist_location"],
        ]
        for i, meta in enumerate(hierarchy_rows, start=1)
    ]
    write_data_sheet(
        out,
        "Hierarchy",
        HIERARCHY_COLUMNS,
        rows,
        {
            "A": 8, "B": 12, "C": 14, "D": 14, "E": 28,
            "F": 48, "G": 14, "H": 42, "I": 36,
        },
    )
    sheets = out._sheets
    hier = out["Hierarchy"]
    sheets.remove(hier)
    sheets.insert(0, hier)


def write_audit(
    hierarchy_rows: list[dict],
    matched_by_key: dict[str, str],
    extra_cards: list[list[str]],
) -> Path:
    rows = []
    for meta in hierarchy_rows:
        sheet = matched_by_key.get(meta["_key"], "")
        found = bool(sheet)
        rows.append(
            [
                meta["raw_tag"] or "(no tag)",
                meta["department"],
                meta["section"],
                meta["location"],
                meta["main_equipment"],
                meta["sub_equipment"],
                meta["hist_location"],
                "Yes" if found else "No",
                sheet,
            ]
        )
    yes_n = sum(1 for r in rows if r[7] == "Yes")

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_data_sheet(
        out,
        "Summary",
        ["Issue type", "Count"],
        [
            ["Hierarchy source", HIERARCHY_FILE.name],
            ["Hierarchy filter", "Yellow-highlighted (FFFF00) rows only"],
            ["Data source", SOURCE_FILE.name],
            ["Hierarchy equipment rows (yellow)", len(hierarchy_rows)],
            ["Hierarchy rows WITH data", yes_n],
            ["Hierarchy rows WITHOUT data", len(rows) - yes_n],
            ["Cards not matched to hierarchy", len(extra_cards)],
            [
                "Duplicate tags in hierarchy",
                sum(1 for v in group_by_tag(hierarchy_rows).values() if len(v) > 1),
            ],
        ],
        {"A": 42, "B": 60},
    )
    ws = write_data_sheet(
        out,
        "All hierarchy rows",
        [
            "History card Tag Nos.",
            "Department",
            "Section",
            "Location",
            "Main Equipment",
            "Sub Equipment",
            "History card Location",
            "Found in data",
            "Matched sheet / card",
        ],
        rows,
        {
            "A": 42, "B": 14, "C": 18, "D": 16,
            "E": 32, "F": 42, "G": 28, "H": 14, "I": 46,
        },
    )
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 8)
        if cell.value == "Yes":
            cell.fill = YES_FILL
            cell.font = YES_FONT
        else:
            cell.fill = NO_FILL
            cell.font = NO_FONT
    write_data_sheet(
        out,
        "Cards not in hierarchy",
        ["sheet name", "EQUIPMENT TAG NO", "NAME OF EQUIPMENT", "Issue"],
        extra_cards,
        {"A": 36, "B": 42, "C": 40, "D": 40},
    )
    return save_workbook_to(out, AUDIT_OUTPUT)


def main() -> None:
    print(f"Hierarchy: {HIERARCHY_FILE}")
    print(f"Data:      {SOURCE_FILE}")
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(SOURCE_FILE)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_FILE, SOURCE_COPY)
    print(f"Copied source -> {SOURCE_COPY.name}")

    hierarchy_rows = load_hierarchy_rows()
    by_tag = group_by_tag(hierarchy_rows)
    by_loose = group_by_loose_tag(hierarchy_rows)
    print(
        f"Hierarchy rows (yellow): {len(hierarchy_rows)} | unique tags: {len(by_tag)} | "
        f"dup-tag groups: {sum(1 for v in by_tag.values() if len(v) > 1)}"
    )

    src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    source_name = SOURCE_FILE.name

    all_mappings: list[tuple] = []
    all_life: list[list[str]] = []
    all_spec: list[list[str]] = []
    all_schedule: list[list[str]] = []
    all_history: list[list[str]] = []
    used_keys: set[str] = set()
    matched_by_key: dict[str, str] = {}
    extra_cards: list[list[str]] = []
    total_cards = 0

    for sheet_name in src.sheetnames:
        if should_skip_sheet(sheet_name):
            continue
        ws = src[sheet_name]
        blocks = find_equipment_subcards(ws)
        if not blocks:
            continue

        for card_index, block in enumerate(blocks, start=1):
            total_cards += 1
            fields = block_extract_life_fields(ws, block)
            equip_tag = clean_name(
                fields.get("EQUIPMENT TAG NAME/APPLICATION")
                or fields.get("EQUIPMENT TAG NO")
                or ""
            )
            equip_name = clean_name(fields.get("NAME OF EQUIPMENT") or "")
            location = clean_name(fields.get("LOCATION") or "")
            card_label = f"{sheet_name} :: {equip_name or equip_tag or card_index}"

            meta, how = resolve_hierarchy_meta(
                hierarchy_rows,
                by_tag,
                by_loose,
                equip_tag,
                equip_name,
                sheet_name,
                location,
                used_keys,
            )
            if meta is None:
                extra_cards.append(
                    [sheet_name, equip_tag or "(blank)", equip_name, "Not in yellow hierarchy set"]
                )
                continue

            used_keys.add(meta["_key"])
            matched_by_key[meta["_key"]] = f"{card_label} [{how}]"

            sheet_id = uuid.uuid4().hex[:12]
            out_tag = meta["raw_tag"] or equip_tag
            out_name = equip_name or meta["sub_equipment"]
            out_sheet_name = f"{sheet_name} :: {out_name}"

            all_mappings.append(
                (source_name, out_sheet_name, sheet_id, out_tag, meta["sub_equipment"])
            )
            all_life.append(
                [
                    source_name,
                    sheet_id,
                    out_sheet_name,
                    out_tag,
                    out_name,
                    location or meta.get("hist_location", "") or meta.get("location", ""),
                    clean_name(fields.get("DATE OF COMMISSIONING") or ""),
                    meta.get("main_equipment", ""),
                    meta.get("sub_equipment", ""),
                    meta.get("department", "INSTRUMENT"),
                ]
            )
            all_spec.extend(
                prepend_source_file(
                    insert_tag_column(
                        block_extract_specs(ws, block, sheet_id, out_sheet_name),
                        out_tag,
                    ),
                    source_name,
                )
            )
            all_schedule.extend(
                prepend_source_file(
                    insert_tag_column(
                        block_extract_schedule(ws, block, sheet_id, out_sheet_name),
                        out_tag,
                    ),
                    source_name,
                )
            )
            all_history.extend(
                prepend_source_file(
                    insert_tag_column(
                        block_extract_history(ws, block, sheet_id, out_sheet_name),
                        out_tag,
                    ),
                    source_name,
                )
            )
            print(f"  OK ({how}): {card_label} -> {out_tag} | {meta['sub_equipment']}")

    src.close()

    missing = [m for m in hierarchy_rows if m["_key"] not in matched_by_key]
    print(f"\nTotal cards scanned in data file: {total_cards}")
    print(f"Extracted cards (matched to yellow hierarchy): {len(all_mappings)}")
    print(f"Yellow hierarchy rows missing data: {len(missing)}")
    for m in missing:
        print(f"  - {m['raw_tag']}  ({m['sub_equipment']})")
    print(f"Cards not matched to hierarchy: {len(extra_cards)}")

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_hierarchy_sheet(out, hierarchy_rows)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["source file", "sheet name", "id", "EQUIPMENT TAG NO", "Sub Equipment"])
    for row in all_mappings:
        map_ws.append(list(row))
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    for col, width in {"A": 48, "B": 44, "C": 16, "D": 42, "E": 42}.items():
        map_ws.column_dimensions[col].width = width

    write_data_sheet(
        out,
        "EQUIPMENT LIFE HISTORY CARD",
        LIFE_COLUMNS,
        all_life,
        {
            "A": 48, "B": 16, "C": 44, "D": 42, "E": 40,
            "F": 28, "G": 20, "H": 28, "I": 36, "J": 14,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT SPECIFICATION",
        SPEC_COLUMNS,
        all_spec,
        {"A": 48, "B": 16, "C": 44, "D": 42, "E": 14, "F": 18, "G": 28, "H": 36},
    )
    write_data_sheet(
        out,
        "MAINTENANCE SCHEDULE",
        SCHEDULE_COLUMNS,
        all_schedule,
        {
            "A": 48, "B": 16, "C": 44, "D": 42, "E": 8, "F": 22, "G": 42,
            "H": 8, "I": 8, "J": 8, "K": 10, "L": 12, "M": 8,
            "N": 10, "O": 10, "P": 10, "Q": 16,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT MAINTENANCE HISTORY",
        HISTORY_COLUMNS,
        all_history,
        {
            "A": 48, "B": 16, "C": 44, "D": 42, "E": 18, "F": 14, "G": 16,
            "H": 16, "I": 36, "J": 36, "K": 16, "L": 22, "M": 28, "N": 24,
        },
    )

    saved = save_workbook_to(out, OUTPUT)
    audit_path = write_audit(hierarchy_rows, matched_by_key, extra_cards)
    print(f"\nEXTRACTED: {saved}")
    print(f"AUDIT:     {audit_path}")
    print(
        f"Hierarchy ({len(hierarchy_rows)}), Sheet Map ({len(all_mappings)}), "
        f"LIFE ({len(all_life)}), SPEC ({len(all_spec)}), "
        f"SCHEDULE ({len(all_schedule)}), HISTORY ({len(all_history)})"
    )


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Extract Mill House mechanical life history cards.

Hierarchy:
  MILL HOUSE EQUIPMENT HISTORY CARD TAG LIST - 11082026.xlsx
Data:
  MILL HOUSE HISTORY CARD updated aug 26.xlsx

One output card per history *sheet* (not per hierarchy row).

Matching:
  - Tag from the sheet vs hierarchy tag
  - If several hierarchy rows share that tag, also match Sub Equipment
    to the sheet's NAME OF EQUIPMENT (do not copy one card onto every child)

Output (under DigiLog/backend/backlog-data/mill data/):
  mill-house-mechanical-equipment-history-11082026.xlsx
  mill-house-mechanical-extract-audit-11082026.xlsx
"""

from __future__ import annotations

import re
import uuid
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from equipment_history_extract_lib import cell_text, norm
from extract_mill_house_electrical import (
    HISTORY_COLUMNS,
    LIFE_COLUMNS,
    SCHEDULE_COLUMNS,
    SPEC_COLUMNS,
    extract_maintenance_history,
    extract_maintenance_schedule,
    extract_specification_rows,
    insert_tag_column,
    prepend_source_file,
    style_header_row,
    write_data_sheet,
)

ROOT = Path(__file__).resolve().parent.parent.parent
BACKLOG = ROOT / "DigiLog" / "backend" / "backlog-data"
OUTPUT_DIR = BACKLOG / "mill data"

HIERARCHY_FILE = ROOT / "MILL HOUSE EQUIPMENT HISTORY CARD TAG LIST - 11082026.xlsx"
DATA_FILE = ROOT / "MILL HOUSE HISTORY CARD updated aug 26.xlsx"
OUTPUT = OUTPUT_DIR / "mill-house-mechanical-equipment-history-11082026.xlsx"
AUDIT_OUTPUT = OUTPUT_DIR / "mill-house-mechanical-extract-audit-11082026.xlsx"

SKIP_SHEETS = {"index", "summary index", "summary link", "sheet1"}
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill("solid", fgColor="FFC7CE")
YES_FONT = Font(bold=True, color="006100")
NO_FONT = Font(bold=True, color="9C0006")


def norm_tag(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def clean_name(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def should_skip_sheet(name: str) -> bool:
    return norm(name.strip()).lower() in SKIP_SHEETS


def row_value_after_label(ws, row_idx: int, label_col: int) -> str:
    for c in range(label_col + 1, ws.max_column + 1):
        val = cell_text(ws.cell(row_idx, c).value)
        if val:
            return val
    return ""


def load_hierarchy_rows(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"Hierarchy file not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        tag = clean_name(ws.cell(r, 8).value)
        if not tag or "/" not in tag:
            continue
        rows.append(
            {
                "raw_tag": tag,
                "plant": clean_name(ws.cell(r, 2).value),
                "section": clean_name(ws.cell(r, 3).value),
                "location": clean_name(ws.cell(r, 4).value),
                "main_equipment": clean_name(ws.cell(r, 5).value),
                "sub_equipment": clean_name(ws.cell(r, 6).value),
                "department": clean_name(ws.cell(r, 7).value),
                "hist_location": clean_name(ws.cell(r, 9).value),
                "hierarchy_row": r,
            }
        )
    wb.close()
    return rows


def first_tag_map(rows: list[dict]) -> dict[str, dict]:
    tag_map: dict[str, dict] = {}
    for meta in rows:
        tag_map.setdefault(norm_tag(meta["raw_tag"]), meta)
    return tag_map


def rows_by_tag(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for meta in rows:
        grouped[norm_tag(meta["raw_tag"])].append(meta)
    return grouped


def equipment_name_key(value) -> str:
    """Compare equipment names ignoring punctuation, optional 'NO', and Cardian/Carding."""
    s = str(value or "").lower()
    s = s.replace(".", " ").replace("-", " ").replace("_", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\bno\s+", "", s)
    s = s.replace("cardian", "carding")
    return re.sub(r"\s+", " ", s).strip()


def name_key_variants(value) -> set[str]:
    key = equipment_name_key(value)
    if not key:
        return set()
    variants = {key}
    if key.startswith("cane "):
        variants.add(key[5:])
    return variants


def names_match(a, b) -> bool:
    return bool(name_key_variants(a) & name_key_variants(b))


def pick_hierarchy_row_for_sheet(fields: dict, sheet_name: str, metas: list[dict]) -> dict | None:
    """Match one hierarchy leaf: same tag group, same sub-equipment name."""
    card = (fields or {}).get("NAME OF EQUIPMENT") or ""
    title = sheet_name or ""
    for meta in metas:
        sub = meta.get("sub_equipment") or ""
        if names_match(card, sub) or names_match(title, sub):
            return meta
    for meta in metas:
        sub = meta.get("sub_equipment") or ""
        main = meta.get("main_equipment") or ""
        if names_match(sub, main) and (names_match(card, main) or names_match(title, main)):
            return meta
    return None


def extract_life_fields(ws) -> dict[str, str]:
    fields = {
        "NAME OF EQUIPMENT": ws.title.strip(),
        "LOCATION": "",
        "EQUIPMENT TAG NO": "",
        "DATE OF COMMISSIONING": "",
        "DRIVE": "",
    }
    life_row = None
    life_col = 1
    spec_row = ws.max_row + 1

    for r in range(1, min(ws.max_row, 40) + 1):
        for c in (1, 2, 3, 4, 5):
            text = norm(ws.cell(r, c).value)
            if not text:
                continue
            if life_row is None and "EQUIPMENT LIFE HISTORY" in text:
                life_row = r
                life_col = c
            if "EQUIPMENT SPECIFICATION" in text:
                spec_row = r

    scan_start = (life_row + 1) if life_row is not None else 1
    scan_end = spec_row if spec_row <= ws.max_row else min(ws.max_row, 25)
    equip_no = ""

    for r in range(scan_start, scan_end):
        label = ""
        label_col = 1
        for c in range(1, min(8, ws.max_column + 1)):
            raw = norm(ws.cell(r, c).value)
            if raw:
                label = raw
                label_col = c
                break
        if not label:
            continue
        value = row_value_after_label(ws, r, label_col)
        if "NAME OF EQUIPMENT" in label:
            fields["NAME OF EQUIPMENT"] = value or fields["NAME OF EQUIPMENT"]
        elif label.startswith("LOCATION"):
            fields["LOCATION"] = value
        elif "EQUIPMENT TAG" in label or label.startswith("TAG NO") or label == "TAG NAME":
            fields["EQUIPMENT TAG NO"] = value
        elif "EQUIPMENT NO" in label and "TAG" not in label:
            equip_no = value
        elif "DATE OF COMMISSIONING" in label or "COMMISSIONING" in label:
            fields["DATE OF COMMISSIONING"] = value
        elif label.startswith("DRIVE"):
            fields["DRIVE"] = value

    if not fields["EQUIPMENT TAG NO"] and equip_no:
        fields["EQUIPMENT TAG NO"] = equip_no

    if not fields["EQUIPMENT TAG NO"]:
        for r in range(1, min(ws.max_row, 40) + 1):
            for c in range(1, min(ws.max_column, 12) + 1):
                val = cell_text(ws.cell(r, c).value)
                if "ZIL/" in val.upper():
                    fields["EQUIPMENT TAG NO"] = val
                    break
            if fields["EQUIPMENT TAG NO"]:
                break

    if not fields["NAME OF EQUIPMENT"]:
        fields["NAME OF EQUIPMENT"] = ws.title.strip()
    return fields


def save_workbook_to(out: openpyxl.Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fallback = path.with_name(f"{path.stem}-updated{path.suffix}")
    for candidate in (path, fallback):
        try:
            out.save(candidate)
            return candidate
        except PermissionError:
            continue
    raise PermissionError(
        f"Could not write {path.name} or {fallback.name}. Close them in Excel and retry."
    )


def write_audit(
    hierarchy_rows: list[dict],
    matched_by_row: dict[int, str],
    matched_sheets: dict[str, list[str]],
) -> Path:
    rows = []
    for meta in hierarchy_rows:
        nt = norm_tag(meta["raw_tag"])
        sheet_for_row = matched_by_row.get(meta["hierarchy_row"], "")
        tag_sheets = matched_sheets.get(nt, [])
        found = bool(sheet_for_row)
        rows.append(
            [
                DATA_FILE.name,
                HIERARCHY_FILE.name,
                meta["raw_tag"],
                meta["department"],
                meta["section"],
                meta["location"],
                meta["main_equipment"],
                meta["sub_equipment"],
                meta["hist_location"],
                "Yes" if found else "No",
                sheet_for_row or ", ".join(tag_sheets),
            ]
        )

    yes_n = sum(1 for r in rows if r[9] == "Yes")
    unique_yes = len(matched_sheets)
    unique_all = len({norm_tag(m["raw_tag"]) for m in hierarchy_rows})

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_data_sheet(
        out,
        "Summary",
        ["Issue type", "Count"],
        [
            ["Hierarchy equipment rows", len(rows)],
            ["Unique tags in hierarchy", unique_all],
            ["History sheets matched by tag", unique_yes],
            ["Equipment rows WITH history (tag + sub-equipment)", yes_n],
            ["Equipment rows WITHOUT matching sub-equipment sheet", len(rows) - yes_n],
        ],
        {"A": 58, "B": 12},
    )
    ws = write_data_sheet(
        out,
        "All hierarchy rows",
        [
            "source file",
            "hierarchy file",
            "History card Tag No.",
            "Department",
            "Section",
            "Location",
            "Main Equipment",
            "Sub Equipment",
            "History card Location",
            "Found in data",
            "Matched sheet name(s)",
        ],
        rows,
        {
            "A": 42, "B": 56, "C": 28, "D": 20, "E": 14, "F": 28,
            "G": 36, "H": 42, "I": 22, "J": 14, "K": 28,
        },
    )
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 10)
        if cell.value == "Yes":
            cell.fill = YES_FILL
            cell.font = YES_FONT
        else:
            cell.fill = NO_FILL
            cell.font = NO_FONT
    return save_workbook_to(out, AUDIT_OUTPUT)


def main() -> None:
    print(f"Hierarchy: {HIERARCHY_FILE}")
    print(f"Data:      {DATA_FILE}")

    hierarchy_rows = load_hierarchy_rows(HIERARCHY_FILE)
    hierarchy_tags = first_tag_map(hierarchy_rows)
    hierarchy_by_tag = rows_by_tag(hierarchy_rows)
    print(f"Hierarchy rows: {len(hierarchy_rows)} | unique tags: {len(hierarchy_tags)}")

    if not DATA_FILE.exists():
        raise FileNotFoundError(f"Data file not found: {DATA_FILE}")

    src = openpyxl.load_workbook(DATA_FILE, data_only=True)
    source_name = DATA_FILE.name

    all_mappings: list[tuple[str, str, str, str, str]] = []
    all_life: list[list[str]] = []
    all_spec: list[list[str]] = []
    all_schedule: list[list[str]] = []
    all_history: list[list[str]] = []
    matched_sheets: dict[str, list[str]] = {}
    matched_by_row: dict[int, str] = {}
    unmatched_sheets: list[str] = []
    unmatched_names: list[str] = []

    for sheet_name in src.sheetnames:
        if should_skip_sheet(sheet_name):
            continue

        ws = src[sheet_name]
        fields = extract_life_fields(ws)
        equip_tag = fields["EQUIPMENT TAG NO"]
        nt = norm_tag(equip_tag)
        metas = hierarchy_by_tag.get(nt) or []
        if not metas:
            unmatched_sheets.append(f"{sheet_name} (tag={equip_tag or '(blank)'})")
            continue

        matched_sheets.setdefault(nt, [])
        if sheet_name in matched_sheets[nt]:
            continue
        matched_sheets[nt].append(sheet_name)

        meta = pick_hierarchy_row_for_sheet(fields, sheet_name, metas)
        out_tag = (meta or metas[0])["raw_tag"]
        sub_equipment = (meta["sub_equipment"] if meta else fields["NAME OF EQUIPMENT"])
        if meta:
            matched_by_row[meta["hierarchy_row"]] = sheet_name
        else:
            unmatched_names.append(
                f"{sheet_name} tag={out_tag} name={fields['NAME OF EQUIPMENT']}"
            )

        sheet_id = uuid.uuid4().hex[:12]
        all_mappings.append((source_name, sheet_name, sheet_id, out_tag, sub_equipment))
        location = fields["LOCATION"]
        if not location and meta:
            location = meta.get("hist_location") or meta.get("location") or ""
        if not location:
            location = metas[0].get("hist_location") or metas[0].get("location") or ""
        all_life.append(
            [
                source_name,
                sheet_id,
                sheet_name,
                out_tag,
                fields["NAME OF EQUIPMENT"] or sub_equipment,
                location,
                fields["DATE OF COMMISSIONING"],
                (meta or metas[0])["main_equipment"],
                sub_equipment,
                (meta or metas[0])["department"],
                (meta or metas[0]).get("hist_location") or "",
            ]
        )
        all_spec.extend(prepend_source_file(
            insert_tag_column(
                extract_specification_rows(ws, sheet_id, sheet_name),
                out_tag,
            ),
            source_name,
        ))
        all_schedule.extend(prepend_source_file(
            insert_tag_column(
                extract_maintenance_schedule(ws, sheet_id, sheet_name),
                out_tag,
            ),
            source_name,
        ))
        all_history.extend(prepend_source_file(
            insert_tag_column(
                extract_maintenance_history(ws, sheet_id, sheet_name),
                out_tag,
            ),
            source_name,
        ))

    src.close()

    missing = sorted(set(hierarchy_tags) - set(matched_sheets))
    print(f"History sheets matched by tag: {len(matched_sheets)} | tags with no sheet: {len(missing)}")
    print(f"Extracted cards (one per sheet): {len(all_mappings)}")
    print(f"Hierarchy rows matched by tag + sub-equipment: {len(matched_by_row)}")
    if unmatched_names:
        print("Sheets whose name did not match a sub-equipment (still extracted):")
        for line in unmatched_names:
            print(f"  - {line}")
    if unmatched_sheets:
        print("Data sheets not in hierarchy:")
        for line in unmatched_sheets:
            print(f"  - {line}")

    out = openpyxl.Workbook()
    out.remove(out.active)

    mech_life_columns = [*LIFE_COLUMNS, "History card Location"]

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(["source file", "sheet name", "id", "EQUIPMENT TAG NO", "Sub Equipment"])
    for source_file_name, sheet_name, sheet_id, equip_tag, sub_equipment in all_mappings:
        map_ws.append([source_file_name, sheet_name, sheet_id, equip_tag, sub_equipment])
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    map_ws.column_dimensions["A"].width = 48
    map_ws.column_dimensions["B"].width = 28
    map_ws.column_dimensions["C"].width = 16
    map_ws.column_dimensions["D"].width = 28
    map_ws.column_dimensions["E"].width = 42

    write_data_sheet(
        out,
        "EQUIPMENT LIFE HISTORY CARD",
        mech_life_columns,
        all_life,
        {
            "A": 48, "B": 16, "C": 28, "D": 28, "E": 36, "F": 22,
            "G": 20, "H": 28, "I": 36, "J": 22, "K": 28,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT SPECIFICATION",
        SPEC_COLUMNS,
        all_spec,
        {"A": 48, "B": 16, "C": 28, "D": 28, "E": 14, "F": 18, "G": 28, "H": 36},
    )
    write_data_sheet(
        out,
        "MAINTENANCE SCHEDULE",
        SCHEDULE_COLUMNS,
        all_schedule,
        {
            "A": 48, "B": 16, "C": 28, "D": 28, "E": 8, "F": 22, "G": 42,
            "H": 8, "I": 8, "J": 8, "K": 10, "L": 12, "M": 8,
            "N": 10, "O": 10, "P": 10, "Q": 16,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT MAINTENANCE HISTORY",
        HISTORY_COLUMNS,
        all_history,
        {
            "A": 48, "B": 16, "C": 28, "D": 28, "E": 18, "F": 14, "G": 16,
            "H": 16, "I": 36, "J": 36, "K": 16, "L": 22, "M": 28, "N": 24,
        },
    )

    saved = save_workbook_to(out, OUTPUT)
    audit_path = write_audit(hierarchy_rows, matched_by_row, matched_sheets)

    print(f"\nEXTRACTED: {saved}")
    print(f"AUDIT:     {audit_path}")
    print(
        f"Sheet Map ({len(all_mappings)}), LIFE ({len(all_life)}), "
        f"SPEC ({len(all_spec)}), SCHEDULE ({len(all_schedule)}), "
        f"HISTORY ({len(all_history)})"
    )


if __name__ == "__main__":
    main()

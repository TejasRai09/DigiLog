#!/usr/bin/env python3
"""
Extract DS House Electrical Motors life-history cards.

Source (hierarchy in INDEX, cards in remaining sheets — often multiple cards per sheet):
  Extraction_files-sugar/Electrical/2_DS Electrical Equipment Life History -MOTORS.xlsx

Matching (hierarchy row ↔ card):
  1. EQUIPMENT NO / TAG vs INDEX "History card Tag Nos."
  2. If several hierarchy rows share that tag, also match Sub Equipment
     to the card's NAME OF EQUIPMENT (one card → one leaf; no fan-out).

Output (DigiLog/backend/backlog-data/mill data/migration files-24-08-26/):
  ds-motors-equipment-history-240826.xlsx
    - Hierarchy (first sheet)
    - Sheet Map
    - EQUIPMENT LIFE HISTORY CARD / SPECIFICATION / SCHEDULE / HISTORY
  ds-motors-extract-audit-240826.xlsx
"""

from __future__ import annotations

import re
import shutil
import uuid
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from extract_mill_house_electrical import (
    HISTORY_COLUMNS,
    LIFE_COLUMNS,
    SCHEDULE_COLUMNS,
    SPEC_COLUMNS,
    style_header_row,
    write_data_sheet,
)
from extract_mill_house_mechanical import equipment_name_key, names_match
from extract_shn_life_history_table import (
    extract_life_fields as block_extract_life_fields,
    extract_maintenance_history as block_extract_history,
    extract_maintenance_schedule as block_extract_schedule,
    extract_specification_rows as block_extract_specs,
    find_equipment_subcards,
    insert_tag_column,
)

ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE_FILE = (
    ROOT
    / "Extraction_files-sugar"
    / "Electrical"
    / "2_DS Electrical Equipment Life History -MOTORS.xlsx"
)
OUTPUT_DIR = (
    ROOT
    / "DigiLog"
    / "backend"
    / "backlog-data"
    / "mill data"
    / "migration files-24-08-26"
)
OUTPUT = OUTPUT_DIR / "ds-motors-equipment-history-240826.xlsx"
AUDIT_OUTPUT = OUTPUT_DIR / "ds-motors-extract-audit-240826.xlsx"
SOURCE_COPY = OUTPUT_DIR / "2_DS Electrical Equipment Life History -MOTORS.xlsx"

SKIP_SHEETS = {"index", "summary index", "summary link", "sheet1", "hierarchy"}
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill("solid", fgColor="FFC7CE")
YES_FONT = Font(bold=True, color="006100")
NO_FONT = Font(bold=True, color="9C0006")

HIERARCHY_COLUMNS = [
    "Sr.No.",
    "Plant",
    "Section",
    "Location",
    "Main Equipment",
    " Sub Equipment",
    "Department",
    "History card Tag Nos.",
    "History card Location",
]


def norm_tag(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def clean_name(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def should_skip_sheet(name: str) -> bool:
    return clean_name(name).lower() in SKIP_SHEETS


def load_hierarchy_rows(wb) -> list[dict]:
    ws = wb[wb.sheetnames[0]]
    header_row = 1
    for r in range(1, min(12, ws.max_row) + 1):
        cells = [clean_name(ws.cell(r, c).value).lower() for c in range(1, 10)]
        if any("tag" in c for c in cells) and any(
            "section" in c or "equipment" in c for c in cells
        ):
            header_row = r
            break

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        tag = clean_name(ws.cell(r, 8).value)
        sub = clean_name(ws.cell(r, 6).value)
        if not tag and not sub:
            continue
        if tag and "/" not in tag:
            continue
        rows.append(
            {
                "sr": clean_name(ws.cell(r, 1).value),
                "raw_tag": tag,
                "plant": clean_name(ws.cell(r, 2).value),
                "section": clean_name(ws.cell(r, 3).value),
                "location": clean_name(ws.cell(r, 4).value),
                "main_equipment": clean_name(ws.cell(r, 5).value),
                "sub_equipment": sub,
                "department": clean_name(ws.cell(r, 7).value),
                "hist_location": clean_name(ws.cell(r, 9).value),
                "hierarchy_row": r,
                "_key": f"{r}:{norm_tag(tag)}:{equipment_name_key(sub)}",
            }
        )
    return rows


def group_by_tag(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for meta in rows:
        if meta["raw_tag"]:
            grouped[norm_tag(meta["raw_tag"])].append(meta)
    return grouped


def pick_hierarchy_meta(
    metas: list[dict],
    equip_name: str,
    used_keys: set[str],
) -> dict | None:
    """Tag group already filtered. Prefer unused row whose Sub Equipment matches name."""
    available = [m for m in metas if m["_key"] not in used_keys]
    if not available:
        return None
    if len(available) == 1 and len(metas) == 1:
        return available[0]

    name_hits = [
        m for m in available
        if names_match(equip_name, m["sub_equipment"])
    ]
    if len(name_hits) == 1:
        return name_hits[0]
    if len(name_hits) > 1:
        return name_hits[0]

    # Unique tag in hierarchy but name text differs slightly — still allow.
    if len(metas) == 1:
        return available[0]
    return None


def prepend_source_file(rows: list[list[str]], source_file: str) -> list[list[str]]:
    return [[source_file] + row for row in rows]


def save_workbook_to(out: openpyxl.Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        out.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {path.name}. Close it in Excel and retry."
        ) from exc
    return path


def write_hierarchy_sheet(out: openpyxl.Workbook, hierarchy_rows: list[dict]) -> None:
    rows = [
        [
            meta.get("sr") or str(i),
            meta["plant"],
            meta["section"],
            meta["location"],
            meta["main_equipment"],
            meta["sub_equipment"],
            meta["department"],
            meta["raw_tag"],
            meta["hist_location"],
        ]
        for i, meta in enumerate(hierarchy_rows, start=1)
    ]
    write_data_sheet(
        out,
        "Hierarchy",
        HIERARCHY_COLUMNS,
        rows,
        {
            "A": 8, "B": 12, "C": 14, "D": 14, "E": 28,
            "F": 48, "G": 14, "H": 42, "I": 36,
        },
    )
    sheets = out._sheets
    hier = out["Hierarchy"]
    sheets.remove(hier)
    sheets.insert(0, hier)


def write_audit(
    hierarchy_rows: list[dict],
    matched_by_key: dict[str, str],
    extra_cards: list[list[str]],
    ambiguous: list[list[str]],
) -> Path:
    rows = []
    for meta in hierarchy_rows:
        sheet = matched_by_key.get(meta["_key"], "")
        found = bool(sheet)
        rows.append(
            [
                SOURCE_FILE.name,
                meta["raw_tag"] or "(no tag)",
                meta["department"],
                meta["section"],
                meta["location"],
                meta["main_equipment"],
                meta["sub_equipment"],
                meta["hist_location"],
                "Yes" if found else "No",
                sheet,
            ]
        )
    yes_n = sum(1 for r in rows if r[8] == "Yes")

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_data_sheet(
        out,
        "Summary",
        ["Issue type", "Count"],
        [
            ["Hierarchy equipment rows", len(hierarchy_rows)],
            ["Hierarchy rows WITH data", yes_n],
            ["Hierarchy rows WITHOUT data", len(rows) - yes_n],
            ["Cards not matched to hierarchy", len(extra_cards)],
            ["Ambiguous tag+name (skipped)", len(ambiguous)],
            [
                "Duplicate tags in hierarchy",
                sum(
                    1
                    for metas in group_by_tag(hierarchy_rows).values()
                    if len(metas) > 1
                ),
            ],
        ],
        {"A": 48, "B": 12},
    )
    ws = write_data_sheet(
        out,
        "All hierarchy rows",
        [
            "source file",
            "History card Tag Nos.",
            "Department",
            "Section",
            "Location",
            "Main Equipment",
            "Sub Equipment",
            "History card Location",
            "Found in data",
            "Matched sheet / card",
        ],
        rows,
        {
            "A": 48, "B": 42, "C": 14, "D": 18, "E": 16,
            "F": 32, "G": 42, "H": 28, "I": 14, "J": 40,
        },
    )
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 9)
        if cell.value == "Yes":
            cell.fill = YES_FILL
            cell.font = YES_FONT
        else:
            cell.fill = NO_FILL
            cell.font = NO_FONT
    write_data_sheet(
        out,
        "Cards not in hierarchy",
        [
            "source file",
            "sheet name",
            "EQUIPMENT TAG NO",
            "NAME OF EQUIPMENT",
            "Issue",
        ],
        extra_cards,
        {"A": 48, "B": 36, "C": 42, "D": 40, "E": 40},
    )
    write_data_sheet(
        out,
        "Ambiguous matches",
        [
            "source file",
            "sheet name",
            "EQUIPMENT TAG NO",
            "NAME OF EQUIPMENT",
            "Issue",
        ],
        ambiguous,
        {"A": 48, "B": 36, "C": 42, "D": 40, "E": 50},
    )
    return save_workbook_to(out, AUDIT_OUTPUT)


def main() -> None:
    print(f"Source: {SOURCE_FILE}")
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(SOURCE_FILE)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_FILE, SOURCE_COPY)
    print(f"Copied source -> {SOURCE_COPY.name}")

    src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    source_name = SOURCE_FILE.name
    hierarchy_rows = load_hierarchy_rows(src)
    by_tag = group_by_tag(hierarchy_rows)
    print(
        f"Hierarchy rows: {len(hierarchy_rows)} | unique tags: {len(by_tag)} | "
        f"dup-tag groups: {sum(1 for v in by_tag.values() if len(v) > 1)}"
    )

    all_mappings: list[tuple] = []
    all_life: list[list[str]] = []
    all_spec: list[list[str]] = []
    all_schedule: list[list[str]] = []
    all_history: list[list[str]] = []
    used_keys: set[str] = set()
    matched_by_key: dict[str, str] = {}
    extra_cards: list[list[str]] = []
    ambiguous: list[list[str]] = []

    for sheet_name in src.sheetnames[1:]:
        if should_skip_sheet(sheet_name):
            continue
        ws = src[sheet_name]
        blocks = find_equipment_subcards(ws)
        if not blocks:
            print(f"  SKIP (no cards): {sheet_name}")
            continue

        for card_index, block in enumerate(blocks, start=1):
            fields = block_extract_life_fields(ws, block)
            equip_tag = clean_name(
                fields.get("EQUIPMENT TAG NAME/APPLICATION")
                or fields.get("EQUIPMENT TAG NO")
                or ""
            )
            equip_name = clean_name(fields.get("NAME OF EQUIPMENT") or "")
            card_label = f"{sheet_name} :: {equip_name or equip_tag or card_index}"

            if not equip_tag or "/" not in equip_tag:
                extra_cards.append(
                    [source_name, sheet_name, equip_tag or "(blank)", equip_name, "No tag on card"]
                )
                print(f"  SKIP (no tag): {card_label}")
                continue

            nt = norm_tag(equip_tag)
            metas = by_tag.get(nt, [])
            if not metas:
                extra_cards.append(
                    [
                        source_name,
                        sheet_name,
                        equip_tag,
                        equip_name,
                        "Tag not found in hierarchy",
                    ]
                )
                print(f"  SKIP (tag not in hierarchy): {card_label}  tag={equip_tag}")
                continue

            meta = pick_hierarchy_meta(metas, equip_name, used_keys)
            if meta is None:
                if len(metas) > 1:
                    ambiguous.append(
                        [
                            source_name,
                            sheet_name,
                            equip_tag,
                            equip_name,
                            "Shared tag — Sub Equipment name did not uniquely match",
                        ]
                    )
                    print(f"  SKIP (ambiguous name): {card_label}  tag={equip_tag}")
                else:
                    extra_cards.append(
                        [
                            source_name,
                            sheet_name,
                            equip_tag,
                            equip_name,
                            "Tag already matched to another card",
                        ]
                    )
                    print(f"  SKIP (tag already used): {card_label}")
                continue

            used_keys.add(meta["_key"])
            matched_by_key[meta["_key"]] = card_label

            sheet_id = uuid.uuid4().hex[:12]
            out_tag = meta["raw_tag"]
            out_name = equip_name or meta["sub_equipment"]
            out_sheet_name = f"{sheet_name} :: {out_name}"

            all_mappings.append(
                (source_name, out_sheet_name, sheet_id, out_tag, meta["sub_equipment"])
            )
            all_life.append(
                [
                    source_name,
                    sheet_id,
                    out_sheet_name,
                    out_tag,
                    out_name,
                    clean_name(fields.get("LOCATION") or "")
                    or meta.get("hist_location", "")
                    or meta.get("location", ""),
                    clean_name(fields.get("DATE OF COMMISSIONING") or ""),
                    meta.get("main_equipment", ""),
                    meta.get("sub_equipment", ""),
                    meta.get("department", "ELECTRICAL"),
                ]
            )
            all_spec.extend(
                prepend_source_file(
                    insert_tag_column(
                        block_extract_specs(ws, block, sheet_id, out_sheet_name),
                        out_tag,
                    ),
                    source_name,
                )
            )
            all_schedule.extend(
                prepend_source_file(
                    insert_tag_column(
                        block_extract_schedule(ws, block, sheet_id, out_sheet_name),
                        out_tag,
                    ),
                    source_name,
                )
            )
            all_history.extend(
                prepend_source_file(
                    insert_tag_column(
                        block_extract_history(ws, block, sheet_id, out_sheet_name),
                        out_tag,
                    ),
                    source_name,
                )
            )
            print(f"  OK: {card_label} -> {out_tag} | {meta['sub_equipment']}")

    src.close()

    missing = [m for m in hierarchy_rows if m["_key"] not in matched_by_key]
    print(f"\nExtracted cards: {len(all_mappings)}")
    print(f"Hierarchy rows missing data: {len(missing)}")
    for m in missing[:15]:
        print(f"  - {m['raw_tag']}  ({m['sub_equipment']})")
    if len(missing) > 15:
        print(f"  … and {len(missing) - 15} more")
    print(f"Cards not matched: {len(extra_cards)} | Ambiguous: {len(ambiguous)}")

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_hierarchy_sheet(out, hierarchy_rows)

    map_ws = out.create_sheet("Sheet Map")
    map_ws.append(
        ["source file", "sheet name", "id", "EQUIPMENT TAG NO", "Sub Equipment"]
    )
    for row in all_mappings:
        map_ws.append(list(row))
    style_header_row(map_ws)
    map_ws.freeze_panes = "A2"
    for col, width in {"A": 48, "B": 44, "C": 16, "D": 42, "E": 42}.items():
        map_ws.column_dimensions[col].width = width

    write_data_sheet(
        out,
        "EQUIPMENT LIFE HISTORY CARD",
        LIFE_COLUMNS,
        all_life,
        {
            "A": 48, "B": 16, "C": 44, "D": 42, "E": 40,
            "F": 28, "G": 20, "H": 28, "I": 36, "J": 14,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT SPECIFICATION",
        SPEC_COLUMNS,
        all_spec,
        {"A": 48, "B": 16, "C": 44, "D": 42, "E": 14, "F": 18, "G": 28, "H": 36},
    )
    write_data_sheet(
        out,
        "MAINTENANCE SCHEDULE",
        SCHEDULE_COLUMNS,
        all_schedule,
        {
            "A": 48, "B": 16, "C": 44, "D": 42, "E": 8, "F": 22, "G": 42,
            "H": 8, "I": 8, "J": 8, "K": 10, "L": 12, "M": 8,
            "N": 10, "O": 10, "P": 10, "Q": 16,
        },
    )
    write_data_sheet(
        out,
        "EQUIPMENT MAINTENANCE HISTORY",
        HISTORY_COLUMNS,
        all_history,
        {
            "A": 48, "B": 16, "C": 44, "D": 42, "E": 18, "F": 14, "G": 16,
            "H": 16, "I": 36, "J": 36, "K": 16, "L": 22, "M": 28, "N": 24,
        },
    )

    saved = save_workbook_to(out, OUTPUT)
    audit_path = write_audit(hierarchy_rows, matched_by_key, extra_cards, ambiguous)
    print(f"\nEXTRACTED: {saved}")
    print(f"AUDIT:     {audit_path}")
    print(
        f"Hierarchy ({len(hierarchy_rows)}), Sheet Map ({len(all_mappings)}), "
        f"LIFE ({len(all_life)}), SPEC ({len(all_spec)}), "
        f"SCHEDULE ({len(all_schedule)}), HISTORY ({len(all_history)})"
    )


if __name__ == "__main__":
    main()
